"""
Central Error Logging Utility for Validation Tools

This module provides functions for all processors to log errors
in a consistent format to a central Excel file.

Usage:
    from utils.data_logger import log_errors

    log_errors(
        hall="JPB15",
        rack_type="T0-to-Host",
        building="B11",
        error_category="Mismatch",
        count=14,
        source_file="JPB15_T0_Host_B11.xlsx"
    )
"""

import pandas as pd
from datetime import datetime
import os
from pathlib import Path

# Central log file location (relative to repo root)
LOG_FILE = Path(__file__).parent.parent / "data" / "validation_error_log.xlsx"

def ensure_log_exists():
    """Create the log file with proper columns if it doesn't exist."""
    LOG_FILE.parent.mkdir(parents=True, exist_ok=True)
    
    if not LOG_FILE.exists():
        df = pd.DataFrame(columns=[
            "timestamp",
            "hall",
            "rack_type",
            "building",
            "error_category",
            "count",
            "source_file",
            "processed_by"
        ])
        df.to_excel(LOG_FILE, index=False)
        print(f"Created new error log at: {LOG_FILE}")

def log_errors(
    hall: str,
    rack_type: str,
    building: str,
    error_category: str,
    count: int,
    source_file: str = "",
    processed_by: str = "system"
):
    """
    Log error counts from a validation processor.
    
    Args:
        hall: e.g. "JPB15", "JPB19", "SYD20"
        rack_type: e.g. "T0-to-Host", "T1-to-T0", "CFAB", "QFAB", "HOPS"
        building: Building or rack identifier (e.g. "B11", "R12101")
        error_category: Type of error (e.g. "Mismatch", "Downlink", "Optics", "FEC")
        count: Number of occurrences
        source_file: Original filename processed
        processed_by: User or system that processed it
    """
    ensure_log_exists()
    
    new_row = {
        "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "hall": hall,
        "rack_type": rack_type,
        "building": building,
        "error_category": error_category,
        "count": count,
        "source_file": source_file,
        "processed_by": processed_by
    }
    
    # Extra safe Windows-friendly logging:
    # 1. Write to a brand new timestamped temp file
    # 2. Then replace the main file
    import shutil
    from openpyxl import load_workbook, Workbook

    try:
        # Always write to a fresh temp file first
        temp_path = LOG_FILE.parent / f"validation_error_log_{datetime.now().strftime('%Y%m%d_%H%M%S_%f')}.xlsx"

        if LOG_FILE.exists():
            # Copy existing data to the temp file
            shutil.copy2(LOG_FILE, temp_path)
            wb = load_workbook(temp_path)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "Error Log"
            headers = ["timestamp", "hall", "rack_type", "building", 
                       "error_category", "count", "source_file", "processed_by"]
            ws.append(headers)

        # Append the new row
        ws.append([
            datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            hall,
            rack_type,
            building,
            error_category,
            count,
            source_file,
            processed_by
        ])

        wb.save(temp_path)
        wb.close()

        # Replace the main log file
        if LOG_FILE.exists():
            try:
                LOG_FILE.unlink()
            except PermissionError:
                # File might still be locked briefly on Windows
                time.sleep(0.5)
                LOG_FILE.unlink()

        shutil.move(str(temp_path), str(LOG_FILE))

        print(f"Logged error → {LOG_FILE}")
        print("NOTE: If you want to open this Excel file, you must fully stop the Streamlit app first (Ctrl+C).")
        time.sleep(0.3)  # small grace period on Windows
        return True

    except Exception as e:
        print(f"Failed to write to error log: {e}")
        # Try to clean up any leftover temp file
        try:
            if 'temp_path' in locals() and temp_path.exists():
                temp_path.unlink()
        except:
            pass
        return False

def get_error_log():
    """Return the full error log as a DataFrame."""
    ensure_log_exists()
    return pd.read_excel(LOG_FILE)

def get_summary_by_hall_and_type():
    """Return aggregated summary grouped by Hall and Rack Type."""
    df = get_error_log()
    if df.empty:
        return pd.DataFrame()
    
    summary = df.groupby(["hall", "rack_type"]).agg({
        "count": "sum",
        "building": "nunique"
    }).reset_index()
    summary.columns = ["Hall", "Rack Type", "Total Errors", "Unique Buildings"]
    return summary.sort_values("Total Errors", ascending=False)