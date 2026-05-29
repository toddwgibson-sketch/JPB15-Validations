"""
lv_portal_formatter.py
Formats LV Portal validation exports into the same layout as the Slack report formatter.
Filters out compute (t0,host) rows — only processes t0,t1 links.
Outputs: Mispatches | Downlinks | Optics | FEC tabs with patch panel lookup from cutsheet.
"""

import sys, os, re, copy, json, time
from pathlib import Path
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

try:
    import tkinter as tk
    from tkinter import filedialog, messagebox
    HAS_TK = True
except ImportError:
    HAS_TK = False

# ── Colours ───────────────────────────────────────────────────────────────────
WHITE   = "FFFFFF"
YELLOW  = "FFFF00"
LOG_BG  = "EADCF8"
HDR_BG  = "1F4E79"
HDR_FG  = "FFFFFF"
SRC_BG  = "FCE4D6"
D1_BG   = "FFF2CC"
D2_BG   = "E2F0D9"
DEST_BG = "D9EAF7"
Z_BG    = "DDEBF7"
ACT_BG  = "FFC7CE"
EXP_BG  = "C6EFCE"
LR_BG   = "FFFFFF"
LR_LOG  = "FFFFFF"
PP_BG   = "FCE4D6"
PD_BG   = "FFF2CC"
TAB_MISS = "FF0000"
TAB_DOWN = "FFA500"
TAB_OPT  = "9933FF"
TAB_FEC  = "0070C0"

def fill(h):  return PatternFill("solid", fgColor=h)
def font(color="000000", bold=False, sz=9, italic=False):
    return Font(bold=bold, italic=italic, color=color, name="Arial", size=sz)
def center(): return Alignment(horizontal="center", vertical="center")
def left():   return Alignment(horizontal="left",   vertical="center")
def vcenter(): return Alignment(horizontal="left", vertical="center")

# ── Config ────────────────────────────────────────────────────────────────────
CONFIG_FILE = os.path.expanduser("~/.lv_portal_config.json")

def load_config():
    try:
        with open(CONFIG_FILE) as f: return json.load(f)
    except: return {}

def save_config(cfg):
    try:
        with open(CONFIG_FILE, 'w') as f: json.dump(cfg, f, indent=2)
    except: pass

# ── File pickers ──────────────────────────────────────────────────────────────
def pick_file(title, filetypes=None):
    if not HAS_TK:
        path = input(f"{title}\nEnter file path: ").strip().strip('"').strip("'")
        return path if os.path.isfile(path) else None
    root = tk.Tk(); root.withdraw(); root.attributes('-topmost', True)
    path = filedialog.askopenfilename(title=title,
           filetypes=filetypes or [("Excel files", "*.xlsx"), ("All files", "*.*")])
    root.destroy()
    return path or None

def pick_multiple_files(title, filetypes=None):
    if not HAS_TK:
        paths = []
        print(f"{title}\nEnter paths one per line, blank to finish:")
        while True:
            p = input("  Path: ").strip().strip('"').strip("'")
            if not p: break
            if os.path.isfile(p): paths.append(p)
        return paths
    root = tk.Tk(); root.withdraw(); root.attributes('-topmost', True)
    paths = filedialog.askopenfilenames(title=title,
            filetypes=filetypes or [("Excel files", "*.xlsx"), ("All files", "*.*")])
    root.destroy()
    return list(paths) if paths else []

def show_msg(title, msg, error=False):
    print(f"{'ERROR: ' if error else ''}{msg}")
    if HAS_TK:
        root = tk.Tk(); root.withdraw(); root.attributes('-topmost', True)
        (messagebox.showerror if error else messagebox.showinfo)(title, msg)
        root.destroy()

# ── Cutsheet loader ───────────────────────────────────────────────────────────
def build_lookup(paths):
    """Build T1 reverse lookup from one or more cutsheets."""
    if isinstance(paths, str): paths = [paths]
    t0     = {}   # (host, iface) -> t0_lbl  e.g. '30L'
    t1     = {}   # (host, iface) -> t1_lbl
    t1_rev = {}   # (t1_host, t1_iface) -> full dict

    t0_to_pp = {}  # (t0_host, t0_iface) -> PP dict
    for path in paths:
        wb = load_workbook(path, read_only=True)
        sheet = next((wb[n] for n in wb.sheetnames if 'installation' in n.lower()), wb[wb.sheetnames[0]])
        count = 0
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not row or len(row) < 9: continue
            lbl    = str(row[0]  or '').strip()
            dev_a  = str(row[1]  or '').strip()
            rack_a = str(row[2]  or '').strip()
            src    = str(row[3]  or '').strip()
            dmarc1 = str(row[4]  or '').strip()
            dmarc2 = str(row[5]  or '').strip()
            dest   = str(row[6]  or '').strip()
            dev_b  = str(row[7]  or '').strip()
            rack_b = str(row[8]  or '').strip()
            t1_lbl = str(row[10] or '').strip() if len(row) > 10 else ''

            if dev_a and lbl and re.match(r'\d+[LR]$', lbl):
                parts = dev_a.split()
                if len(parts) == 2:
                    k = (parts[0], parts[1])
                    t0[k] = lbl; t1[k] = t1_lbl
                    # Store PP data keyed by T0 host+iface for fallback lookup
                    t0_to_pp[k] = {
                        'source_port': src,
                        'dmarc1':      dmarc1,
                        'dmarc2':      dmarc2,
                        'dest_port':   dest,
                        'rack_b':      rack_b,
                        't1_lbl_pp':   t1_lbl,
                    }

            if dev_b and ' ' in dev_b:
                parts = dev_b.split()
                if len(parts) == 2:
                    t1_rev[(parts[0], parts[1])] = {
                        't0_lbl':      lbl,
                        'rack_a':      rack_a,
                        'source_port': src,
                        'dmarc1':      dmarc1,
                        'dmarc2':      dmarc2,
                        'dest_port':   dest,
                        'rack_b':      rack_b,
                        't1_lbl':      t1_lbl,
                    }
                    count += 1
        wb.close()
        print(f"  Loaded: {os.path.basename(path)} ({count} T1 entries)")
    return t0, t1, t1_rev, t0_to_pp

# ── Helpers ───────────────────────────────────────────────────────────────────
def is_compute(name):
    """True if this is a compute/host device — skip it."""
    return 'compute' in str(name or '').lower()


def is_real_compute_host(name):
    """True iff `name` looks like a real, fully-resolved compute hostname
    (e.g. 'jbp15-c1-b87-t0-r1-compute6'). Returns False for placeholders the LLDP
    neighbour returns when a host is unprovisioned or unreachable, such as
    'instance20260522182923', bare serial numbers ('2609XN10MR'), 'missing', or
    empty. These rows are ghost errors — the cable may be fine; the remote just
    hasn't been named yet.
    """
    s = str(name or '').strip().lower()
    if not s or s == 'missing': return False
    return 'compute' in s and s.startswith('jbp15-')
def cs_lookup(compute_lookup, host, port):
    """Lookup cutsheet entry.

    Two cutsheet conventions are supported:

    V1 ("one host owns all 4 lanes"): each physical compute hostname owns all 4 lanes
    of its ports (-1, -2, -3, -4). Logical even lanes (-2, -4) may not have their own
    rows; they're inherited from the physical odd partner on the same MPO head
    (-2 from -1, -4 from -3). NEVER cross MPO heads (-2 -> -3 would be wrong).

    V2 ("lanes split across neighbour trays"): the same MPO pair is split across two
    physically-separate compute trays at adjacent U's. One host owns -1/-3, its
    neighbour at U±1 in the same rack owns -2/-4. The validation report still uses
    the v1 naming (all lanes under one host), so when a (host, lane) misses we look
    at the same-rack neighbour at U±1 for that lane.

    Order of resolution:
      1. Direct hit.
      2. V2 neighbour (if host doesn't own this lane and a same-rack ±1U host does).
      3. V1 partner: same host, MPO-head partner lane (-1<->-2, -3<->-4).
      4. Host-level fallback (compute rack/U only, no T0 data).
    """
    if not compute_lookup: return {}
    cs = compute_lookup.get((host, port), {})
    if cs: return cs
    import re as _csl
    m = _csl.match(r'(slot\d+/port\d+)-(\d+)', port)
    if m:
        port_grp = m.group(1)
        lane = int(m.group(2))

        # Does `host` own this lane (per v2 meta index)? If not, prefer v2 neighbour
        # substitution before falling back to v1 partner — because in v2 the v1
        # partner would point to data from a different MPO on the same host, not the
        # actual physical connection.
        meta_all = compute_lookup.get('_v2_host_meta', {})
        rack_idx = compute_lookup.get('_v2_rack_u_host', {})
        meta = meta_all.get(host)
        host_owns_lane = bool(meta) and lane in meta['lanes_by_port'].get(port_grp, set())

        # V2 neighbour: same rack, ±1/±2 U, host that actually owns this lane.
        if not host_owns_lane and meta and rack_idx:
            for du in (1, -1, 2, -2):
                nh = rack_idx.get((meta['rack'], meta['u'] + du))
                if not nh: continue
                nmeta = meta_all.get(nh, {})
                if lane in nmeta.get('lanes_by_port', {}).get(port_grp, set()):
                    ncs = compute_lookup.get((nh, f"{port_grp}-{lane}"), {})
                    if ncs: return ncs

        # V1 partner: same host, MPO-head partner lane only (-1<->-2, -3<->-4).
        # Safe in v1 cutsheets; in v2 only reached when no host meta exists (i.e.
        # build_compute_lookup couldn't index this host).
        partner = lane - 1 if lane % 2 == 0 else lane + 1
        cs = compute_lookup.get((host, f"{port_grp}-{partner}"), {})
        if cs: return cs

    fb = compute_lookup.get('_host_fallback', {})
    return fb.get(host, {})



def compute_port_group(compute_host, compute_port):
    """Group key for compute port pairs: slot1/port2-1 and slot1/port2-2 share a key.
    Pairs: -1/-2 and -3/-4 within the same slotX/portY."""
    import re as _cpg
    m = _cpg.match(r'(slot\d+/port\d+-)(\d+)', str(compute_port or ''))
    if not m: return (compute_host, compute_port)
    lane = int(m.group(2))
    pair_base = lane if lane % 2 == 1 else lane - 1  # 1->1, 2->1, 3->3, 4->3
    return (compute_host, f"{m.group(1)}{pair_base}")


def parse_rack(rack_str):
    """Parse LV portal rack format '4909:9' -> ('Rack 4909', 'U9')"""
    s = str(rack_str or '').strip()
    if ':' in s:
        parts = s.split(':')
        return f"Rack {parts[0]}", f"U{parts[1]}"
    return s, ''

def get_t0_labels(host, iface, t0, t1):
    """Get L&R labels for a T0 host+interface pair.
    Physical = lane that appears directly in the cutsheet as DeviceA.
    Logical  = partner lane on the same MPO head.
    """
    key = (host, iface)
    t0_lbl = t0.get(key, '')
    t1_lbl = t1.get(key, '')
    is_phys = bool(t0_lbl)  # directly in cutsheet = physical

    if not is_phys:
        # Not in cutsheet directly — must be logical, get labels from partner
        m = re.match(r'(swp\d+)s(\d+)', iface)
        if m:
            base, lane = m.group(1), int(m.group(2))
            partner = {0:1,1:0,2:3,3:2}.get(lane)
            if partner is not None:
                pk = (host, f"{base}s{partner}")
                t0_lbl = t0.get(pk, '')
                t1_lbl = t1.get(pk, '')
    return t0_lbl, t1_lbl, is_phys

def get_mismatch_info(act_host, act_iface, t1_rev):
    """Look up possible patch panel from active interface."""
    mi = t1_rev.get((act_host, act_iface), {})
    if not mi:
        m = re.match(r'(swp\d+)s(\d+)', act_iface)
        if m:
            base, lane = m.group(1), int(m.group(2))
            partner = {0:1,1:0,2:3,3:2}.get(lane)
            if partner is not None:
                mi = t1_rev.get((act_host, f"{base}s{partner}"), {})
    return mi

# ── Border drawing ────────────────────────────────────────────────────────────
def draw_pair_borders(ws, lr_col=2, iface_col=1):
    """Draw medium borders around each physical+logical pair.
    Groups by (L&R value, base port) so same L&R from different hosts get separate borders.
    e.g. swp10s2 and swp10s3 on the same host group together,
    but swp10s2 from a different host starts a new group.
    """
    thin  = Side(style="thin",   color="AAAAAA")
    thick = Side(style="medium", color="555555")

    def group_key(row):
        lr  = ws.cell(row, lr_col).value or ''
        iface = str(ws.cell(row, iface_col).value or '')
        m = re.match(r'(swp\d+)s\d+', iface)
        base = m.group(1) if m else iface
        return (lr, base)

    dr = 2
    while dr <= ws.max_row:
        key = group_key(dr)
        grp_end = dr
        while grp_end + 1 <= ws.max_row and group_key(grp_end + 1) == key and key[0]:
            grp_end += 1
        for rr in range(dr, grp_end + 1):
            is_top = (rr == dr); is_bot = (rr == grp_end)
            for cc in range(1, ws.max_column + 1):
                ws.cell(rr, cc).border = Border(
                    top    = thick if is_top else thin,
                    bottom = thick if is_bot else Side(style=None),
                    left   = thick if cc == 1 else thin,
                    right  = thick if cc == ws.max_column else thin,
                )
        dr = grp_end + 1

# ── Sheet builders ────────────────────────────────────────────────────────────
def write_header_row(ws, headers, widths):
    """Write a styled header row. headers = [(label, bg_hex), ...], widths = [int, ...]"""
    for col, ((label, bg), w) in enumerate(zip(headers, widths), start=1):
        c = ws.cell(1, col)
        c.value = label; c.fill = fill(bg)
        c.font = Font(bold=True, color=WHITE, name="Arial", size=9)
        c.alignment = center()
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 20
    ws.freeze_panes = "A2"

def write_data_cell(ws, row, col, value, bg="FFFFFF", bold=False, sz=9, align=None):
    c = ws.cell(row, col)
    c.value = value
    c.fill  = fill(bg)
    c.font  = Font(bold=bold, color="000000", name="Arial", size=sz)
    c.alignment = align or center()

# ── Process LLDP rows ─────────────────────────────────────────────────────────
def process_lldp(ws_src, t0, t1, t1_rev, t0_to_pp=None):
    t0_to_pp = t0_to_pp or {}
    """
    Parse LLDP Mismatch + Link Down sheet.
    Returns (miss_rows, down_rows) — each a list of dicts.
    Skips compute rows.
    Pairs physical+logical by L&R value.
    """
    # LV portal columns:
    # 1=Device A Name, 2=Device B Port, 3=Device B Rack, 4=Expected Device B Name
    # 5=Device A Rack, 6=Expected Device B Rack, 7=Device B Name
    # 8=Device A Port, 9=LLDP Status, 10=Expected Device B Port, 11=Patch Panel Matrix

    raw = []
    for row in range(2, ws_src.max_row + 1):
        dev_a      = str(ws_src.cell(row, 1).value or '').strip()
        dev_b_port = str(ws_src.cell(row, 2).value or '').strip()
        dev_b_rack = str(ws_src.cell(row, 3).value or '').strip()
        exp_dev_b  = str(ws_src.cell(row, 4).value or '').strip()
        dev_a_rack = str(ws_src.cell(row, 5).value or '').strip()
        exp_rack_b = str(ws_src.cell(row, 6).value or '').strip()
        dev_b_name = str(ws_src.cell(row, 7).value or '').strip()
        dev_a_port = str(ws_src.cell(row, 8).value or '').strip()
        status     = str(ws_src.cell(row, 9).value or '').strip()
        exp_port_b = str(ws_src.cell(row, 10).value or '').strip()

        # Skip compute/host links
        if is_compute(dev_a) or is_compute(dev_b_name) or is_compute(exp_dev_b):
            continue
        if not dev_a or not dev_a_port:
            continue

        # Parse rack/elevation
        rack, elev = parse_rack(dev_a_rack)
        exp_rack, exp_elev = parse_rack(exp_rack_b)
        act_rack, act_elev = parse_rack(dev_b_rack)

        # L&R labels
        t0_lbl, t1_lbl, is_phys = get_t0_labels(dev_a, dev_a_port, t0, t1)

        # PP lookup — only on physical rows, logical rows copy from partner in second pass
        pp_own = {'source_port':'','dmarc1':'','dmarc2':'','dest_port':'','rack_b':'','t1_lbl':''}
        if is_phys and exp_dev_b and exp_port_b:
            mi_exp = t1_rev.get((exp_dev_b, exp_port_b), {})
            if not mi_exp:
                m = re.match(r'(swp\d+)s(\d+)', exp_port_b)
                if m:
                    base, lane = m.group(1), int(m.group(2))
                    partner = {0:1,1:0,2:3,3:2}.get(lane)
                    if partner is not None:
                        mi_exp = t1_rev.get((exp_dev_b, f"{base}s{partner}"), {})
            pp_own.update(mi_exp)

        # Mismatch info (where the cable actually is)
        # Only look up on physical rows — logical rows copy from partner in second pass
        mi = {}
        is_down = (status == 'INTERFACE_DOWN')
        if not is_down and is_phys and dev_b_name and dev_b_name != 'Unknown' and dev_b_port and dev_b_port != 'Unknown':
            mi = get_mismatch_info(dev_b_name, dev_b_port, t1_rev)

        raw.append({
            'host':       dev_a,
            'iface':      dev_a_port,
            'rack':       rack,
            'elev':       elev,
            't0_lbl':     t0_lbl,
            't1_lbl':     t1_lbl,
            'is_phys':    is_phys,
            'row_type':   'downlink' if is_down else 'mismatch',
            'status':     status,
            # Expected (Z side)
            'exp_host':   exp_dev_b,
            'exp_port':   exp_port_b,
            'exp_rack':   exp_rack,
            'exp_elev':   exp_elev,
            # Actual (what it connected to)
            'act_host':   dev_b_name,
            'act_port':   dev_b_port,
            'act_rack':   act_rack,
            'act_elev':   act_elev,
            # Patch panel (expected route)
            'source_port': pp_own.get('source_port', ''),
            'dmarc1':      pp_own.get('dmarc1', ''),
            'dmarc2':      pp_own.get('dmarc2', ''),
            'dest_port':   pp_own.get('dest_port', ''),
            'rack_b':      pp_own.get('rack_b', ''),
            't1_lbl_pp':   pp_own.get('t1_lbl', ''),
            # Possible (where cable actually is)
            'mi':          mi,
        })

    # Second pass: each logical row copies PP from its exact physical partner
    # Key = (host, physical_partner_iface) to avoid cross-contamination between
    # different MPO heads (s0↔s1 and s2↔s3 are separate pairs)
    pp_by_partner = {}  # (host, partner_iface) -> pp dict from physical row
    mi_by_partner = {}  # (host, partner_iface) -> mi dict

    for rd in raw:
        if not rd['is_phys']: continue  # only physical rows contribute
        m = re.match(r'(swp\d+)s(\d+)', rd['iface'])
        if m:
            lane = int(m.group(2))
            partner = {0:1,1:0,2:3,3:2}.get(lane)
            if partner is not None:
                # Key is the logical partner's iface
                key = (rd['host'], f"{m.group(1)}s{partner}")
                if rd['source_port']:
                    pp_by_partner[key] = {k: rd[k] for k in ['source_port','dmarc1','dmarc2','dest_port','rack_b','t1_lbl_pp']}
                if rd['mi']:
                    mi_by_partner[key] = rd['mi']

    for rd in raw:
        key = (rd['host'], rd['iface'])
        if not rd['is_phys'] and key in pp_by_partner:
            # Logical row — always copy from physical partner
            rd.update(pp_by_partner[key])
        elif not rd['is_phys'] and not rd['source_port']:
            # No physical partner in report — try cutsheet fallbacks
            mi_iface = re.match(r'(swp\d+)s(\d+)', rd['iface'])
            if mi_iface:
                base_i, lane_i = mi_iface.group(1), int(mi_iface.group(2))
                partner_i = {0:1,1:0,2:3,3:2}.get(lane_i)
                if partner_i is not None:
                    ck = (rd['host'], f"{base_i}s{partner_i}")
                    ct_pp = t0_to_pp.get(ck, {})
                    if ct_pp:
                        rd.update(ct_pp)
            if not rd['source_port'] and rd['exp_host'] and rd['exp_port']:
                em = re.match(r'(swp\d+)s(\d+)', rd['exp_port'])
                if em:
                    elane = int(em.group(2))
                    epartner = {0:1,1:0,2:3,3:2}.get(elane)
                    if epartner is not None:
                        pk = (rd['exp_host'], f"{em.group(1)}s{epartner}")
                        mi_exp = t1_rev.get(pk, {})
                        if mi_exp:
                            rd['source_port'] = mi_exp.get('source_port','')
                            rd['dmarc1']      = mi_exp.get('dmarc1','')
                            rd['dmarc2']      = mi_exp.get('dmarc2','')
                            rd['dest_port']   = mi_exp.get('dest_port','')
                            rd['rack_b']      = mi_exp.get('rack_b','')
                            rd['t1_lbl_pp']   = mi_exp.get('t1_lbl','')
        # Always copy mi from physical partner
        if not rd['is_phys'] and key in mi_by_partner:
            rd['mi'] = mi_by_partner[key]


    miss_rows = [r for r in raw if r['row_type'] == 'mismatch']
    down_rows = [r for r in raw if r['row_type'] == 'downlink']
    return miss_rows, down_rows

# ── Previous report comparison ────────────────────────────────────────────────
def get_prev_issues_lv(report_path):
    """Extract all issues from a previous LV portal report for recurring detection."""
    try:
        wb = load_workbook(report_path, read_only=True)
    except Exception as e:
        print(f"  Warning: could not load previous report: {e}")
        return set(), set(), set(), {}

    ws = next((wb[n] for n in wb.sheetnames if 'lldp' in n.lower() or 'mismatch' in n.lower()), None)
    prev_miss = set(); prev_down = set(); prev_rack_map = {}

    if ws:
        # Find cols by header
        hc = next((c for c in range(1,ws.max_column+1) if str(ws.cell(1,c).value or '').strip()=='Device A Name'), None)
        pc = next((c for c in range(1,ws.max_column+1) if str(ws.cell(1,c).value or '').strip()=='Device A Port'), None)
        sc = next((c for c in range(1,ws.max_column+1) if str(ws.cell(1,c).value or '').strip()=='LLDP Status'), None)
        rc = next((c for c in range(1,ws.max_column+1) if str(ws.cell(1,c).value or '').strip()=='Device A Rack'), None)
        if hc and pc and sc:
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row: continue
                h = str(row[hc-1] or '').strip()
                i = str(row[pc-1] or '').strip()
                st = str(row[sc-1] or '').strip()
                rack_raw = str(row[rc-1] or '').strip() if rc else ''
                rack, _ = parse_rack(rack_raw)
                if not h or not i: continue
                if is_compute(h): continue
                if st == 'INTERFACE_DOWN': prev_down.add((h,i))
                else:                      prev_miss.add((h,i))
                prev_rack_map[(h,i)] = rack

    # Optics
    ws_opt = next((wb[n] for n in wb.sheetnames if 'optic' in n.lower()), None)
    prev_opt = set()
    if ws_opt:
        dn = next((c for c in range(1,ws_opt.max_column+1) if str(ws_opt.cell(1,c).value or '').strip()=='Device Name'), None)
        dp = next((c for c in range(1,ws_opt.max_column+1) if str(ws_opt.cell(1,c).value or '').strip()=='Device Port'), None)
        if dn and dp:
            for row in ws_opt.iter_rows(min_row=2, values_only=True):
                if not row: continue
                h = str(row[dn-1] or '').strip(); i = str(row[dp-1] or '').strip()
                if h and i and not is_compute(h): prev_opt.add((h,i))

    wb.close()
    print(f"  Previous: {len(prev_miss)} mismatches, {len(prev_down)} downlinks, {len(prev_opt)} optics")
    return prev_miss, prev_down, prev_opt, prev_rack_map


def get_history_flag(host, iface, current_type, prev_miss, prev_down, prev_opt, rack=None):
    """Return (flag_text, flag_colour) based on previous report.
    Checks (rack, iface) and (host, iface) key formats."""
    prev_miss = prev_miss or set(); prev_down = prev_down or set(); prev_opt = prev_opt or set()
    rack_num = (rack.replace("Rack ","").split()[0] if rack else None)
    def _in(s):
        if not s or not isinstance(s, (set, frozenset)): return False
        try:
            return (rack_num, iface) in s or (host, iface) in s
        except TypeError:
            return False
    if current_type == 'mismatch':
        if _in(prev_miss): return "🔁 Recurring mismatch",  "FF6B6B"
        if _in(prev_down): return "⬆️ Was downlink",        "FFB347"
    elif current_type == 'downlink':
        if _in(prev_down): return "🔁 Recurring downlink",  "FF6B6B"
        if _in(prev_opt):  return "⚡ Was optic error",      "D35400"
        if _in(prev_miss): return "⬇️ Was mismatch",        "FFB347"
    elif current_type == 'optic':
        if _in(prev_opt):  return "🔁 Recurring optic",     "FF6B6B"
        if _in(prev_down): return "⬆️ Was downlink",        "FFB347"
        if _in(prev_miss): return "⬇️ Was mismatch",        "FFB347"
    return "", ""


# ── GPU / Compute cutsheet loader ────────────────────────────────────────────
def build_compute_lookup(paths):
    """Build lookup from GPU cutsheet. Auto-detects format by header names.
    Returns dict keyed by (t0_host, t0_iface) AND (comp_host, comp_port) -> enrichment dict.
    """
    import re as _fbr
    if isinstance(paths, str): paths = [paths]
    lookup = {}
    for path in paths:
        wb  = load_workbook(path, read_only=True, data_only=True)  # read-only: ~10x faster on large cutsheets
        ws  = wb[wb.sheetnames[0]]
        # Pull every row once (read-only random .cell access is O(n^2) — iterate instead)
        all_rows = list(ws.iter_rows(values_only=True))
        wb.close()
        if not all_rows:
            print(f"  Loaded GPU cutsheet: {os.path.basename(path)} (0 entries)")
            continue
        header_row = all_rows[0]
        ncol = len(header_row)
        hdr = {str(v or '').strip(): i + 1 for i, v in enumerate(header_row)}  # 1-indexed
        count = 0

        def _cell(rowvals, c):
            # c is 1-indexed; return '' if out of range or None
            if not c or c > len(rowvals): return ''
            v = rowvals[c - 1]
            return str(v).strip() if v is not None else ''

        if 'PortA' in hdr and 'PortB' in hdr:
            # New format: DeviceA=compute, DeviceB=T0 switch, all named columns
            c_comp_h = hdr['DeviceA']
            c_comp_p = hdr.get('PortA', 0)
            c_nic    = hdr.get('DeviceA Physical Port', 0)
            c_rack_a = hdr.get('RackA', 0)
            c_ohr    = hdr.get('OHR', 0)
            c_fdf    = hdr.get('FDF', 0)
            c_t0pp   = hdr.get('T0', 0)
            c_t0h    = hdr.get('DeviceB', 0)
            c_t0i    = hdr.get('PortB', 0)
            c_rack_b = hdr.get('RackB', 0)
            c_t0lr   = hdr.get('DeviceB Physical Port', 0)
            for rv in all_rows[1:]:
                comp_host = _cell(rv, c_comp_h); comp_port = _cell(rv, c_comp_p)
                t0_host   = _cell(rv, c_t0h);    t0_iface  = _cell(rv, c_t0i)
                if not comp_host or not t0_host or not t0_iface: continue
                entry = {
                    'lr':        _cell(rv, c_t0lr),  'nic_pos':   _cell(rv, c_nic),
                    'comp_host': comp_host,           'comp_port': comp_port,
                    'comp_rack': _cell(rv, c_rack_a), 'ohr':       _cell(rv, c_ohr),
                    'fdf':       _cell(rv, c_fdf),    't0_pp':     _cell(rv, c_t0pp),
                    't0_rack':   _cell(rv, c_rack_b), 't0_host':   t0_host, 't0_iface': t0_iface,
                }
                lookup[(t0_host, t0_iface)] = entry
                lookup[(comp_host, comp_port)] = entry
                count += 1
        else:
            # Old format: positional columns (col 4 = T0 Switch Port, col 27 = DeviceA, etc.)
            for rv in all_rows[1:]:
                if ncol < 36: break
                t0_col     = _cell(rv,  4)
                comp_dev_a = _cell(rv, 27)
                rack_a     = _cell(rv, 28)
                ohr        = _cell(rv, 29)
                fdf        = _cell(rv, 30)
                t0_pp      = _cell(rv, 31)
                rack_b     = _cell(rv, 33)
                nic_pos    = _cell(rv, 35)
                lr_label   = _cell(rv, 36)
                if not t0_col or ' ' not in t0_col: continue
                t0_host, t0_iface = t0_col.split()[0], t0_col.split()[1]
                comp_host = comp_dev_a.split()[0] if ' ' in comp_dev_a else comp_dev_a
                comp_port = comp_dev_a.split()[-1] if ' ' in comp_dev_a else ''
                entry = {
                    'lr': lr_label, 'nic_pos': nic_pos,
                    'comp_host': comp_host, 'comp_port': comp_port,
                    'comp_rack': rack_a, 'ohr': ohr, 'fdf': fdf,
                    't0_pp': t0_pp, 't0_rack': rack_b,
                    't0_host': t0_host, 't0_iface': t0_iface,
                }
                lookup[(t0_host, t0_iface)] = entry
                if comp_host and comp_port:
                    lookup[(comp_host, comp_port)] = entry
                count += 1
        print(f"  Loaded GPU cutsheet: {os.path.basename(path)} ({count} entries)")

    # Host-level fallback
    host_fallback = {}
    for (h, p), v in list(lookup.items()):
        if not isinstance(p, str): continue
        rack = v.get('comp_rack', '')
        if is_compute(h) and _fbr.match(r'Rack \d+ U\d+', rack):
            host_fallback[h] = v
    lookup['_host_fallback'] = host_fallback

    # V2 neighbour-tray index — for cutsheets where logical lanes live on a separate
    # compute hostname at an adjacent U in the same rack. Validation reports use the
    # v1 convention (all lanes named under one host), so we substitute when a
    # logical-lane lookup misses on the named host. Built once here for speed.
    # host -> {'rack': 'Rack 0701', 'u': 17, 'lanes_by_port': {slot/portN: set([1,3])}}
    # plus reverse index: (rack, u) -> host.
    host_meta = {}
    rack_u_to_host = {}
    for k, v in lookup.items():
        if not isinstance(k, tuple) or len(k) != 2: continue
        h, p = k
        if not isinstance(p, str): continue
        if not is_compute(h): continue
        m = _fbr.match(r'(slot\d+/port\d+)-(\d+)', p)
        if not m: continue
        port_grp, lane = m.group(1), int(m.group(2))
        ru = _fbr.match(r'(Rack \d+) U(\d+)', v.get('comp_rack', ''))
        if not ru: continue
        rack, u = ru.group(1), int(ru.group(2))
        meta = host_meta.setdefault(h, {'rack': rack, 'u': u, 'lanes_by_port': {}})
        meta['lanes_by_port'].setdefault(port_grp, set()).add(lane)
        rack_u_to_host[(rack, u)] = h
    lookup['_v2_host_meta']   = host_meta
    lookup['_v2_rack_u_host'] = rack_u_to_host

    # T0-host -> rack/U index. A single T0 switch always lives at one physical rack/U
    # regardless of which port you query, so any cutsheet row for that host gives the
    # right answer. Useful when the report's location field is blank AND the specific
    # (T0 host, iface) isn't directly in the cutsheet (e.g. only the shuffle partner
    # iface is stored).
    t0_rack_by_host = {}
    for k, v in lookup.items():
        if not isinstance(k, tuple) or len(k) != 2: continue
        th = v.get('t0_host', '')
        tr = v.get('t0_rack', '')
        if th and tr and th not in t0_rack_by_host:
            t0_rack_by_host[th] = tr
    lookup['_t0_rack_by_host'] = t0_rack_by_host
    return lookup


def draw_compute_borders(ws, host_col, port_col):
    """Draw borders grouping rows by compute port pairs (slot/port -1/-2 and -3/-4)."""
    thin  = Side(style="thin",   color="AAAAAA")
    thick = Side(style="medium", color="555555")
    def key(row):
        h = str(ws.cell(row, host_col).value or '')
        p = str(ws.cell(row, port_col).value or '')
        return compute_port_group(h, p)
    dr = 2
    while dr <= ws.max_row:
        k = key(dr)
        grp_end = dr
        while grp_end + 1 <= ws.max_row and key(grp_end + 1) == k and k[1]:
            grp_end += 1
        for rr in range(dr, grp_end + 1):
            is_top = (rr == dr); is_bot = (rr == grp_end)
            for cc in range(1, ws.max_column + 1):
                ws.cell(rr, cc).border = Border(
                    top    = thick if is_top else thin,
                    bottom = thick if is_bot else Side(style=None),
                    left   = thick if cc == 1 else thin,
                    right  = thick if cc == ws.max_column else thin,
                )
        dr = grp_end + 1


def draw_row_borders(ws):
    """Draw simple thin borders around each individual row — no grouping."""
    thin   = Side(style="thin",   color="AAAAAA")
    medium = Side(style="medium", color="555555")
    for rr in range(2, ws.max_row + 1):
        for cc in range(1, ws.max_column + 1):
            ws.cell(rr, cc).border = Border(
                top    = medium if rr == 2 else thin,
                bottom = medium if rr == ws.max_row else thin,
                left   = medium if cc == 1 else thin,
                right  = medium if cc == ws.max_column else thin,
            )



# ── Build Mispatches sheet ────────────────────────────────────────────────────
def build_mispatches_sheet(wb_out, rows, prev_miss=None, prev_down=None):
    prev_miss = prev_miss or set(); prev_down = prev_down or set()
    if not rows: return
    ws = wb_out.create_sheet("Mispatches")
    ws.sheet_properties.tabColor = TAB_MISS

    headers = [
        ("Interface",        HDR_BG),
        ("L&R",              HDR_BG),
        ("Rack",             HDR_BG),
        ("Elevation",        HDR_BG),
        ("Source_port",      "C0504D"),
        ("DMARC1",           "7F6000"),
        ("DMARC2",           "375623"),
        ("Destination_port", "17375E"),
        ("Z Interface",      "17375E"),
        ("Z L&R",            "17375E"),
        ("Z Rack",           "17375E"),
        ("Z Elevation",      "17375E"),
        ("Act. Interface",   "9C0006"),
        ("Act. Rack",        "9C0006"),
        ("Act. Elevation",   "9C0006"),
        ("Exp. Interface",   "375623"),
        ("Exp. Rack",        "375623"),
        ("Exp. Elevation",   "375623"),
        ("History",          "595959"),
    ]
    widths = [12,6,10,6, 30,28,28,30, 12,6,10,6, 12,10,6, 12,10,6, 22]
    write_header_row(ws, headers, widths)

    for r_idx, rd in enumerate(rows, start=2):
        ws.row_dimensions[r_idx].height = 15
        p     = rd['is_phys']
        bg    = "FFFFFF"
        lr_bg = LR_BG if p else LR_LOG

        hist_flag, hist_col = get_history_flag(rd['host'], rd['iface'], 'mismatch', prev_miss, prev_down, set())
        hist_bg = hist_col if hist_flag else bg
        vals = [
            rd['iface'], rd['t0_lbl'], rd['rack'], rd['elev'],
            rd['source_port'], rd['dmarc1'], rd['dmarc2'], rd['dest_port'],
            rd['exp_port'], rd['t1_lbl'], rd['exp_rack'], rd['exp_elev'],
            rd['act_port'], rd['act_rack'], rd['act_elev'],
            rd['exp_port'], rd['exp_rack'], rd['exp_elev'],
            hist_flag,
        ]
        all_bgs = [bg,lr_bg,bg,bg, bg,bg,bg,bg, bg,lr_bg,bg,bg,
                   ACT_BG,ACT_BG,ACT_BG, EXP_BG,EXP_BG,EXP_BG, hist_bg]

        for col, (val, cell_bg) in enumerate(zip(vals, all_bgs), start=1):
            c = ws.cell(r_idx, col)
            c.value = val; c.fill = fill(cell_bg)
            fg = WHITE if (col == len(vals) and hist_flag) else "000000"
            c.font = Font(bold=(col==2 or (col==len(vals) and hist_flag)), color=fg, name="Arial", size=9)
            c.alignment = center()

    draw_pair_borders(ws, lr_col=2)

# ── Build Downlinks sheet ─────────────────────────────────────────────────────
def build_downlinks_sheet(wb_out, rows, prev_miss=None, prev_down=None, prev_opt=None):
    prev_miss = prev_miss or set(); prev_down = prev_down or set(); prev_opt = prev_opt or set()
    if not rows: return
    ws = wb_out.create_sheet("Downlinks")
    ws.sheet_properties.tabColor = TAB_DOWN

    headers = [
        ("Interface",        HDR_BG),
        ("L&R",              HDR_BG),
        ("Rack",             HDR_BG),
        ("Elevation",        HDR_BG),
        ("Source_port",      "C0504D"),
        ("DMARC1",           "7F6000"),
        ("DMARC2",           "375623"),
        ("Destination_port", "17375E"),
        ("Z Interface",      "17375E"),
        ("Z L&R",            "17375E"),
        ("Z Rack",           "17375E"),
        ("Z Elevation",      "17375E"),
        ("History",          "595959"),
    ]
    widths = [12,6,10,6, 30,28,28,30, 12,6,10,6, 22]
    write_header_row(ws, headers, widths)

    for r_idx, rd in enumerate(rows, start=2):
        ws.row_dimensions[r_idx].height = 15
        p  = rd['is_phys']
        bg = "FFFFFF"; lr_bg = LR_BG if p else LR_LOG
        hist_flag, hist_col = get_history_flag(rd['host'], rd['iface'], 'downlink', prev_miss, prev_down, prev_opt)
        hist_bg = hist_col if hist_flag else bg
        vals = [
            rd['iface'], rd['t0_lbl'], rd['rack'], rd['elev'],
            rd['source_port'], rd['dmarc1'], rd['dmarc2'], rd['dest_port'],
            rd['exp_port'], rd['t1_lbl'], rd['exp_rack'], rd['exp_elev'],
            hist_flag,
        ]
        all_bgs = [bg, lr_bg, bg, bg, bg, bg, bg, bg, bg, lr_bg, bg, bg, hist_bg]
        for col, (val, cell_bg) in enumerate(zip(vals, all_bgs), start=1):
            c = ws.cell(r_idx, col)
            c.value = val; c.fill = fill(cell_bg)
            c.font = Font(bold=(col==2), color="000000", name="Arial", size=9)
            c.alignment = center()

    draw_pair_borders(ws, lr_col=2)

def parse_compute_pp_matrix(pp_str):
    """Parse the t0,host (compute) Patch Panel Matrix — the authoritative physical path
    the validation tool already resolved. Fixed 7-part bullet format:
      [0] compute host + port    e.g. 'jbp15-...-compute7 slot1/port1-[1+2]'
      [1] compute Rack/U         e.g. 'Rack 0701 U17'
      [2] OHR patch panel        e.g. 'PP.JBP15:8.DH8.R0791.U6.MPO43'
      [3] FDF patch panel        e.g. 'PP.JBP15:8.DH8.R0910.U2.MPO43'
      [4] T0 patch panel         e.g. 'PP.JBP15:8.DH8.R0909.U25.S9.B1'
      [5] T0 host + iface        e.g. 'jbp15-q2-p[1+2]-t0-r49 swp4s2'
      [6] T0 Rack/U              e.g. 'Rack 0909 U1'
    Returns {} if not parseable.
    """
    if not pp_str or pp_str.strip().upper() == 'N/A': return {}
    if '•' not in pp_str: return {}
    parts = [p.strip() for p in pp_str.split('•')]
    if len(parts) < 7: return {}
    comp_hp = parts[0]
    comp_host = comp_hp.split()[0] if ' ' in comp_hp else comp_hp
    comp_port = comp_hp.split()[-1] if ' ' in comp_hp else ''
    t0_hp = parts[5]
    t0_host  = t0_hp.split()[0] if ' ' in t0_hp else t0_hp
    t0_iface = t0_hp.split()[-1] if ' ' in t0_hp else ''
    return {
        'comp_host_port': comp_hp,
        'comp_host':      comp_host,
        'comp_port':      comp_port,
        'comp_rack_full': parts[1],
        'ohr':            parts[2],
        'fdf':            parts[3],
        't0_pp':          parts[4],
        't0_host_port':   t0_hp,
        't0_host':        t0_host,
        't0_iface':       t0_iface,
        't0_rack_full':   parts[6],
    }


def parse_pp_matrix(pp_str):
    """Parse LV portal PP Matrix: 'host port • RackA • src • dmarc1 • dmarc2 • dest • T1host T1port • RackB'"""
    if not pp_str or pp_str.strip().upper() == 'N/A': return {}
    parts = [p.strip() for p in pp_str.split('•')]
    if len(parts) < 6: return {}
    # parts[6] = 't1_host t1_iface' e.g. 'jbp15-q2-p1-t1-r45 swp64s2'
    t1_host_port = parts[6].strip() if len(parts) > 6 else ''
    t1_host = ''; t1_iface = ''
    if ' ' in t1_host_port:
        t1_host  = t1_host_port.split()[0]
        t1_iface = t1_host_port.split()[-1]

    return {
        'rack_a':      parts[1] if len(parts) > 1 else '',
        'source_port': parts[2] if len(parts) > 2 else '',
        'dmarc1':      parts[3] if len(parts) > 3 else '',
        'dmarc2':      parts[4] if len(parts) > 4 else '',
        'dest_port':   parts[5] if len(parts) > 5 else '',
        't1_host_port':t1_host_port,
        't1_host':     t1_host,
        't1_iface':    t1_iface,
        'rack_b':      parts[7] if len(parts) > 7 else '',
    }

# ── Build Optics sheet ────────────────────────────────────────────────────────
def build_optics_sheet(wb_out, ws_src, t0, t1, t1_rev, downlink_set, t0_to_pp=None, prev_miss=None, prev_down=None, prev_opt=None):
    t0_to_pp = t0_to_pp or {}
    prev_miss = prev_miss or set(); prev_down = prev_down or set(); prev_opt = prev_opt or set()
    if not ws_src: return
    # LV portal optics cols:
    # 1=Device Port, 2=Rx Power, 3=Device Name, 4=Tx Power, 5=Transceiver, 6=Patch Panel Matrix

    # Find columns by header name — handles different LV portal versions
    ncols = ws_src.max_column
    col_port = col_rx = col_dev = col_tx = col_pp = None
    for c in range(1, ncols+1):
        hv = str(ws_src.cell(1,c).value or '').strip()
        if hv == 'Device Port':         col_port = c
        elif hv == 'Rx Power':          col_rx   = c
        elif hv == 'Device Name':       col_dev  = c
        elif hv == 'Tx Power':          col_tx   = c
        elif hv == 'Patch Panel Matrix':col_pp   = c
    if not col_port or not col_dev:
        print(f"  Optics: could not find required columns"); return

    # First pass — collect all rows, parse PP, skip N/A and compute
    opt_rows = []
    for row in range(2, ws_src.max_row + 1):
        port     = str(ws_src.cell(row, col_port).value or '').strip()
        rx_power = str(ws_src.cell(row, col_rx).value   or '').strip() if col_rx else ''
        dev_name = str(ws_src.cell(row, col_dev).value  or '').strip()
        tx_power = str(ws_src.cell(row, col_tx).value   or '').strip() if col_tx else ''
        pp_str   = str(ws_src.cell(row, col_pp).value   or '').strip() if col_pp else ''
        # Skip non-swp ports (e.g. Ethernet50/1) — only process swp interfaces
        if not re.match(r'swp', port): continue

        if not port or not dev_name: continue
        if is_compute(dev_name) or is_compute(pp_str): continue
        if pp_str.upper() == 'N/A' or not pp_str: continue  # skip N/A

        pp = parse_pp_matrix(pp_str)
        if not pp.get('source_port'): continue  # no useful PP data

        t0_lbl, _, is_p = get_t0_labels(dev_name, port, t0, t1)
        is_dl = (dev_name, port) in downlink_set

        # Look up T1 physical port label from t1_rev
        t1_lbl_z = ''
        t1_h = pp.get('t1_host',''); t1_i = pp.get('t1_iface','')
        if t1_h and t1_i:
            t1_lbl_z = t1_rev.get((t1_h, t1_i), {}).get('t1_lbl', '')
            if not t1_lbl_z:
                m2 = re.match(r'(swp\d+)s(\d+)', t1_i)
                if m2:
                    partner2 = {0:1,1:0,2:3,3:2}.get(int(m2.group(2)))
                    if partner2 is not None:
                        t1_lbl_z = t1_rev.get((t1_h, f"{m2.group(1)}s{partner2}"), {}).get('t1_lbl', '')

        opt_rows.append({
            'host':       dev_name, 'iface': port,
            't0_lbl':     t0_lbl,  'is_phys': is_p,
            'rx_power':   rx_power, 'tx_power': tx_power,
            'rack_a':     pp.get('rack_a',''),
            'source_port':pp.get('source_port',''),
            'dmarc1':     pp.get('dmarc1',''),
            'dmarc2':     pp.get('dmarc2',''),
            'dest_port':  pp.get('dest_port',''),
            't1_iface':   t1_i,
            't1_lbl_z':   t1_lbl_z,
            'rack_b':     pp.get('rack_b',''),
            'is_dl':      is_dl,
        })

    # Second pass — logical rows copy PP from physical partner by base port
    pp_base = {}
    for rd in opt_rows:
        if rd['source_port']:
            m = re.match(r'(swp\d+)s\d+', rd['iface'])
            if m: pp_base[(rd['host'], m.group(1))] = {
                k: rd[k] for k in ['rack_a','source_port','dmarc1','dmarc2','dest_port','t1_iface','t1_lbl_z','rack_b']}
    for rd in opt_rows:
        if not rd['source_port']:
            m = re.match(r'(swp\d+)s\d+', rd['iface'])
            if m:
                k = (rd['host'], m.group(1))
                if k in pp_base: rd.update(pp_base[k])

    if not opt_rows: return

    ws = wb_out.create_sheet("Optics")
    ws.sheet_properties.tabColor = TAB_OPT

    headers = [
        ("Interface",        HDR_BG),
        ("L&R",              HDR_BG),
        ("Rack",             HDR_BG),
        ("Rx Power",         "7030A0"),
        ("Source_port",      "C0504D"),
        ("DMARC1",           "7F6000"),
        ("DMARC2",           "375623"),
        ("Destination_port", "17375E"),
        ("Z Interface",      "17375E"),
        ("Z L&R",            "17375E"),
        ("Z Rack",           "17375E"),
        ("DL Flag",          "595959"),
        ("History",          "595959"),
    ]
    widths = [12,6,14, 30, 30,28,28,30, 12,6,14, 24,22]
    write_header_row(ws, headers, widths)

    for out_row, rd in enumerate(opt_rows, start=2):
        ws.row_dimensions[out_row].height = 15
        is_dl = rd['is_dl']
        row_bg = "C8C8C8" if is_dl else "FFFFFF"
        dl_flag = "⬇️ Also Downlink — skip" if is_dl else ''
        txt_fg = "888888" if is_dl else "000000"

        vals = [rd['iface'], rd['t0_lbl'], rd['rack_a'],
                rd['rx_power'],
                rd['source_port'], rd['dmarc1'], rd['dmarc2'], rd['dest_port'],
                rd['t1_iface'], rd['t1_lbl_z'], rd['rack_b'], dl_flag, '']
        bgs = [row_bg,"FFFFFF",row_bg, row_bg,
               row_bg,row_bg,row_bg,row_bg, row_bg,"FFFFFF",row_bg,
               "C8C8C8" if is_dl else row_bg, row_bg]

        for col, (val, bg) in enumerate(zip(vals, bgs), start=1):
            c = ws.cell(out_row, col)
            c.value = val; c.fill = fill(bg)
            c.font = Font(color=txt_fg, name="Arial", size=9)
            c.alignment = center()

    draw_pair_borders(ws, lr_col=2)

# ── Build FEC sheet ───────────────────────────────────────────────────────────
def build_fec_sheet(wb_out, ws_src, t0, t1, downlink_set):
    if not ws_src: return
    # LV portal FEC cols:
    # 1=Device Port, 2=PRE_FEC_BER, 3=Remote Device, 4=Remote Interface
    # 5=Device Rack, 6=Device Name, 7=Lock Status, 8=Patch Panel Matrix
    ws = wb_out.create_sheet("FEC Errors")
    ws.sheet_properties.tabColor = TAB_FEC

    headers = [
        ("Interface",     HDR_BG),
        ("L&R",           HDR_BG),
        ("Rack",          HDR_BG),
        ("Elevation",     HDR_BG),
        ("Pre-FEC BER",   "7030A0"),
        ("Lock Status",   "7030A0"),
        ("Z Interface",   "17375E"),
        ("DL Flag",       "595959"),
        ("History",       "595959"),
    ]
    widths = [12,6,10,6, 20,40, 12, 24,22]
    write_header_row(ws, headers, widths)

    out_row = 2
    for row in range(2, ws_src.max_row + 1):
        port       = str(ws_src.cell(row, 1).value or '').strip()
        ber        = str(ws_src.cell(row, 2).value or '').strip()
        remote_dev = str(ws_src.cell(row, 3).value or '').strip()
        remote_if  = str(ws_src.cell(row, 4).value or '').strip()
        dev_rack   = str(ws_src.cell(row, 5).value or '').strip()
        dev_name   = str(ws_src.cell(row, 6).value or '').strip()
        lock_status= str(ws_src.cell(row, 7).value or '').strip()

        if not port or not dev_name: continue
        if is_compute(dev_name) or is_compute(remote_dev): continue

        t0_lbl, _, is_p = get_t0_labels(dev_name, port, t0, t1)
        rack, elev = parse_rack(dev_rack)
        is_dl = (dev_name, port) in downlink_set
        row_bg = "C8C8C8" if is_dl else "FFFFFF"

        ws.row_dimensions[out_row].height = 15
        dl_flag = "⬇️ Also Downlink — skip" if is_dl else ''
        vals = [port, t0_lbl, rack, elev, ber, lock_status, remote_if, dl_flag, '']
        bgs  = [row_bg,"FFFFFF",row_bg,row_bg, row_bg,row_bg, row_bg,
                "C8C8C8" if is_dl else row_bg, row_bg]

        for col, (val, bg) in enumerate(zip(vals, bgs), start=1):
            c = ws.cell(out_row, col)
            c.value = val; c.fill = fill(bg)
            c.font = Font(color="888888" if is_dl else "000000", name="Arial", size=9)
            c.alignment = center()
        out_row += 1

# ── Main ──────────────────────────────────────────────────────────────────────
# ── Ghost host detection ─────────────────────────────────────────────────────
GHOST_THRESHOLD = 16  # hosts with this many errors are ghost/not-yet-online switches

def get_ghost_hosts(ws_src):
    """Return set of compute hostnames that appear 16+ times — ghost switches."""
    from collections import Counter
    counts = Counter()
    # Detect column: new format uses 'Expected Device B Name', old uses col 4
    hdr = {str(ws_src.cell(1,c).value or '').strip(): c for c in range(1, ws_src.max_column+1)}
    exp_b_col = hdr.get('Expected Device B Name', 4)
    for row in range(2, ws_src.max_row + 1):
        exp_b = str(ws_src.cell(row, exp_b_col).value or '').strip()
        # Skip fake 'missing' entries — real downlinks now in Interface Down tab
        cur_b_col = hdr.get('Current Device B Name', exp_b_col - 3)
        cur_b = str(ws_src.cell(row, cur_b_col).value or '').strip()
        if is_compute(exp_b) and cur_b.lower() != 'missing':
            counts[exp_b] += 1
    ghosts = {host for host, cnt in counts.items() if cnt >= GHOST_THRESHOLD}
    if ghosts:
        print(f"  Ghost hosts detected ({GHOST_THRESHOLD}+ errors): {ghosts}")
    return ghosts


# ── Process compute (T0,Host) LLDP rows ──────────────────────────────────────
def process_interface_down(ws_src, compute_lookup=None):
    """Process Interface Down Errors tab — same column layout as Optics tab.
    Returns (real_rows, ghost_rows, ghost_hosts) in the same format as process_compute_lldp.
    Columns: Remote Device Name, Remote Device Port, Source Device Name,
             Source Device Location, Source Device Port, Issue, Patch Panel Matrix
    """
    import re
    compute_lookup = compute_lookup or {}
    if not ws_src: return [], [], set()

    hdr = {str(ws_src.cell(1,c).value or '').strip(): c for c in range(1, ws_src.max_column+1)}
    c_remote_dev  = hdr.get('Remote Device Name', 1)
    c_remote_port = hdr.get('Remote Device Port', 2)
    c_src_dev     = hdr.get('Source Device Name', 3)
    c_src_loc     = hdr.get('Source Device Location', 4)
    c_src_port    = hdr.get('Source Device Port', 5)
    c_pp          = hdr.get('Patch Panel Matrix', 7)

    # Count per compute host to detect ghosts
    from collections import Counter
    counts = Counter()
    for row in range(2, ws_src.max_row + 1):
        remote = str(ws_src.cell(row, c_remote_dev).value or '').strip()
        if is_compute(remote): counts[remote] += 1
    ghost_hosts = {h for h,c in counts.items() if c >= GHOST_THRESHOLD}
    if ghost_hosts:
        print(f"  Ghost hosts detected ({GHOST_THRESHOLD}+ errors): {ghost_hosts}")

    real_rows = []; ghost_rows = []
    for row in range(2, ws_src.max_row + 1):
        exp_b    = str(ws_src.cell(row, c_remote_dev).value  or '').strip()
        comp_port= str(ws_src.cell(row, c_remote_port).value or '').strip()
        dev_a    = str(ws_src.cell(row, c_src_dev).value     or '').strip()
        loc_a    = str(ws_src.cell(row, c_src_loc).value     or '').strip()
        t0_iface = str(ws_src.cell(row, c_src_port).value    or '').strip()
        pp       = str(ws_src.cell(row, c_pp).value          or '').strip() if c_pp else ''

        if not exp_b or not dev_a: continue
        if not is_compute(exp_b): continue

        t0_rack, t0_elev = _parse_location(loc_a)
        t0_lbl, _, is_phys = get_t0_labels(dev_a, t0_iface, {}, {})

        # Cutsheet enrichment
        cs = cs_lookup(compute_lookup, dev_a, t0_iface)
        # See process_compute_lldp for full explanation: don't copy T0 location
        # from a cutsheet row that matches on compute port but a different T0.
        cs_t0_match = bool(cs) and cs.get('t0_host') == dev_a and cs.get('t0_iface') == t0_iface

        if not cs.get('ohr') and comp_port:
            cs_rev = cs_lookup(compute_lookup, exp_b, comp_port)
            if cs_rev.get('ohr'):
                cs = {**cs_rev, **{k:v for k,v in cs.items() if v}}
        if not cs.get('comp_rack'):
            fb = compute_lookup.get('_host_fallback', {})
            cs_fb = fb.get(exp_b, {})
            if cs_fb.get('comp_rack'): cs = {**cs, 'comp_rack': cs_fb['comp_rack']}

        if cs_t0_match and cs.get('t0_rack'):
            t0_rack = cs['t0_rack']
        # Final fallback: if T0 rack is still blank (report blank + specific iface not
        # in cutsheet), use any row for the same T0 host — all lanes share rack/U.
        if not t0_rack:
            t0_rack = compute_lookup.get('_t0_rack_by_host', {}).get(dev_a, '')
        import re as _rlre2
        _rls2 = _rlre2.match(r'(Rack \d+) (U\d+)', t0_rack)
        if _rls2: t0_rack, t0_elev = _rls2.group(1), _rls2.group(2)

        comp_rack_full = cs.get('comp_rack', '')
        import re as _cr2re
        _crm = _cr2re.match(r'Rack (\d+) U(\d+)', comp_rack_full)
        comp_rack = f"Rack {_crm.group(1)}" if _crm else comp_rack_full
        comp_elev = f"U{_crm.group(2)}"      if _crm else ''

        if cs_t0_match and cs.get('lr'): t0_lbl = cs['lr']
        # L&R is per-MPO-pair: inherit from the T0 partner lane (swpXs0<->s1, s2<->s3)
        # when this lane's label wasn't a direct match. Same MPO pair = same label.
        if not t0_lbl:
            _plm = _rlre2.match(r'(swp\d+s)(\d+)', t0_iface)
            if _plm:
                _pl = int(_plm.group(2)); _pp = {0:1,1:0,2:3,3:2}.get(_pl)
                if _pp is not None:
                    _ppcs = compute_lookup.get((dev_a, f"{_plm.group(1)}{_pp}"), {})
                    if _ppcs.get('lr'): t0_lbl = _ppcs['lr']

        import re as _dre2
        _dm2 = _dre2.match(r'slot\d+/port\d+-(\d+)', comp_port)
        comp_is_phys = (int(_dm2.group(1)) % 2 == 1) if _dm2 else True

        # Patch Panel Matrix is authoritative for the physical path; cutsheet is fallback
        ohr = cs.get('ohr', ''); fdf = cs.get('fdf', ''); t0_pp = cs.get('t0_pp', '')
        ppm = parse_compute_pp_matrix(pp)
        if ppm:
            if ppm.get('ohr'):   ohr   = ppm['ohr']
            if ppm.get('fdf'):   fdf   = ppm['fdf']
            if ppm.get('t0_pp'): t0_pp = ppm['t0_pp']

        rd = {
            'host': dev_a, 'iface': t0_iface, 't0_lbl': t0_lbl, 'is_phys': is_phys,
            'rack': t0_rack, 'elev': t0_elev,
            # v2 substitution: when cutsheet's compute host differs from reported (the
            # lane actually lives on a neighbour tray), show the cutsheet host so the
            # tech goes to the right place.
            'exp_host': cs.get('comp_host') or exp_b,
            'comp_port': comp_port,
            'comp_rack': comp_rack, 'comp_elev': comp_elev,
            'cur_b': 'interface down',
            'comp_is_phys': comp_is_phys,
            'nic_pos':  cs.get('nic_pos', ''),
            'ohr':      ohr,
            'fdf':      fdf,
            't0_pp':    t0_pp,
        }
        if exp_b in ghost_hosts:
            ghost_rows.append(rd)
        else:
            real_rows.append(rd)

    return real_rows, ghost_rows, ghost_hosts


def _parse_location(loc_str):
    """Parse rack/U from either format:
    New: 'jbp15:9114:5'  -> Rack 9114, U5
    Old: 'rackNumber: 9114\nrackElevation: 5' -> Rack 9114, U5
    """
    import re as _re
    # New format: site:rack:elevation
    m = _re.match(r'[^:]+:(\d+):(\d+)', str(loc_str or '').strip())
    if m: return f"Rack {m.group(1)}", f"U{m.group(2)}"
    # Old format: multiline rackNumber/rackElevation
    rack_m = _re.search(r'rackNumber: *(\d+)', str(loc_str or ''))
    elev_m = _re.search(r'rackElevation: *(\d+)', str(loc_str or ''))
    rack = f"Rack {rack_m.group(1)}" if rack_m else ''
    elev = f"U{elev_m.group(1)}"     if elev_m else ''
    return rack, elev


def process_compute_lldp(ws_src, compute_lookup=None):
    """Returns (real_rows, ghost_rows, ghost_hosts). Handles both old and new column formats."""
    import re
    compute_lookup = compute_lookup or {}
    ghost_hosts = get_ghost_hosts(ws_src)
    real_rows = []; ghost_rows = []; mispatch_rows = []

    # Detect format by header names
    hdr = {str(ws_src.cell(1,c).value or '').strip(): c for c in range(1, ws_src.max_column+1)}
    new_fmt = 'Device A Name' in hdr  # New format has named headers

    if new_fmt:
        col_dev_a   = hdr.get('Device A Name', 1)
        col_loc_a   = hdr.get('Device A Location', 2)
        col_port_a  = hdr.get('Device A Port', 3)
        col_cur_b   = hdr.get('Current Device B Name', 4)
        col_cur_loc = hdr.get('Current B Location', 5)
        col_cur_port= hdr.get('Current Device B Port', 6)
        col_exp_b   = hdr.get('Expected Device B Name', 7)
        col_exp_loc = hdr.get('Expected B Location', 8)
        col_exp_port= hdr.get('Expected Device B Port', 9)
        col_pp      = hdr.get('Patch Panel Matrix', 10)
        col_err     = hdr.get('Error Message', 11)
    else:
        col_dev_a=1; col_loc_a=3; col_port_a=2; col_cur_b=8
        col_cur_loc=None; col_cur_port=None
        col_exp_b=4; col_exp_loc=6; col_exp_port=5; col_pp=None; col_err=10

    for row in range(2, ws_src.max_row + 1):
        dev_a   = str(ws_src.cell(row, col_dev_a).value or '').strip()
        loc_a   = str(ws_src.cell(row, col_loc_a).value or '').strip()
        port_a  = str(ws_src.cell(row, col_port_a).value or '').strip()
        cur_b   = str(ws_src.cell(row, col_cur_b).value or '').strip()
        cur_port= str(ws_src.cell(row, col_cur_port).value or '').strip() if col_cur_port else ''
        cur_loc = str(ws_src.cell(row, col_cur_loc).value or '').strip() if col_cur_loc else ''
        exp_b   = str(ws_src.cell(row, col_exp_b).value or '').strip()
        exp_loc = str(ws_src.cell(row, col_exp_loc).value or '').strip()
        exp_port= str(ws_src.cell(row, col_exp_port).value or '').strip()
        err_msg = str(ws_src.cell(row, col_err).value or '').strip()
        pp      = str(ws_src.cell(row, col_pp).value or '').strip() if col_pp else ''
        if not dev_a: continue
        if not (is_compute(exp_b) or is_compute(cur_b)): continue
        # Mispatch = a *real* (non-placeholder) compute host is connected, but it's the
        # WRONG one (host or port differs from expected). Placeholders like
        # 'instance20260522182923', bare serial numbers, or 'missing' are unprovisioned/
        # ghost neighbours — the cable may well be correct, the remote just hasn't been
        # named yet. Route those to Ghost Links below instead.
        is_mispatch = (is_real_compute_host(cur_b)
                       and (cur_b != exp_b or (cur_port and cur_port != exp_port)))
        # Placeholder current device (present but not a real compute name) → ghost
        cur_is_placeholder = (cur_b != '' and cur_b.lower() != 'missing'
                              and not is_real_compute_host(cur_b))
        # Skip fake downlinks (current = 'missing') — handled by Interface Down tab
        if cur_b.lower() == 'missing': continue

        t0_iface = port_a if re.match(r'swp', port_a) else (
            re.search(r'(swp\d+s\d+)', err_msg).group(1)
            if re.search(r'(swp\d+s\d+)', err_msg) else 'Unknown')

        t0_rack, t0_elev = _parse_location(loc_a)
        comp_port = exp_port
        comp_rack, comp_elev = _parse_location(exp_loc)
        t0_lbl, _, is_phys = get_t0_labels(dev_a, t0_iface, {}, {})

        # Enrich from GPU cutsheet — T0 forward lookup then compute reverse
        cs = cs_lookup(compute_lookup, dev_a, t0_iface)
        # cs_t0_match: True if cs's T0 host/iface matches the validation report row.
        # When False, the row was found via compute-port fallback and points to a
        # DIFFERENT T0 switch — its t0_rack / t0_iface / lr must NOT be copied.
        # OHR / FDF / T0 PP describe the physical MPO trunk pair and are still valid.
        cs_t0_match = bool(cs) and cs.get('t0_host') == dev_a and cs.get('t0_iface') == t0_iface

        # If T0 lookup missing PP data, try compute-side reverse
        if not cs.get('ohr') and comp_port:
            cs_rev = cs_lookup(compute_lookup, exp_b, comp_port)
            if cs_rev.get('ohr'):
                cs = {**cs_rev, **{k:v for k,v in cs.items() if v}}
                # cs_t0_match unchanged: still depends on whether the original cs aligned

        # 3. If still no rack, use host-level fallback
        if not cs.get('comp_rack'):
            fb = compute_lookup.get('_host_fallback', {})
            cs_fb = fb.get(exp_b, {})
            if cs_fb.get('comp_rack'): cs = {**cs, 'comp_rack': cs_fb['comp_rack']}
        if cs:
            # Compute-side fields: always safe to copy
            comp_rack = cs.get('comp_rack', comp_rack)
            comp_port = cs.get('comp_port', comp_port)
            # T0-side location: only when cutsheet row actually matches this T0 host+iface
            if cs_t0_match:
                t0_lbl  = cs.get('lr',      t0_lbl)
                t0_rack = cs.get('t0_rack', t0_rack)
            is_phys = bool(compute_lookup.get((dev_a, t0_iface)))  # physical = directly in CS

        # Final fallback: if T0 rack still blank (report had no location AND no direct
        # cutsheet hit for this iface), use any cutsheet row for the same T0 host —
        # all lanes share rack/U because a T0 switch occupies one physical slot.
        if not t0_rack:
            t0_rack = compute_lookup.get('_t0_rack_by_host', {}).get(dev_a, '')

        # L&R is per-MPO-pair: a logical T0 lane (swpXs1/s3) shares its label with the
        # physical partner lane (swpXs0/s2). If we didn't get L&R from a direct match,
        # inherit it from the T0 partner lane on the SAME switch (safe — same MPO pair).
        if not t0_lbl:
            import re as _lrre
            _lm = _lrre.match(r'(swp\d+s)(\d+)', t0_iface)
            if _lm:
                _lane = int(_lm.group(2))
                _partner = {0:1, 1:0, 2:3, 3:2}.get(_lane)
                if _partner is not None:
                    _pcs = compute_lookup.get((dev_a, f"{_lm.group(1)}{_partner}"), {})
                    if _pcs.get('lr'): t0_lbl = _pcs['lr']
        # Split combined Rack+U for both T0 and compute racks
        import re as _rlre
        _rls = _rlre.match(r'(Rack \d+) (U\d+)', t0_rack)
        if _rls: t0_rack, t0_elev = _rls.group(1), _rls.group(2)
        _rlc = _rlre.match(r'(Rack \d+) (U\d+)', comp_rack)
        if _rlc: comp_rack, comp_elev = _rlc.group(1), _rlc.group(2)

        # Compute port physical: odd lane = physical
        import re as _dre
        _dm = _dre.match(r'slot\d+/port\d+-(\d+)', comp_port)
        comp_is_phys = (int(_dm.group(1)) % 2 == 1) if _dm else True

        # Authoritative physical path from the report's Patch Panel Matrix (when present).
        # It already resolved the exact MPO trunk, so prefer it for OHR/FDF/T0 PP over the
        # cutsheet. Match the T0 host loosely (matrix uses 'p[1+2]' shorthand for the pair).
        ohr = cs.get('ohr', ''); fdf = cs.get('fdf', ''); t0_pp = cs.get('t0_pp', '')
        ppm = parse_compute_pp_matrix(pp)
        if ppm:
            if ppm.get('ohr'):   ohr   = ppm['ohr']
            if ppm.get('fdf'):   fdf   = ppm['fdf']
            if ppm.get('t0_pp'): t0_pp = ppm['t0_pp']

        # For a mispatch, also resolve the ACTUAL (wrongly-plugged) port's physical path
        # from the cutsheet: its 3 PP hops plus the T0 it really lands on. This is the
        # cable a tech will physically find at the actual compute port.
        act_ohr = act_fdf = act_t0pp = act_t0_host = act_t0_iface = act_t0_rack = act_t0_elev = ''
        act_lr = act_nic = act_comp_rack = act_comp_elev = ''
        if is_mispatch and cur_b and cur_port:
            acs = cs_lookup(compute_lookup, cur_b, cur_port)
            if acs:
                act_ohr     = acs.get('ohr', '')
                act_fdf     = acs.get('fdf', '')
                act_t0pp    = acs.get('t0_pp', '')
                act_t0_host = acs.get('t0_host', '')
                act_t0_iface= acs.get('t0_iface', '')
                act_lr      = acs.get('lr', '')
                act_nic     = acs.get('nic_pos', '')
                _ar = _rlre.match(r'(Rack \d+) (U\d+)', acs.get('t0_rack', ''))
                if _ar: act_t0_rack, act_t0_elev = _ar.group(1), _ar.group(2)
                else:   act_t0_rack = acs.get('t0_rack', '')
                # actual cable's compute-end rack/U (its real home rack)
                _acr = _rlre.match(r'(Rack \d+) (U\d+)', acs.get('comp_rack', ''))
                if _acr: act_comp_rack, act_comp_elev = _acr.group(1), _acr.group(2)
                else:    act_comp_rack = acs.get('comp_rack', '')

        rd = {
            'host': dev_a, 'iface': t0_iface, 't0_lbl': t0_lbl, 'is_phys': is_phys,
            'rack': t0_rack, 'elev': t0_elev,
            # v2 substitution: prefer cutsheet's compute host when it differs from report
            'exp_host': cs.get('comp_host') or exp_b,
            'comp_port': comp_port,
            'comp_rack': comp_rack, 'comp_elev': comp_elev,
            'cur_b': cur_b,
            'cur_port': cur_port,
            'cur_loc': cur_loc,
            'is_mispatch': is_mispatch,
            'comp_is_phys': comp_is_phys,
            # Physical path: Patch Panel Matrix (authoritative) with cutsheet fallback
            'nic_pos':  cs.get('nic_pos', ''),
            'ohr':      ohr,
            'fdf':      fdf,
            't0_pp':    t0_pp,
            # Actual (wrongly-plugged) cable's resolved path — populated for mispatches only
            'act_ohr':      act_ohr,
            'act_fdf':      act_fdf,
            'act_t0_pp':    act_t0pp,
            'act_t0_host':  act_t0_host,
            'act_t0_iface': act_t0_iface,
            'act_t0_rack':  act_t0_rack,
            'act_t0_elev':  act_t0_elev,
            'act_lr':       act_lr,
            'act_nic':      act_nic,
            'act_comp_rack':act_comp_rack,
            'act_comp_elev':act_comp_elev,
        }
        if is_mispatch:
            mispatch_rows.append(rd)
        elif cur_is_placeholder or exp_b in ghost_hosts:
            # Placeholder current (instance.../serial/etc) — host hasn't been named yet
            ghost_rows.append(rd)
        else:
            real_rows.append(rd)
    return real_rows, ghost_rows, mispatch_rows, ghost_hosts


def build_compute_sheet(wb_out, rows, tab_name="Downlinks", tab_colour="70AD47",
                        prev_miss=None, prev_down=None):
    prev_miss = prev_miss or set(); prev_down = prev_down or set()
    if not rows: return
    ws = wb_out.create_sheet(tab_name)
    ws.sheet_properties.tabColor = tab_colour
    headers = [
        ("Interface",     HDR_BG),
        ("L&R",           HDR_BG),
        ("T0 Rack",       HDR_BG),
        ("Elevation",     HDR_BG),
        ("T0 PP",         "17375E"),
        ("FDF",           "375623"),
        ("OHR",           "7F6000"),
        ("Compute Host",  "375623"),
        ("Compute Port",  "375623"),
        ("NIC Position",  "833C00"),
        ("Compute Rack",  "375623"),
        ("Compute U",     "375623"),
        ("Current Device","9C0006"),
        ("History",       "595959"),
    ]
    widths = [12, 6, 12, 6, 30, 28, 28, 40, 16, 14, 12, 8, 40, 22]
    write_header_row(ws, headers, widths)

    # Sort by compute host, then port group pair, then port for visual grouping
    rows = sorted(rows, key=lambda r: (
        r.get('exp_host',''),
        compute_port_group(r.get('exp_host',''), r.get('comp_port','')),
        r.get('comp_port','')
    ))

    for r_idx, rd in enumerate(rows, start=2):
        ws.row_dimensions[r_idx].height = 15
        p = rd['is_phys']; bg="FFFFFF"; lr_bg=LR_BG if p else LR_LOG
        hist_flag, hist_col = get_history_flag(
            rd['host'], rd['iface'], 'mismatch', prev_miss, prev_down, set())
        hist_bg = hist_col if hist_flag else bg
        comp_p = rd.get('comp_is_phys', True)
        comp_port_bg = "E2F0D9" if comp_p else "D5F5E3"
        vals = [rd['iface'], rd['t0_lbl'], rd['rack'], rd['elev'],
                rd.get('t0_pp',''), rd.get('fdf',''), rd.get('ohr',''),
                rd['exp_host'], rd['comp_port'], rd.get('nic_pos',''), rd['comp_rack'],
                rd.get('comp_elev',''), rd['cur_b'], hist_flag]
        all_bgs = [bg,lr_bg,bg,bg,
                   "D9EAF7","E2F0D9","FFF2CC",
                   "E2F0D9",comp_port_bg,"FDDCB5","E2F0D9","E2F0D9",
                   ACT_BG, hist_bg]
        for col,(val,cell_bg) in enumerate(zip(vals,all_bgs),start=1):
            c=ws.cell(r_idx,col); c.value=val; c.fill=fill(cell_bg)
            fg=WHITE if (col==len(vals) and hist_flag) else "000000"
            c.font=Font(bold=(col==2 or (col==len(vals) and hist_flag)),color=fg,name="Arial",size=9)
            c.alignment=center()
    draw_compute_borders(ws, host_col=8, port_col=9)


def build_compute_mispatch_sheet(wb_out, rows, tab_name="Mispatches", tab_colour="C00000",
                                 prev_miss=None, prev_down=None):
    """Wrong-cable rows: a real compute host is connected but not the expected one.
    Shows the T0 + expected compute side, then what is ACTUALLY plugged in."""
    prev_miss = prev_miss or set(); prev_down = prev_down or set()
    if not rows: return
    ws = wb_out.create_sheet(tab_name)
    ws.sheet_properties.tabColor = tab_colour
    headers = [
        ("Interface",        HDR_BG),
        ("L&R",              HDR_BG),
        ("T0 Rack",          HDR_BG),
        ("Elevation",        HDR_BG),
        ("Expected T0 PP",   "17375E"),
        ("Expected FDF",     "375623"),
        ("Expected OHR",     "7F6000"),
        ("NIC Position",     "833C00"),
        ("Expected Host",    "375623"),
        ("Expected Port",    "375623"),
        ("Expected Rack",    "375623"),
        ("Expected U",       "375623"),
        ("Actual Interface", "9C0006"),
        ("Actual L&R",       "9C0006"),
        ("Actual T0 Rack",   "9C0006"),
        ("Actual T0 U",      "9C0006"),
        ("Actual T0 PP",     "843C0C"),
        ("Actual FDF",       "843C0C"),
        ("Actual OHR",       "9C5700"),
        ("Actual NIC",       "9C5700"),
        ("Actual Rack",      "9C0006"),
        ("Actual U",         "9C0006"),
        ("History",          "595959"),
    ]
    widths = [12, 6, 12, 9, 30, 28, 28, 14, 40, 16, 12, 9, 12, 8, 12, 9, 30, 28, 28, 14, 12, 9, 22]
    write_header_row(ws, headers, widths)

    rows = sorted(rows, key=lambda r: (
        r.get('exp_host',''),
        compute_port_group(r.get('exp_host',''), r.get('comp_port','')),
        r.get('comp_port','')
    ))

    for r_idx, rd in enumerate(rows, start=2):
        ws.row_dimensions[r_idx].height = 15
        p = rd['is_phys']; bg="FFFFFF"; lr_bg=LR_BG if p else LR_LOG
        hist_flag, hist_col = get_history_flag(
            rd['host'], rd['iface'], 'mismatch', prev_miss, prev_down, set())
        hist_bg = hist_col if hist_flag else bg
        comp_p = rd.get('comp_is_phys', True)
        comp_port_bg = "E2F0D9" if comp_p else "D5F5E3"
        # actual cable physical flag (odd lane = physical) for L&R shading
        import re as _amre
        _adm = _amre.match(r'slot\d+/port\d+-(\d+)', rd.get('cur_port',''))
        act_phys = (int(_adm.group(1)) % 2 == 1) if _adm else True
        act_lr_bg = LR_BG if act_phys else LR_LOG
        ACT2="FCE4D6"  # light orange for actual-path cells
        vals = [rd['iface'], rd['t0_lbl'], rd['rack'], rd['elev'],
                rd.get('t0_pp',''), rd.get('fdf',''), rd.get('ohr',''),
                rd.get('nic_pos',''),
                rd['exp_host'], rd['comp_port'],
                rd['comp_rack'], rd.get('comp_elev',''),
                rd.get('act_t0_iface',''), rd.get('act_lr',''),
                rd.get('act_t0_rack',''), rd.get('act_t0_elev',''),
                rd.get('act_t0_pp',''), rd.get('act_fdf',''), rd.get('act_ohr',''),
                rd.get('act_nic',''),
                rd.get('act_comp_rack',''), rd.get('act_comp_elev',''),
                hist_flag]
        all_bgs = [bg,lr_bg,bg,bg,
                   "D9EAF7","E2F0D9","FFF2CC",
                   "FDDCB5",
                   "E2F0D9",comp_port_bg,"E2F0D9","E2F0D9",
                   ACT_BG, act_lr_bg, ACT2, ACT2,
                   ACT2, ACT2, ACT2,
                   "FDDCB5", ACT2, ACT2,
                   hist_bg]
        n=len(vals)
        for col,(val,cell_bg) in enumerate(zip(vals,all_bgs),start=1):
            c=ws.cell(r_idx,col); c.value=val; c.fill=fill(cell_bg)
            fg=WHITE if (col==n and hist_flag) else "000000"
            c.font=Font(bold=(col in (2,14) or col in (9,10) or (col==n and hist_flag)),
                        color=fg,name="Arial",size=9)
            c.alignment=center()
    draw_compute_borders(ws, host_col=9, port_col=10)


def build_compute_optics_sheet(wb_out, ws_src, ghost_hosts, t0, t1, compute_lookup=None):
    compute_lookup = compute_lookup or {}
    """Build Compute Optics tab — optics errors on T0,Host links, excluding ghost hosts."""
    if not ws_src: return []
    # Optics cols: 1=Remote Device Name, 2=Remote Device Port, 3=Source Device Name,
    #              4=Source Device Port, 5=Source Device Location, 6=Rx Power, 7=PP Matrix
    rows = []
    import re
    # Detect column layout from headers
    _ohdr = {str(ws_src.cell(1,c).value or '').strip(): c for c in range(1, ws_src.max_column+1)}
    _col_remote_dev  = _ohdr.get('Remote Device Name', 1)
    _col_remote_port = _ohdr.get('Remote Device Port', 2)
    _col_src_dev     = _ohdr.get('Source Device Name', 3)
    _col_src_loc     = _ohdr.get('Source Device Location', _ohdr.get('Source Location', 5))
    _col_src_port    = _ohdr.get('Source Device Port', _ohdr.get('Source Port', 4))
    _col_rx          = _ohdr.get('Rx Power', 6)
    for row in range(2, ws_src.max_row + 1):
        remote_dev  = str(ws_src.cell(row, _col_remote_dev).value  or '').strip()
        remote_port = str(ws_src.cell(row, _col_remote_port).value or '').strip()
        src_dev     = str(ws_src.cell(row, _col_src_dev).value     or '').strip()
        src_port    = str(ws_src.cell(row, _col_src_port).value    or '').strip()
        src_loc     = str(ws_src.cell(row, _col_src_loc).value     or '').strip()
        rx_power    = str(ws_src.cell(row, _col_rx).value          or '').strip()

        if not src_dev or not src_port: continue
        if not is_compute(remote_dev): continue        # T0,T1 optics handled elsewhere
        if remote_dev in ghost_hosts: continue         # skip ghost hosts

        # Parse T0 rack/elev from Source Device Location
        t0_rack, t0_elev = _parse_location(src_loc)

        t0_lbl, _, is_phys = get_t0_labels(src_dev, src_port, t0, t1)


        # Compute-side lookup first (most reliable for compute rack/U)
        import re as _opr
        comp_cs = cs_lookup(compute_lookup, remote_dev, remote_port)
        comp_rack_full = comp_cs.get('comp_rack', '')
        _cr = _opr.match(r'Rack (\d+) U(\d+)', comp_rack_full)
        comp_rack_num = f"Rack {_cr.group(1)}" if _cr else ''
        comp_u        = f"U{_cr.group(2)}"      if _cr else ''

        # T0-side lookup for lr/t0_rack — only trust direct hits, not partner fallback.
        # Partner fallback can return a row for the same MPO pair (same compute port group)
        # whose T0 host/iface point at a DIFFERENT switch in the shuffle. That row's
        # t0_rack/lr would be wrong for the row we're enriching.
        cs = compute_lookup.get((src_dev, src_port), {})
        cs_t0_match_direct = bool(cs)  # exact T0-side hit
        if not cs:
            _ml = _opr.match(r'(swp\d+s)(\d+)', src_port)
            if _ml:
                partner = {'0':'1','1':'0','2':'3','3':'2'}.get(_ml.group(2))
                if partner:
                    cs_partner = compute_lookup.get((src_dev, f"{_ml.group(1)}{partner}"), {})
                    if cs_partner:
                        # T0 partner lane on same physical port — same MPO head, same T0 rack/U.
                        # Safe to copy.
                        cs = cs_partner
                        cs_t0_match_direct = True
        # Compute-side comp_cs is NOT safe for T0 fields (compute-port partner ≠ same T0)
        if cs_t0_match_direct and cs:
            t0_lbl  = cs.get('lr', t0_lbl)
            t0_rack = cs.get('t0_rack', t0_rack)
        # Split combined Rack+U if cutsheet stored them together
        _ru2 = _opr.match(r'(Rack \d+) (U\d+)', t0_rack)
        if _ru2: t0_rack, t0_elev = _ru2.group(1), _ru2.group(2)

        is_flat    = '-40' in rx_power
        is_missing = 'missing' in rx_power.lower() and '-40' not in rx_power
        if is_missing: continue  # skip — transceiver not responding, not actionable

        # Compute port physical/logical: odd lane = physical
        _cm = _opr.match(r'slot\d+/port\d+-(\d+)', remote_port)
        comp_is_phys = (int(_cm.group(1)) % 2 == 1) if _cm else True

        # Normalise t0_rack: split 'Rack 9114 U5' -> 'Rack 9114', 'U5'
        import re as _nrx
        _rux = _nrx.match(r'(Rack \d+) (U\d+)', t0_rack)
        if _rux: t0_rack, t0_elev = _rux.group(1), _rux.group(2)

        rows.append({
            'src_dev': src_dev, 'src_port': src_port,
            't0_lbl': t0_lbl, 'is_phys': is_phys,
            'rack': t0_rack, 'elev': t0_elev,
            'remote_dev': remote_dev, 'remote_port': remote_port,
            'rx_power': rx_power, 'is_flat': is_flat, 'is_missing': is_missing,
            'nic_pos':    comp_cs.get('nic_pos', '') or cs.get('nic_pos', ''),
            'ohr':        comp_cs.get('ohr',     '') or cs.get('ohr',     ''),
            'fdf':        comp_cs.get('fdf',     '') or cs.get('fdf',     ''),
            't0_pp':      comp_cs.get('t0_pp',   '') or cs.get('t0_pp',   ''),
            'comp_rack':  comp_rack_num,
            'comp_u':     comp_u,
            'comp_is_phys': comp_is_phys,
        })

    if not rows: return []
    ws = wb_out.create_sheet("Compute Optics")
    ws.sheet_properties.tabColor = "7030A0"

    headers = [
        ("Interface",       HDR_BG),
        ("L&R",             HDR_BG),
        ("T0 Rack",         HDR_BG),
        ("Elevation",       HDR_BG),
        ("OHR",             "7F6000"),
        ("FDF",             "375623"),
        ("T0 PP",           "17375E"),
        ("Rx Power",        "7030A0"),
        ("Compute Host",    "375623"),
        ("Compute Port",    "375623"),
        ("NIC Position",    "833C00"),
        ("Compute Rack",    "375623"),
        ("Compute U",       "375623"),
        ("Flag",            "595959"),
    ]
    widths = [12, 6, 12, 6, 28, 28, 30, 30, 40, 16, 14, 12, 8, 26]
    write_header_row(ws, headers, widths)

    for r_idx, rd in enumerate(rows, start=2):
        ws.row_dimensions[r_idx].height = 15
        p = rd['is_phys']; bg = "FFFFFF"; lr_bg = LR_BG if p else LR_LOG
        if rd.get('is_missing'):    flag_txt = "⚠️ Transceiver not responding"
        elif rd.get('is_flat'):      flag_txt = "⬇️ Likely downlink (-40dBm)"
        else:                        flag_txt = ""
        flag_bg = "FFE0B2" if rd.get('is_missing') else ("C8C8C8" if rd.get('is_flat') else "FFFFFF")
        rx_bg   = "FFE0B2" if rd.get('is_missing') else ("C8C8C8" if rd.get('is_flat') else "EAD1F8")
        comp_p  = rd.get('comp_is_phys', True)
        comp_port_bg = "E2F0D9" if comp_p else "D5F5E3"
        vals = [rd['src_port'], rd['t0_lbl'], rd['rack'], rd['elev'],
                rd.get('ohr',''), rd.get('fdf',''), rd.get('t0_pp',''),
                rd['rx_power'], rd['remote_dev'], rd['remote_port'], rd.get('nic_pos',''),
                rd.get('comp_rack',''), rd.get('comp_u',''), flag_txt]
        bgs  = [bg, lr_bg, bg, bg,
                "FFF2CC","E2F0D9","D9EAF7",
                rx_bg, "E2F0D9", comp_port_bg, "FDDCB5",
                "E2F0D9","E2F0D9", flag_bg]
        for col, (val, cell_bg) in enumerate(zip(vals, bgs), start=1):
            c = ws.cell(r_idx, col); c.value = val; c.fill = fill(cell_bg)
            txt_fg = "888888" if (rd.get('is_flat') or rd.get('is_missing')) else "000000"
            bold = col==2 or (col==len(vals) and bool(flag_txt))
            c.font = Font(bold=bold, color=txt_fg, name="Arial", size=9)
            c.alignment = center()
    print(f"  Compute Optics — {len(rows)} rows ({len(ghost_hosts)} ghost host(s) excluded)")
    draw_compute_borders(ws, host_col=9, port_col=10)
    return rows


def build_compute_summary(wb_out, compute_real, compute_ghost, ghost_hosts,
                          opt_rows, report_name, fec_rows=None, mispatch_rows=None):
    fec_rows = fec_rows or []
    mispatch_rows = mispatch_rows or []
    """Summary tab listing totals and all ghost compute trays."""
    from datetime import datetime
    ws = wb_out.create_sheet("Summary", 0)
    ws.sheet_properties.tabColor = "1F4E79"
    NAVY="1F4E79"; WHITE="FFFFFF"; RED="C00000"; GREEN="1E8449"
    AMBER="B7770D"; GREY="595959"; LGRY="F2F2F2"

    def fill(h):    return PatternFill("solid", fgColor=h)
    def center():   return Alignment(horizontal="center", vertical="center")
    def left():     return Alignment(horizontal="left",   vertical="center")

    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 16
    for ltr, w in zip("CDEFGHI", [14,14,14,14,14,14,14]):
        ws.column_dimensions[ltr].width = w

    # Title
    ws.merge_cells("B1:H1"); c=ws["B1"]
    c.value = "COMPUTE LINK VALIDATION — SUMMARY"
    c.fill=fill(NAVY); c.font=Font(bold=True,color=WHITE,name="Arial",size=13)
    c.alignment=center(); ws.row_dimensions[1].height=28

    ws.merge_cells("B2:H2"); c=ws["B2"]
    c.value = f"Report: {report_name}   |   Generated: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    c.fill=fill("0D7377"); c.font=Font(italic=True,color=WHITE,name="Arial",size=9)
    c.alignment=center(); ws.row_dimensions[2].height=16

    # KPI row
    ws.row_dimensions[3].height=8
    flat_count = sum(1 for r in opt_rows if r.get('is_flat'))
    total_issues = len(compute_real) + len(mispatch_rows) + len(opt_rows) + len(fec_rows)
    kpi = [
        ("TOTAL ISSUES",     total_issues,        NAVY),
        ("HOST LINK ERRORS", len(compute_real),   "375623"),
        ("MISPATCHES",       len(mispatch_rows),  RED),
        ("GHOST LINKS",      len(compute_ghost),   GREY),
        ("COMPUTE OPTICS",   len(opt_rows),         "7030A0"),
        ("FEC ERRORS",       len(fec_rows),         "0070C0"),
        ("⬇️ -40dBm FLAGS",  flat_count,            "808080"),
    ]
    for i, (lbl, val, bg) in enumerate(kpi):
        col = i+2
        ws.row_dimensions[4].height=16; ws.row_dimensions[5].height=30
        c=ws.cell(4,col); c.value=lbl; c.fill=fill(bg)
        c.font=Font(bold=True,color=WHITE,name="Arial",size=8); c.alignment=center()
        c=ws.cell(5,col); c.value=val; c.fill=fill(bg)
        c.font=Font(bold=True,color=WHITE,name="Arial",size=20); c.alignment=center()

    # Ghost trays section
    ws.row_dimensions[6].height=10
    ws.merge_cells("B7:H7"); c=ws["B7"]
    c.value = f"GHOST COMPUTE TRAYS  ({GHOST_THRESHOLD}+ LLDP errors — not yet online)"
    c.fill=fill(GREY); c.font=Font(bold=True,color=WHITE,name="Arial",size=10)
    c.alignment=center(); ws.row_dimensions[7].height=20

    sub_hdrs = ["Compute Host", "Error Count", "Rack / U", "Status"]
    sub_bgs  = [NAVY, NAVY, NAVY, NAVY]
    for i,(h,bg) in enumerate(zip(sub_hdrs,sub_bgs)):
        c=ws.cell(8,i+2); c.value=h; c.fill=fill(bg)
        c.font=Font(bold=True,color=WHITE,name="Arial",size=9); c.alignment=center()
    ws.row_dimensions[8].height=16

    # Count per ghost host
    from collections import Counter
    ghost_counts = Counter(rd['exp_host'] for rd in compute_ghost)
    # Build host -> rack lookup from ghost rows
    ghost_rack_map = {}
    for rd in compute_ghost:
        h = rd.get('exp_host','')
        r = rd.get('comp_rack','')
        if h and r: ghost_rack_map[h] = r

    for row_i, host in enumerate(sorted(ghost_hosts)):
        row = 9 + row_i; ws.row_dimensions[row].height=18
        rack_u = ghost_rack_map.get(host, '')
        for col, (val, bg) in enumerate([
            (host,                     LGRY),
            (ghost_counts.get(host,0), LGRY),
            (rack_u,                   LGRY),
            ("Ghost — not online",     "FFE0E0"),
        ], start=2):
            c=ws.cell(row,col); c.value=val; c.fill=fill(bg)
            c.font=Font(name="Arial",size=10,color="000000")
            c.alignment=center() if col>2 else left()

    next_row = 9 + len(ghost_hosts) + 1
    if not ghost_hosts:
        ws.merge_cells(f"B9:H9"); c=ws["B9"]
        c.value="No ghost hosts detected in this report"
        c.fill=fill("E2F0D9"); c.font=Font(italic=True,name="Arial",size=10,color="1E8449")
        c.alignment=center(); ws.row_dimensions[9].height=18

    # Host link errors summary
    ws.row_dimensions[next_row].height=10
    ws.merge_cells(f"B{next_row+1}:H{next_row+1}"); c=ws.cell(next_row+1,2)
    c.value="REAL HOST LINK ERRORS BY COMPUTE HOST"
    c.fill=fill("375623"); c.font=Font(bold=True,color=WHITE,name="Arial",size=10)
    c.alignment=center(); ws.row_dimensions[next_row+1].height=20

    for i,(h,bg) in enumerate(zip(["Compute Host","Error Count","Rack / U"],[NAVY,NAVY,NAVY])):
        c=ws.cell(next_row+2,i+2); c.value=h; c.fill=fill(bg)
        c.font=Font(bold=True,color=WHITE,name="Arial",size=9); c.alignment=center()
    ws.row_dimensions[next_row+2].height=16

    real_counts = Counter(rd['exp_host'] for rd in compute_real)
    # Build host -> rack from real rows
    real_rack_map = {}
    for rd in compute_real:
        h = rd.get('exp_host','')
        r = rd.get('comp_rack','')
        if h and r: real_rack_map[h] = r
    for ri, (host, cnt) in enumerate(sorted(real_counts.items(), key=lambda x:-x[1])):
        row = next_row+3+ri; ws.row_dimensions[row].height=18
        rack_u = real_rack_map.get(host, '')
        for col, (val,bg) in enumerate([(host,LGRY),(cnt,LGRY),(rack_u,LGRY)], start=2):
            c=ws.cell(row,col); c.value=val; c.fill=fill(bg)
            c.font=Font(name="Arial",size=10); c.alignment=center() if col>2 else left()

    ws.freeze_panes="B2"
    print(f"  Summary tab built — {len(ghost_hosts)} ghost host(s), {len(compute_real)} real errors")


def build_compute_fec_sheet(wb_out, ws_src, ghost_hosts, compute_lookup=None):
    """FEC errors with full T0 + compute patch panel info matching Downlinks layout."""
    import re as _fre
    compute_lookup = compute_lookup or {}
    if not ws_src: return []
    rows = []
    for row in range(2, ws_src.max_row + 1):
        dev   = str(ws_src.cell(row, 1).value or '').strip()
        iface = str(ws_src.cell(row, 2).value or '').strip()
        lanes = str(ws_src.cell(row, 3).value or '').strip()
        issue = str(ws_src.cell(row, 4).value or '').strip()
        if not dev or not issue: continue
        if not is_compute(dev): continue
        if dev in ghost_hosts: continue
        ber_vals = [float(m) for m in _fre.findall(r'raw-ber=([\d.e+-]+)', issue)]
        if not ber_vals or max(ber_vals) < 1e-07: continue
        max_ber = max(ber_vals)
        if max_ber >= 1e-04:   severity = "\U0001f534 Critical"
        elif max_ber >= 1e-05: severity = "\U0001f7e0 High"
        elif max_ber >= 1e-06: severity = "\U0001f7e1 Elevated"
        else:                   severity = "\U0001f7e2 Marginal"
        # Reverse lookup by compute port
        cs = cs_lookup(compute_lookup, dev, iface)
        _cr = _fre.match(r'Rack (\d+) U(\d+)', cs.get('comp_rack',''))
        comp_rack = f"Rack {_cr.group(1)}" if _cr else ''
        comp_u    = f"U{_cr.group(2)}"     if _cr else ''
        _t0r_raw = cs.get('t0_rack', '')
        _t0r_m = _fre.match(r'(Rack \d+)(?: (U\d+))?', _t0r_raw)
        t0_rack_s = _t0r_m.group(1)           if _t0r_m else ''
        t0_elev_s = _t0r_m.group(2) or ''  if _t0r_m else ''
        rows.append({
            'dev': dev, 'iface': iface, 'lanes': lanes, 'issue': issue,
            'max_ber': max_ber, 'severity': severity,
            't0_iface': cs.get('t0_iface',''), 'lr': cs.get('lr',''),
            't0_rack':  t0_rack_s, 't0_elev': t0_elev_s,
            'ohr':      cs.get('ohr',''),  'fdf': cs.get('fdf',''),
            't0_pp':    cs.get('t0_pp',''), 'nic_pos': cs.get('nic_pos',''),
            'comp_rack': comp_rack, 'comp_u': comp_u,
        })
    if not rows: return []
    ws = wb_out.create_sheet("FEC Errors")
    ws.sheet_properties.tabColor = "0070C0"
    headers = [
        ("T0 Interface",  HDR_BG),
        ("L&R",           HDR_BG),
        ("T0 Rack",       HDR_BG),
        ("T0 Elev",       HDR_BG),
        ("OHR",           "7F6000"),
        ("FDF",           "375623"),
        ("T0 PP",         "17375E"),
        ("Compute Host",  "17375E"),
        ("Compute Port",  "17375E"),
        ("NIC Position",  "833C00"),
        ("Compute Rack",  "17375E"),
        ("Compute U",     "17375E"),
        ("Severity",      HDR_BG),
        ("Max BER",       HDR_BG),
        ("Issue Detail",  HDR_BG),
    ]
    widths = [14, 6, 14, 6, 28, 28, 30, 40, 18, 14, 12, 8, 14, 10, 60]
    write_header_row(ws, headers, widths)

    rows = sorted(rows, key=lambda r: (
        r.get('dev',''),
        compute_port_group(r.get('dev',''), r.get('iface','')),
        r.get('iface','')
    ))

    for r_idx, rd in enumerate(rows, start=2):
        ws.row_dimensions[r_idx].height = 15
        sev_bg = {"\U0001f534 Critical":"FFCCCC","\U0001f7e0 High":"FFE0CC",
                  "\U0001f7e1 Elevated":"FFF2CC","\U0001f7e2 Marginal":"E2F0D9"}.get(rd['severity'],"FFFFFF")
        vals = [rd['t0_iface'], rd['lr'], rd['t0_rack'], rd.get('t0_elev',''),
                rd['ohr'], rd['fdf'], rd['t0_pp'],
                rd['dev'], rd['iface'], rd['nic_pos'],
                rd['comp_rack'], rd['comp_u'],
                rd['severity'], f"{rd['max_ber']:.2e}", rd['issue']]
        bgs = ["FFFFFF","D9EAF7","FFFFFF","FFFFFF",
               "FFF2CC","E2F0D9","D9EAF7",
               "E2F0D9","E2F0D9","FDDCB5",
               "E2F0D9","E2F0D9",
               sev_bg,"FFFFFF","FFFFFF"]
        for col,(val,bg) in enumerate(zip(vals,bgs),start=1):
            c = ws.cell(r_idx,col); c.value=val; c.fill=fill(bg)
            c.font = Font(name="Arial",size=9,color="000000")
            c.alignment = Alignment(
                horizontal="left" if col in (4,5,6,14) else "center",
                vertical="center")
    draw_compute_borders(ws, host_col=8, port_col=9)
    print(f"  FEC Errors — {len(rows)} rows ({ws_src.max_row-1-len(rows)} below threshold skipped)")
    return rows

def main():
    cfg = load_config()

    print("=" * 60)
    print("  LV Portal Validation Formatter")
    print("=" * 60)

    # Step 1: cutsheets
    saved_paths = cfg.get('cutsheet_paths', [])
    if not saved_paths and cfg.get('cutsheet_path'):
        saved_paths = [cfg['cutsheet_path']]
    saved_paths = [p for p in saved_paths if os.path.isfile(p)]

    # Step 1: GPU/compute cutsheet
    compute_lookup = {}
    t0 = {}; t1 = {}; t1_rev = {}; t0_to_pp = {}
    gpu_paths = pick_multiple_files("Select GPU/Compute Cutsheet(s) — Ctrl/Cmd for multiple")
    if not gpu_paths:
        show_msg("Cancelled", "No cutsheet selected.", error=True); sys.exit(0)
    compute_lookup = build_compute_lookup(gpu_paths)
    cfg['gpu_paths'] = gpu_paths
    save_config(cfg)

    # Step 2: Portal report — single pick, no other prompts
    prev_report_path = None
    report_path = pick_file("Select LV Portal Validation Export (.xlsx)")
    if not report_path:
        show_msg("Cancelled", "No file selected.", error=True); sys.exit(0)

    prev_miss = set(); prev_down = set(); prev_opt = set(); prev_rack_map = {}
    if prev_report_path:
        print("Loading previous report...")
        prev_miss, prev_down, prev_opt, prev_rack_map = get_prev_issues_lv(prev_report_path)

    print(f"Processing: {os.path.basename(report_path)}")
    wb_src = load_workbook(report_path)

    # Find sheets
    def find_sheet(wb, *patterns):
        for name in wb.sheetnames:
            for p in patterns:
                if p.lower() in name.lower(): return wb[name]
        return None

    ws_lldp     = find_sheet(wb_src, 'lldp mismatch', 'lldp', 'mismatch')
    ws_iface_down = find_sheet(wb_src, 'interface down', 'interface_down')
    ws_optics = find_sheet(wb_src, 'optic errors', 'optic')

    if not ws_lldp:
        show_msg("Error", "Could not find LLDP sheet.", error=True); sys.exit(0)

    # Process LLDP
    print("Processing LLDP...")
    miss_rows, down_rows = process_lldp(ws_lldp, t0, t1, t1_rev, t0_to_pp)
    compute_real, compute_ghost, compute_mispatch, ghost_hosts = process_compute_lldp(ws_lldp, compute_lookup)
    # Merge Interface Down tab if present (new format separates downlinks from LLDP mismatches)
    if ws_iface_down:
        id_real, id_ghost, id_ghosts = process_interface_down(ws_iface_down, compute_lookup)
        compute_real  = compute_real  + id_real
        compute_ghost = compute_ghost + id_ghost
        ghost_hosts   = ghost_hosts | id_ghosts
        if id_real or id_ghost:
            print(f"  Interface Down: {len(id_real)} real, {len(id_ghost)} ghost rows merged")
    print(f"  Mismatches: {len(miss_rows)} | Downlinks: {len(down_rows)} | Host Links: {len(compute_real)} | Mispatches: {len(compute_mispatch)} | Ghost: {len(compute_ghost)}")

    # Build downlink set for optics/FEC cross-reference
    downlink_set = set()
    for rd in down_rows:
        downlink_set.add((rd['host'], rd['iface']))

    # Build output
    wb_out = Workbook()
    wb_out.remove(wb_out.active)

    if miss_rows: build_mispatches_sheet(wb_out, miss_rows, prev_miss=prev_miss, prev_down=prev_down)
    if down_rows: build_downlinks_sheet(wb_out, down_rows, prev_miss=prev_miss, prev_down=prev_down, prev_opt=prev_opt)
    ws_comp_fec = find_sheet(wb_src, 'fec_ber', 'fec')
    fec_rows = build_compute_fec_sheet(wb_out, ws_comp_fec, ghost_hosts, compute_lookup)
    ws_comp_opt = find_sheet(wb_src, 'optic errors', 'optic')
    # Build optics first so the summary KPI uses the exact same row set the tab shows
    # (the tab excludes 'transceiver not responding' rows, so a separate count drifts).
    opt_rows = build_compute_optics_sheet(wb_out, ws_comp_opt, ghost_hosts, t0, t1, compute_lookup)
    build_compute_summary(wb_out, compute_real, compute_ghost, ghost_hosts,
                          opt_rows, os.path.basename(report_path),
                          fec_rows=fec_rows, mispatch_rows=compute_mispatch)
    if compute_real:  build_compute_sheet(wb_out, compute_real,  "Downlinks",  "70AD47", prev_miss=prev_miss, prev_down=prev_down)
    if compute_mispatch: build_compute_mispatch_sheet(wb_out, compute_mispatch, "Mispatches", "C00000", prev_miss=prev_miss, prev_down=prev_down)
    if compute_ghost: build_compute_sheet(wb_out, compute_ghost, "Ghost Links", "808080", prev_miss=prev_miss, prev_down=prev_down)
    # Optics tab already built above; move it after the compute tabs for ordering
    if 'Compute Optics' in wb_out.sheetnames:
        wb_out.move_sheet('Compute Optics', offset=len(wb_out.sheetnames))
    if 'FEC Errors' in wb_out.sheetnames:
        wb_out.move_sheet('FEC Errors', offset=len(wb_out.sheetnames))
    if ws_optics: build_optics_sheet(wb_out, ws_optics, t0, t1, t1_rev, downlink_set, t0_to_pp, prev_miss=prev_miss, prev_down=prev_down, prev_opt=prev_opt)

    # Save
    base, ext = os.path.splitext(report_path)
    out_path = base + "_formatted" + ext
    wb_out.save(out_path)
    print(f"\nSaved: {out_path}")
    show_msg("Done ✅", f"Saved to:\n{out_path}")

if __name__ == "__main__":
    main()
