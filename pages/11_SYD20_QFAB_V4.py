#!/usr/bin/env python3
"""
SYD20 QFAB / Slack Report Formatter — Streamlit V4
Full conversion of SY20_QFAB_SLACK_No_PP.py
- swap_mismatch_groups and sort_mismatch_pairs are verbatim from the original
- Column ordering (Expected before Active in Mismatch tabs) and cluster coloring
  exactly match the original's "suggested mismatch" presentation
- All other original features and quality preserved
"""
import streamlit as st
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io
import tempfile
import os
import re
from collections import Counter, defaultdict
from pathlib import Path

# --- Auth + Central Logging (GitHub repo data folder) ---
from utils.auth import require_login
from utils.data_logger import log_errors
require_login()

# ================== Styles & Constants ==================
YELLOW_FILL = PatternFill("solid", start_color="FFFF00")
PINK_FILL = PatternFill("solid", start_color="FFC0CB")
ORANGE_FILL = PatternFill("solid", start_color="FFA500")
HEADER_FONT = Font(name="Arial", bold=True, color="000000")
BODY_FONT = Font(name="Arial")
GREY_FONT = Font(name="Arial", color="808080")
DOWN_PORT_THRESHOLD = 32
NOTE_COL_WIDTH = 25
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=False)
THIN = Side(style="thin", color="000000")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

PINK_COLS = [
    "Active Host Act. Interface", "Act. Rack", "Act. Elevation",
    "Cutsheet PP_A", "Cutsheet PP_B", "Cutsheet Other End", "Cutsheet Other End Rack",
]
CUT_COLS = ["Cutsheet PP_A", "Cutsheet PP_B", "Cutsheet Other End", "Cutsheet Other End Rack"]

MERGE_PAIRS = [
    ("Hostname", "Interface", "Hostname Interface"),
    ("Expected Hostname", "Exp. Interface", "Expected Hostname Exp. Interface"),
    ("Active Host", "Act. Interface", "Active Host Act. Interface"),
]

def extract_label(filename):
    m = re.search(r"b(\d+)", Path(filename).name, re.IGNORECASE)
    return f"B{m.group(1)}" if m else Path(filename).stem

PP_NOT_FOUND = "PP_info_not_found"
FULL_PATH_SHEET = "full_path_lldp_with_int_down"
RAW_LLDP_SHEET = "lldp_with_int_down"

# ================== Core Logic ==================
def load_combined(input_files, lookup=None):
    TARGETS = [FULL_PATH_SHEET, "optics", "fec_ber"]
    combined = {s: {"headers": None, "rows": []} for s in TARGETS}
    for uploaded_file in input_files:
        tag = extract_label(uploaded_file.name)
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
            tmp.write(uploaded_file.getvalue())
            tmp_path = tmp.name
        try:
            wb = load_workbook(tmp_path, data_only=True)
            for s in TARGETS:
                if s in wb.sheetnames:
                    src_sheet, raw = s, False
                elif s == FULL_PATH_SHEET and RAW_LLDP_SHEET in wb.sheetnames:
                    src_sheet, raw = RAW_LLDP_SHEET, True
                else:
                    continue
                ws = wb[src_sheet]
                headers = [c.value for c in ws[1]]
                pp_idx = None
                if raw:
                    hi = headers.index("Hostname") if "Hostname" in headers else None
                    ii = headers.index("Interface") if "Interface" in headers else None
                    if "PP_A" not in headers and "PP_B" not in headers:
                        at = (headers.index("Elevation") + 1) if "Elevation" in headers else len(headers)
                        headers = headers[:at] + ["PP_A", "PP_B"] + headers[at:]
                    else:
                        at = headers.index("PP_A")
                    pp_idx = (hi, ii, at)
                if combined[s]["headers"] is None:
                    combined[s]["headers"] = ["Source"] + list(headers)
                for r in ws.iter_rows(min_row=2, values_only=True):
                    if all(v is None or v == "" for v in r):
                        continue
                    row = list(r)
                    if raw and pp_idx:
                        hi, ii, at = pp_idx
                        pp_a = pp_b = PP_NOT_FOUND
                        if lookup is not None and hi is not None and row[hi] is not None:
                            key = f"{row[hi]} {row[ii]}".strip()
                            hit = lookup.get(key)
                            if hit and hit[0] is not None:
                                pp_a, pp_b = hit[0], hit[1]
                        row = row[:at] + [pp_a, pp_b] + row[at:]
                    combined[s]["rows"].append([tag] + row)
        finally:
            os.unlink(tmp_path)
    return combined

def merge_columns(headers, rows):
    for col_a, col_b, new_hdr in MERGE_PAIRS:
        if col_a not in headers or col_b not in headers:
            continue
        ia = headers.index(col_a)
        ib = headers.index(col_b)
        new_headers = [new_hdr if i == ia else h for i, h in enumerate(headers) if i != ib]
        new_rows = []
        for r in rows:
            new_row = []
            for i, v in enumerate(r):
                if i == ib: continue
                if i == ia:
                    va = r[ia] or ''
                    vb = r[ib] or ''
                    new_row.append(f"{va} {vb}".strip())
                else:
                    new_row.append(v)
            new_rows.append(new_row)
        headers, rows = new_headers, new_rows
    return headers, rows

# ================== MISMATCH CLUSTERING (verbatim from original SY20_QFAB_SLACK_No_PP.py) ==================
def sort_mismatch_pairs(headers, rows):
    """Group rows that are connected via Expected<->Active swaps into clusters.
    Cluster 1 → orange, Cluster 2 → yellow, Cluster 3 → orange, ... (alternating).
    Unpaired rows (no match) are moved to the end with no highlight.
    Returns (new_rows, row_colors) where row_colors is a dict {0-based index: fill}.
    """
    from collections import defaultdict

    exp_i = headers.index("Expected Hostname Exp. Interface")
    act_i = headers.index("Active Host Act. Interface")

    # Build adjacency: row i <-> row j if rows[i][exp] == rows[j][act] or vice versa
    act_to_idxs = defaultdict(list)
    for i, r in enumerate(rows):
        v = r[act_i]
        if v:
            act_to_idxs[v].append(i)

    adj = defaultdict(set)
    for i, r in enumerate(rows):
        exp_val = r[exp_i]
        if exp_val and exp_val in act_to_idxs:
            for j in act_to_idxs[exp_val]:
                if j != i:
                    adj[i].add(j)
                    adj[j].add(i)

    # Find connected components
    visited = set()
    groups = []
    for start in range(len(rows)):
        if start in visited or start not in adj:
            continue
        group = []
        stack = [start]
        while stack:
            n = stack.pop()
            if n in visited:
                continue
            visited.add(n)
            group.append(n)
            stack.extend(adj[n] - visited)
        groups.append(sorted(group))

    # Unpaired rows (not in any group)
    grouped_idxs = {i for g in groups for i in g}
    unpaired = [i for i in range(len(rows)) if i not in grouped_idxs]

    # Assign alternating colors: group 1=orange, group 2=yellow, group 3=orange ...
    group_color = []
    for gi, group in enumerate(groups):
        group_color.append(ORANGE_FILL if gi % 2 == 0 else YELLOW_FILL)

    # Build new row order: all grouped rows first (in group order), then unpaired
    new_rows = []
    row_colors = {}
    for gi, group in enumerate(groups):
        fill = group_color[gi]
        for orig_idx in group:
            row_colors[len(new_rows)] = fill
            new_rows.append(rows[orig_idx])

    for orig_idx in unpaired:
        new_rows.append(rows[orig_idx])

    return new_rows, row_colors

# ================== Remaining original functions (unchanged) ==================
def drop_columns(headers, rows, drop_names):
    keep = [i for i, h in enumerate(headers) if h not in drop_names]
    return [headers[i] for i in keep], [[r[i] for i in keep] for r in rows]

def reorder_columns(headers, rows, new_order_names):
    idxs = [headers.index(n) for n in new_order_names]
    extras = [i for i in range(len(headers)) if i not in idxs]
    final = idxs + extras
    return [headers[i] for i in final], [[r[i] for i in final] for r in rows]

def swap_mismatch_groups(headers, rows):
    """Put Expected group before Active group in Mismatch tabs.
    Must be called BEFORE merge_columns (uses the pre-merged column names).
    Matches the original SY20_QFAB_SLACK_No_PP.py logic.
    """
    ACT = ["Active Host", "Act. Interface", "Act. Rack", "Act. Elevation"]
    EXP = ["Expected Hostname", "Exp. Interface", "Exp. Rack", "Exp. Elevation"]
    if not all(h in headers for h in ACT + EXP):
        return headers, rows
    act_idxs = [headers.index(h) for h in ACT]
    exp_idxs = [headers.index(h) for h in EXP]
    pre = [i for i in range(len(headers)) if i not in set(act_idxs + exp_idxs)]
    final = pre + exp_idxs + act_idxs
    return [headers[i] for i in final], [[r[i] for i in final] for r in rows]

def dedup_bidirectional(headers, rows):
    h_i = headers.index("Hostname")
    i_i = headers.index("Interface")
    eh_i = headers.index("Expected Hostname")
    ei_i = headers.index("Exp. Interface")
    src_i = headers.index("Source") if "Source" in headers else None
    seen = {}
    order = []
    for r in rows:
        key = frozenset([(r[h_i], r[i_i]), (r[eh_i], r[ei_i])])
        if key in seen:
            if src_i is not None:
                existing = str(seen[key][src_i])
                parts = [p.strip() for p in existing.split(",")]
                new_src = str(r[src_i])
                if new_src not in parts:
                    seen[key][src_i] = f"{existing},{new_src}"
        else:
            seen[key] = list(r)
            order.append(key)
    return [seen[k] for k in order]

def build_cutsheet_lookup(path_bytes):
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        tmp.write(path_bytes)
        tmp_path = tmp.name
    try:
        wb = load_workbook(tmp_path, data_only=True)
        ws = wb.active
        lookup = {}
        for r in ws.iter_rows(values_only=True):
            if not r or len(r) < 6: continue
            c1, c2, c3, c4, c5, c6 = r[:6]
            if not c5:
                if c1: lookup[str(c1).strip()] = (None, None, c3, c4)
                if c3: lookup[str(c3).strip()] = (None, None, c1, c2)
            else:
                if c1: lookup[str(c1).strip()] = (c3, c4, c5, c6)
                if c5: lookup[str(c5).strip()] = (c3, c4, c1, c2)
        return lookup
    finally:
        os.unlink(tmp_path)

def enrich(headers, rows, key_pairs, lookup):
    new_headers = list(headers) + CUT_COLS
    new_rows = []
    for r in rows:
        hit = None
        for host_col, intf_col in key_pairs:
            if host_col not in headers: continue
            h_idx = headers.index(host_col)
            i_idx = headers.index(intf_col)
            if r[h_idx] is None: continue
            key = f"{r[h_idx]} {r[i_idx]}".strip()
            h = lookup.get(key)
            if h:
                hit = h
                break
        new_rows.append(list(r) + list(hit) if hit else list(r) + [None]*len(CUT_COLS))
    return new_headers, new_rows

def fill_empty_pp(headers, rows):
    if "Cutsheet PP_A" not in headers or "Cutsheet PP_B" not in headers:
        return rows
    pa = headers.index("Cutsheet PP_A")
    pb = headers.index("Cutsheet PP_B")
    for r in rows:
        if r[pa] in (None, ""): r[pa] = "<==>"
        if r[pb] in (None, ""): r[pb] = "<==>"
    return rows

def split_full_path(headers, rows):
    ai = headers.index("Act. Interface")
    down, mis = [], []
    for r in rows:
        (down if r[ai] == "interface down" else mis).append(r)
    return down, mis

def split_by_pp(headers, rows):
    a = headers.index("PP_A")
    t2, t1 = [], []
    for r in rows:
        (t1 if r[a] == "PP_info_not_found" else t2).append(r)
    return t2, t1

def split_by_cutsheet_pp(headers, rows):
    if "Cutsheet PP_A" not in headers:
        return list(rows), []
    pa = headers.index("Cutsheet PP_A")
    t2, t1 = [], []
    for r in rows:
        v = r[pa] if pa < len(r) else None
        if v not in (None, "") and str(v).strip() != "<==>":
            t2.append(r)
        else:
            t1.append(r)
    return t2, t1

def write_sheet(wb, name, headers, rows):
    if name in wb.sheetnames:
        del wb[name]
    ws = wb.create_sheet(name)
    ws.append(headers)
    for r in rows:
        ws.append(r)
    return ws

def style_sheet(ws, pink_col_names=(), freeze_at=None, row_colors=None, col_widths=None):
    if ws.max_row == 0: return
    max_col = ws.max_column
    max_row = ws.max_row
    headers = [c.value for c in ws[1]]
    pink_idxs = {headers.index(n) + 1 for n in pink_col_names if n in headers}
    ws_row_colors = {i + 2: fill for i, fill in (row_colors or {}).items()}
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = YELLOW_FILL
        cell.alignment = CENTER
        cell.border = BORDER
    act_col = headers.index("Active Host Act. Interface") + 1 if "Active Host Act. Interface" in headers else None
    for row in ws.iter_rows(min_row=2, max_row=max_row, max_col=max_col):
        row_fill = ws_row_colors.get(row[0].row)
        for cell in row:
            cell.font = BODY_FONT
            cell.alignment = CENTER
            cell.border = BORDER
            if row_fill and act_col and cell.column < act_col:
                cell.fill = row_fill
            elif cell.column in pink_idxs:
                cell.fill = PINK_FILL
    ws.auto_filter.ref = f"A1:{get_column_letter(max_col)}{max_row}"
    if col_widths:
        for ci, w in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(ci)].width = w
    else:
        for col_cells in ws.iter_cols(min_row=1, max_row=max_row, max_col=max_col):
            col_letter = get_column_letter(col_cells[0].column)
            max_len = max((len(str(c.value)) for c in col_cells if c.value is not None), default=10)
            ws.column_dimensions[col_letter].width = max(max_len + 2, 10)
    ws.freeze_panes = freeze_at or "A2"

def compute_col_widths(headers, rows):
    widths = [len(str(h)) if h is not None else 0 for h in headers]
    for row in rows:
        for i, v in enumerate(row):
            if v is not None:
                widths[i] = max(widths[i], len(str(v)))
    return [min(max(w + 2, 10), 60) for w in widths]

def annotate_downlink_switches(ws, threshold=DOWN_PORT_THRESHOLD):
    headers = [c.value for c in ws[1]]
    if "Note" not in headers: return
    note_col = headers.index("Note") + 1
    target_cols = [h for h in ("Hostname Interface", "Expected Hostname Exp. Interface") if h in headers]
    if not target_cols: return
    col_idxs = [headers.index(h) for h in target_cols]
    def host_of(value):
        return str(value).split(" ", 1)[0] if value else None
    down_hosts = set()
    for ci in col_idxs:
        counts = defaultdict(int)
        for row in ws.iter_rows(min_row=2):
            h = host_of(row[ci].value)
            if h: counts[h] += 1
        down_hosts |= {h for h, n in counts.items() if n == threshold}
    if not down_hosts: return
    for row in ws.iter_rows(min_row=2):
        if any(host_of(row[ci].value) in down_hosts for ci in col_idxs):
            for cell in row:
                cell.font = GREY_FONT
            ws.cell(row=row[0].row, column=note_col).value = "maybe switch off"

def annotate_optics_in_downlinks(wb):
    def hi_values(name):
        if name not in wb.sheetnames: return set()
        ws = wb[name]
        headers = [c.value for c in ws[1]]
        if "Hostname Interface" not in headers: return set()
        i = headers.index("Hostname Interface")
        return {str(r[i]).strip() for r in ws.iter_rows(min_row=2, values_only=True) if r[i]}
    dl = hi_values("T2-T1 Downlink") | hi_values("T1-T0 Downlink")
    if not dl: return
    for opt_tab in ("T2-T1 Optics", "T1-T0 Optics"):
        if opt_tab not in wb.sheetnames: continue
        ws = wb[opt_tab]
        headers = [c.value for c in ws[1]]
        if "Hostname Interface" not in headers or "Note" not in headers: continue
        hi = headers.index("Hostname Interface")
        note_col = headers.index("Note") + 1
        for row in ws.iter_rows(min_row=2):
            v = row[hi].value
            if v and str(v).strip() in dl:
                for cell in row: cell.font = GREY_FONT
                ws.cell(row=row[0].row, column=note_col).value = "also in downlinks"

def widen_note_columns(wb, width=NOTE_COL_WIDTH):
    for name in wb.sheetnames:
        ws = wb[name]
        headers = [c.value for c in ws[1]]
        if "Note" in headers:
            col = get_column_letter(headers.index("Note") + 1)
            ws.column_dimensions[col].width = width

def capitalize_b_numbers(ws):
    B_NUM = re.compile(r"\bb(\d+)\b")
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            if isinstance(cell.value, str):
                new_val, n = B_NUM.subn(lambda m: f"B{m.group(1)}", cell.value)
                if n:
                    cell.value = new_val

def build_summary(wb, labels, cat_to_tabs):
    def count_sources(tabs):
        c = Counter()
        for tn in tabs:
            if tn not in wb.sheetnames: continue
            ws = wb[tn]
            headers = [cell.value for cell in ws[1]]
            if "Source" not in headers: continue
            s_idx = headers.index("Source")
            for r in ws.iter_rows(min_row=2, values_only=True):
                src = r[s_idx]
                if src is None: continue
                for part in str(src).split(","):
                    p = part.strip()
                    if p: c[p] += 1
        return c
    if "summary" in wb.sheetnames:
        del wb["summary"]
    ws = wb.create_sheet("summary", 0)
    ws.append(["Category"] + list(labels) + ["Total"])
    for cat, tabs in cat_to_tabs:
        c = count_sources(tabs)
        row_values = [cat] + [c.get(l, 0) for l in labels]
        total = sum(row_values[1:])
        ws.append(row_values + [total])
    return ws

# ================== Main Processing Pipeline ==================
def process_files(cutsheet_bytes, input_files):
    labels = []
    for f in input_files:
        m = re.search(r"b(\d+)", f.name, re.IGNORECASE)
        labels.append(f"B{m.group(1)}" if m else Path(f.name).stem)
    lookup = build_cutsheet_lookup(cutsheet_bytes)
    combined = load_combined(input_files, lookup)
    fp = combined["full_path_lldp_with_int_down"]
    if fp["headers"] is None:
        raise ValueError("No full_path_lldp_with_int_down sheet found in any input file.")
    fp_headers = fp["headers"]
    fp_rows = fp["rows"]
    dl_rows, mis_rows = split_full_path(fp_headers, fp_rows)
    t2_dl_rows, t1_dl_rows = split_by_pp(fp_headers, dl_rows)
    t2_mis_rows, t1_mis_rows = split_by_pp(fp_headers, mis_rows)
    t2_dl_hdr, t2_dl_rows = drop_columns(fp_headers, t2_dl_rows, ["index", "Building", "Exp. Building", "Active Host", "Act. Interface", "Act. Building", "Act. Rack", "Act. Elevation"])
    t1_dl_hdr, t1_dl_rows = drop_columns(fp_headers, t1_dl_rows, ["index", "Building", "Exp. Building", "PP_A", "PP_B", "Active Host", "Act. Interface", "Act. Building", "Act. Rack", "Act. Elevation"])
    t2_mis_hdr, t2_mis_rows = drop_columns(fp_headers, t2_mis_rows, ["index", "Building", "Act. Building", "Exp. Building"])
    t1_mis_hdr, t1_mis_rows = drop_columns(fp_headers, t1_mis_rows, ["index", "Building", "PP_A", "PP_B", "Act. Building", "Exp. Building"])
    t2_mis_hdr, t2_mis_rows = swap_mismatch_groups(t2_mis_hdr, t2_mis_rows)
    t1_mis_hdr, t1_mis_rows = swap_mismatch_groups(t1_mis_hdr, t1_mis_rows)
    t1_dl_rows = dedup_bidirectional(t1_dl_hdr, t1_dl_rows)
    t1_mis_rows = dedup_bidirectional(t1_mis_hdr, t1_mis_rows)
    op = combined["optics"]
    fb = combined["fec_ber"]
    if op["headers"] is None:
        op_hdr, op_rows = [], []
    else:
        op_hdr, op_rows = drop_columns(op["headers"], op["rows"], ["index", "Building"])
        OPTICS_ORDER = ["Source", "Input Power", "Output Power", "Hostname", "Interface", "Rack", "Elevation"]
        op_hdr, op_rows = reorder_columns(op_hdr, op_rows, [c for c in OPTICS_ORDER if c in op_hdr])
    if fb["headers"] is None:
        fb_hdr, fb_rows = [], []
    else:
        fb_hdr, fb_rows = drop_columns(fb["headers"], fb["rows"], ["index", "BER", "Lock", "Remote Host", "Remote Interface", "Reason"])
    t2_mis_hdr, t2_mis_rows = enrich(t2_mis_hdr, t2_mis_rows, [("Active Host", "Act. Interface"), ("Hostname", "Interface")], lookup)
    t1_mis_hdr, t1_mis_rows = enrich(t1_mis_hdr, t1_mis_rows, [("Active Host", "Act. Interface"), ("Hostname", "Interface")], lookup)
    if op_hdr:
        op_hdr, op_rows = enrich(op_hdr, op_rows, [("Hostname", "Interface")], lookup)
    if fb_hdr:
        fb_hdr, fb_rows = enrich(fb_hdr, fb_rows, [("Hostname", "Interface")], lookup)
    for hdr, rows in [(t2_mis_hdr, t2_mis_rows), (t1_mis_hdr, t1_mis_rows), (op_hdr, op_rows), (fb_hdr, fb_rows)]:
        fill_empty_pp(hdr, rows)
    t2_dl_hdr, t2_dl_rows = merge_columns(t2_dl_hdr, t2_dl_rows)
    t1_dl_hdr, t1_dl_rows = merge_columns(t1_dl_hdr, t1_dl_rows)
    t2_mis_hdr, t2_mis_rows = merge_columns(t2_mis_hdr, t2_mis_rows)
    t1_mis_hdr, t1_mis_rows = merge_columns(t1_mis_hdr, t1_mis_rows)
    op_hdr, op_rows = merge_columns(op_hdr, op_rows)
    fb_hdr, fb_rows = merge_columns(fb_hdr, fb_rows)
    op_t2_rows, op_t1_rows = split_by_cutsheet_pp(op_hdr, op_rows)
    fb_t2_rows, fb_t1_rows = split_by_cutsheet_pp(fb_hdr, fb_rows)
    # === Clustering (exact original logic) ===
    t2_mis_rows, t2_mis_colors = sort_mismatch_pairs(t2_mis_hdr, t2_mis_rows)
    t1_mis_rows, t1_mis_colors = sort_mismatch_pairs(t1_mis_hdr, t1_mis_rows)
    sheets_to_write = [
        ("T2-T1 Downlink", t2_dl_hdr + ["Note"], [r + [None] for r in t2_dl_rows]),
        ("T1-T0 Downlink", t1_dl_hdr + ["Note"], [r + [None] for r in t1_dl_rows]),
        ("T2-T1 Mismatch", t2_mis_hdr + ["Note"], [r + [None] for r in t2_mis_rows]),
        ("T1-T0 Mismatch", t1_mis_hdr + ["Note"], [r + [None] for r in t1_mis_rows]),
        ("T2-T1 Optics", op_hdr + ["Note"], [r + [None] for r in op_t2_rows]),
        ("T1-T0 Optics", op_hdr + ["Note"], [r + [None] for r in op_t1_rows]),
        ("T2-T1 fec_ber", fb_hdr + ["Note"], [r + [None] for r in fb_t2_rows]),
        ("T1-T0 fec_ber", fb_hdr + ["Note"], [r + [None] for r in fb_t1_rows]),
    ]
    sheet_col_widths = {name: compute_col_widths(hdr, rows) for name, hdr, rows in sheets_to_write}
    wb = Workbook()
    wb.remove(wb.active)
    for name, hdr, rows in sheets_to_write:
        write_sheet(wb, name, hdr, rows)
    build_summary(wb, labels, [
        ("Downlink", ["T2-T1 Downlink", "T1-T0 Downlink"]),
        ("Mismatch", ["T2-T1 Mismatch", "T1-T0 Mismatch"]),
        ("optics", ["T2-T1 Optics", "T1-T0 Optics"]),
        ("fec_ber", ["T2-T1 fec_ber", "T1-T0 fec_ber"]),
    ])
    canonical = ["summary", "T2-T1 Downlink", "T1-T0 Downlink", "T2-T1 Mismatch", "T1-T0 Mismatch", "T2-T1 Optics", "T1-T0 Optics", "T2-T1 fec_ber", "T1-T0 fec_ber"]
    wb._sheets = [wb[n] for n in canonical if n in wb.sheetnames]
    for name in wb.sheetnames:
        capitalize_b_numbers(wb[name])
    mismatch_colors = {"T2-T1 Mismatch": t2_mis_colors, "T1-T0 Mismatch": t1_mis_colors}
    for name in wb.sheetnames:
        pink = PINK_COLS if name.endswith("Mismatch") else ()
        freeze = "D2" if name.endswith("Optics") else "A2"
        rc = mismatch_colors.get(name)
        cw = sheet_col_widths.get(name)
        style_sheet(wb[name], pink_col_names=pink, freeze_at=freeze, row_colors=rc, col_widths=cw)
    for dl_tab in ("T2-T1 Downlink", "T1-T0 Downlink"):
        if dl_tab in wb.sheetnames:
            annotate_downlink_switches(wb[dl_tab])
    annotate_optics_in_downlinks(wb)
    widen_note_columns(wb)
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue(), "-".join(labels) + "_QFAB_Updated.xlsx"

# ================== Streamlit UI ==================
st.set_page_config(page_title="SYD20 QFAB (V4)", page_icon="📊", layout="wide")
st.title("SYD20 QFAB / Slack Formatter — V4")
st.caption("Improved mismatch logic (V4)")

st.markdown("### 1. Cutsheet")
cutsheet_file = st.file_uploader("Select the CUTSHEET xlsx", type=["xlsx", "xlsm"], key="cutsheet_v4")

st.markdown("### 2. Input Files (one or more per-building files)")
input_files = st.file_uploader("Select input audit files", type=["xlsx", "xlsm"], accept_multiple_files=True, key="inputs_v4")

if st.button("🚀 Process Files", type="primary", disabled=not (cutsheet_file and input_files), key="process_v4"):
    with st.spinner("Processing with V4 (exact original mismatch logic)..."):
        try:
            result_bytes, filename = process_files(cutsheet_file.getvalue(), input_files)
            if result_bytes:
                st.success("✅ Processing complete!")
                st.download_button("📥 Download Formatted Report", data=result_bytes, file_name=filename, mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)
        except Exception as e:
            st.error(f"Error during processing: {e}")
            st.exception(e)

st.caption("All processing happens locally in your browser.")
