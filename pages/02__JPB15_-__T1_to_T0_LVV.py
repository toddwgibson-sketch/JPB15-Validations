import streamlit as st
from pathlib import Path
import tempfile
import shutil
import io
import contextlib
import os
import re
import sys
from datetime import datetime
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Consistent header with logo ──────────────────────────────────────────────
def show_header(title: str, subtitle: str = ""):
    col1, col2 = st.columns([1, 8])
    with col1:
        logo_path = Path(__file__).parent.parent / "assets" / "LOGO.png"
        if logo_path.exists():
            st.image(str(logo_path), width=130)
        else:
            st.markdown("### 🔧")
    with col2:
        st.markdown(f"### {title}")
        if subtitle:
            st.caption(subtitle)

st.set_page_config(page_title="JPB15 T1-T0 LVV", page_icon="🔗", layout="wide")
show_header("")

# ── All original constants and helper functions from the provided script ─────
TAB_SUMM = "1F4E79"
TAB_MISS = "C00000"
TAB_DOWN = "ED7D31"
TAB_OPT  = "833C00"
TAB_FEC  = "7030A0"

HDR_NAVY    = "1F4E79"
HDR_SRC     = "C0504D"
HDR_DM1     = "7F6000"
HDR_DM2     = "375623"
HDR_DEST    = "17375E"
HDR_POSS_A  = "833C00"
HDR_POSS_Z  = "375623"
HDR_ACT     = "9C0006"
HDR_EXP     = "375623"
HDR_GREY    = "595959"

POSS_A_BG   = "FDDCB5"
POSS_DM1_BG = "FFF2CC"
POSS_DM2_BG = "FCE4D6"
POSS_Z_BG   = "D5F5E3"

DL_GREY_BG    = "C8C8C8"
DL_GREY_LR_BG = "A8A8A8"
DL_GREY_FG    = "888888"
DL_FLAG_FG    = "666666"

WHITE = "FFFFFF"
MISS_BG = "FFF2CC"

_THIN  = Side(style="thin",   color="AAAAAA")
_MED   = Side(style="medium", color="555555")

BORDER_LEFTMOST = Border(left=_MED,  right=_THIN, top=_MED, bottom=_MED)
BORDER_MIDDLE   = Border(left=_THIN, right=_THIN, top=_MED, bottom=_MED)
BORDER_RIGHTMOST = Border(left=_THIN, right=_MED, top=_MED, bottom=_MED)

def fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)

def center() -> Alignment:
    return Alignment(horizontal="center", vertical="center", wrap_text=False)

def border_for(col: int, max_col: int) -> Border:
    if col == 1:
        return BORDER_LEFTMOST
    if col == max_col:
        return BORDER_RIGHTMOST
    return BORDER_MIDDLE

RACK_RE = re.compile(r"Rack\s+(\S+)\s+U\s*(\S+)", re.IGNORECASE)
IFACE_RE = re.compile(r"swp(\d+)s(\d+)", re.IGNORECASE)

def parse_rack_u(s: str) -> tuple[str, str]:
    if not s:
        return "", ""
    m = RACK_RE.match(str(s).strip())
    return (m.group(1), m.group(2)) if m else ("", "")

def iface_to_lr(iface: str) -> str:
    if not iface:
        return ""
    m = IFACE_RE.match(str(iface).strip())
    if not m:
        return ""
    port = int(m.group(1))
    lane = int(m.group(2))
    even_port = (port % 2 == 0)
    low_lane  = (lane <= 1)
    side = "L" if even_port == low_lane else "R"
    return f"{port}{side}"

def pair_search_order(iface: str) -> list[str]:
    if not iface:
        return []
    m = IFACE_RE.match(str(iface).strip())
    if not m:
        return [iface]
    port = int(m.group(1))
    lane = int(m.group(2))
    base = f"swp{port}"
    orderings = {
        0: [0, 1],
        1: [1, 2, 0],
        2: [2, 1, 3],
        3: [3, 2],
    }
    return [f"{base}s{n}" for n in orderings.get(lane, [lane])]

# ── Full original build_lookup (new + legacy support) ────────────────────────
def build_lookup(cutsheet_paths: list[str]) -> tuple[dict, dict]:
    forward: dict = {}
    reverse: dict = {}

    for path in cutsheet_paths:
        try:
            wb = load_workbook(path, data_only=True)
        except Exception as e:
            st.warning(f"Could not load {os.path.basename(path)}: {e}")
            continue

        for ws in wb.worksheets:
            if ws.max_row < 2:
                continue

            headers = {}
            for col in range(1, ws.max_column + 1):
                h = ws.cell(1, col).value
                if isinstance(h, str):
                    headers[h.strip().lower()] = col

            def col_of(*names: str):
                for n in names:
                    c = headers.get(n.lower())
                    if c:
                        return c
                return None

            c_phys_a   = col_of("DeviceA Physical Port", "physa", "phys a", "L/R", "L&R", "lr")
            c_dev_a    = col_of("DeviceA", "device a", "a device")
            c_host_a   = col_of("Hostname", "host", "device", "device name")
            c_iface_a  = col_of("Interface", "port", "iface")
            c_rack_a   = col_of("RackA", "rack a", "a rack", "Rack")
            c_elev_a   = col_of("Elevation", "U", "elevation a")
            c_src_pp   = col_of("Source_port", "source port", "source_port")
            c_dm1      = col_of("DMARC1", "dmarc 1")
            c_dm2      = col_of("DMARC2", "dmarc 2")
            c_dst_pp   = col_of("Destination_port", "destination port", "dest_port")
            c_dev_b    = col_of("DeviceB", "device b", "b device")
            c_host_b   = col_of("Z Hostname", "z host", "remote hostname", "remote device")
            c_iface_b  = col_of("Z Interface", "z port", "z iface", "remote interface", "remote port")
            c_rack_b   = col_of("RackB", "rack b", "b rack", "Z Rack")
            c_elev_b   = col_of("Z Elevation", "z u", "elevation b", "remote elevation")
            c_phys_b   = col_of("DeviceB Physical Port", "physb", "phys b", "Z L/R", "Z L&R", "z lr")

            have_a_combined = bool(c_dev_a)
            have_a_separate = bool(c_host_a and c_iface_a)
            if not (have_a_combined or have_a_separate) or not c_src_pp:
                continue

            def read_str(ws, row, col):
                if not col:
                    return ""
                v = ws.cell(row, col).value
                return str(v).strip() if v is not None else ""

            for row in range(2, ws.max_row + 1):
                if have_a_separate:
                    hostname = read_str(ws, row, c_host_a)
                    iface    = read_str(ws, row, c_iface_a)
                else:
                    dev_a = read_str(ws, row, c_dev_a)
                    parts = dev_a.split()
                    if len(parts) < 2:
                        continue
                    hostname, iface = parts[0], parts[1]
                if not hostname or not iface:
                    continue

                rack_raw = read_str(ws, row, c_rack_a)
                rack_a, elev_a = parse_rack_u(rack_raw)
                if not rack_a and rack_raw:
                    rack_a = rack_raw
                    elev_a = read_str(ws, row, c_elev_a)
                rack_a_full = f"Rack {rack_a} U{elev_a}" if rack_a else ""

                if c_host_b and c_iface_b:
                    dev_b   = read_str(ws, row, c_host_b)
                    iface_b = read_str(ws, row, c_iface_b)
                elif c_dev_b:
                    dev_b_full = read_str(ws, row, c_dev_b)
                    db_parts = dev_b_full.split()
                    dev_b   = db_parts[0] if db_parts else ""
                    iface_b = db_parts[1] if len(db_parts) > 1 else ""
                else:
                    dev_b, iface_b = "", ""

                rack_b_raw = read_str(ws, row, c_rack_b)
                rack_b, elev_b = parse_rack_u(rack_b_raw)
                if not rack_b and rack_b_raw:
                    rack_b = rack_b_raw
                    elev_b = read_str(ws, row, c_elev_b)
                rack_b_full = f"Rack {rack_b} U{elev_b}" if rack_b else ""

                forward[(hostname.lower(), iface.lower())] = {
                    "phys_a":      read_str(ws, row, c_phys_a),
                    "rack_a":      rack_a,
                    "elev_a":      elev_a,
                    "rack_a_full": rack_a_full,
                    "src_pp":      read_str(ws, row, c_src_pp),
                    "dmarc1":      read_str(ws, row, c_dm1),
                    "dmarc2":      read_str(ws, row, c_dm2),
                    "dest_pp":     read_str(ws, row, c_dst_pp),
                    "dev_b":       dev_b,
                    "iface_b":     iface_b,
                    "rack_b":      rack_b,
                    "elev_b":      elev_b,
                    "rack_b_full": rack_b_full,
                    "phys_b":      read_str(ws, row, c_phys_b),
                }

                if dev_b and iface_b:
                    reverse[(dev_b.lower(), iface_b.lower())] = {
                        "rack": rack_b,
                        "elev": elev_b,
                    }
    return forward, reverse

def parse_matrix(matrix: str) -> dict:
    out = {"rack_a": "", "elev_a": "", "rack_b": "", "elev_b": "",
           "src_pp": "", "dmarc1": "", "dmarc2": "", "dest_pp": ""}
    if not matrix or str(matrix).strip().lower() == "missing":
        return out
    lines = [ln.strip() for ln in str(matrix).splitlines() if ln.strip()]
    rack_idxs = [i for i, ln in enumerate(lines) if RACK_RE.match(ln)]
    if rack_idxs:
        m = RACK_RE.match(lines[rack_idxs[0]])
        if m:
            out["rack_a"], out["elev_a"] = m.group(1), m.group(2)
    if len(rack_idxs) >= 2:
        m = RACK_RE.match(lines[rack_idxs[-1]])
        if m:
            out["rack_b"], out["elev_b"] = m.group(1), m.group(2)
    for ln in lines:
        u = ln.upper()
        if not u.startswith("PP."):
            continue
        if ".DH" in u and "MPO" in u:
            if ".DH1." in u or "DH10" in u:
                if not out["dmarc1"]:
                    out["dmarc1"] = ln
            elif ".DH2." in u:
                if not out["dmarc2"]:
                    out["dmarc2"] = ln
        else:
            if not out["src_pp"]:
                out["src_pp"] = ln
            else:
                out["dest_pp"] = ln
    return out

def lookup_or_matrix(lookup, hostname: str, iface: str, matrix: str,
                     remote_dev: str = "", remote_port: str = "") -> dict:
    forward, reverse = lookup
    host_lower = str(hostname).strip().lower()
    iface_lower = str(iface).strip().lower()

    if (host_lower, iface_lower) in forward:
        info = dict(forward[(host_lower, iface_lower)])
        info["cutsheet_miss"] = False
        return info

    candidates = pair_search_order(iface_lower)
    pair_entry = None
    for cand in candidates[1:]:
        if (host_lower, cand) in forward:
            pair_entry = forward[(host_lower, cand)]
            break

    if pair_entry is not None:
        a_rack = pair_entry["rack_a"]
        a_elev = pair_entry["elev_a"]
        a_rack_full = pair_entry["rack_a_full"]

        z_dev   = str(remote_dev).strip()
        z_iface = str(remote_port).strip()
        z_rack, z_elev = "", ""
        if z_dev and z_iface:
            for z_cand in pair_search_order(z_iface.lower()):
                r = reverse.get((z_dev.lower(), z_cand))
                if r:
                    z_rack, z_elev = r["rack"], r["elev"]
                    break

        return {
            "phys_a":        iface_to_lr(iface),
            "rack_a":        a_rack,
            "elev_a":        a_elev,
            "rack_a_full":   a_rack_full,
            "src_pp":        pair_entry["src_pp"],
            "dmarc1":        pair_entry["dmarc1"],
            "dmarc2":        pair_entry["dmarc2"],
            "dest_pp":       pair_entry["dest_pp"],
            "dev_b":         z_dev,
            "iface_b":       z_iface,
            "rack_b":        z_rack,
            "elev_b":        z_elev,
            "rack_b_full":   f"Rack {z_rack} U{z_elev}" if z_rack else "",
            "phys_b":        iface_to_lr(z_iface),
            "cutsheet_miss": False,
        }

    parsed = parse_matrix(matrix)
    has_matrix = any(parsed.values())
    return {
        "phys_a":        iface_to_lr(iface),
        "rack_a":        parsed["rack_a"],
        "elev_a":        parsed["elev_a"],
        "rack_a_full":   f"Rack {parsed['rack_a']} U{parsed['elev_a']}" if parsed["rack_a"] else "",
        "src_pp":        parsed["src_pp"],
        "dmarc1":        parsed["dmarc1"],
        "dmarc2":        parsed["dmarc2"],
        "dest_pp":       parsed["dest_pp"],
        "dev_b":         "",
        "iface_b":       "",
        "rack_b":        parsed["rack_b"],
        "elev_b":        parsed["elev_b"],
        "rack_b_full":   f"Rack {parsed['rack_b']} U{parsed['elev_b']}" if parsed["rack_b"] else "",
        "phys_b":        "",
        "cutsheet_miss": True,
        "matrix_only":   has_matrix,
    }

# ── Sheet building helpers (full original logic) ─────────────────────────────
def style_header(cell, fill_hex: str, font_size: int = 10) -> None:
    cell.fill = fill(fill_hex)
    cell.font = Font(bold=True, color=WHITE, name="Arial", size=font_size)
    cell.alignment = center()

def write_headers(ws, headers: list[tuple[str, str]]) -> None:
    for col, (text, fhex) in enumerate(headers, start=1):
        style_header(ws.cell(1, col, text), fhex)
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"

def write_data_cell(ws, row: int, col: int, val, fg: str = "000000",
                    bold: bool = False, bg: str = WHITE,
                    max_col: int | None = None) -> None:
    c = ws.cell(row, col, val if val != "" else None)
    c.fill = fill(bg)
    c.font = Font(color=fg, name="Arial", size=9, bold=bold)
    c.alignment = center()
    if max_col is not None:
        c.border = border_for(col, max_col)

def set_widths(ws, widths: list[int]) -> None:
    for col, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = w

def draw_pair_borders(ws, iface_col: int = 1, rack_col: int = 3, u_col: int = 4) -> None:
    if ws.max_row < 2:
        return
    thin  = Side(style="thin",   color="AAAAAA")
    thick = Side(style="medium", color="555555")

    def group_key(row: int) -> tuple:
        iface = str(ws.cell(row, iface_col).value or "")
        rack  = str(ws.cell(row, rack_col).value  or "")
        u     = str(ws.cell(row, u_col).value     or "")
        m = IFACE_RE.match(iface)
        if not m:
            return (rack, u, iface, -1)
        base_port = f"swp{m.group(1)}"
        lane = int(m.group(2))
        cable_group = 1 if lane in (1, 2) else lane
        return (rack, u, base_port, cable_group)

    max_col = ws.max_column
    dr = 2
    while dr <= ws.max_row:
        key = group_key(dr)
        if not key or key[3] < 0:
            dr += 1
            continue
        grp_end = dr
        while grp_end + 1 <= ws.max_row and group_key(grp_end + 1) == key:
            grp_end += 1
        if grp_end > dr:
            for rr in range(dr, grp_end + 1):
                is_top = (rr == dr)
                is_bot = (rr == grp_end)
                for cc in range(1, max_col + 1):
                    is_left  = (cc == 1)
                    is_right = (cc == max_col)
                    ws.cell(rr, cc).border = Border(
                        top    = thick if is_top   else thin,
                        bottom = thick if is_bot   else thin,
                        left   = thick if is_left  else thin,
                        right  = thick if is_right else thin,
                    )
        dr = grp_end + 1

CHAN_LINE_RE = re.compile(r"channel_?(\d+)\s*:\s*([-+]?\d+\.?\d*)\s*(\(failed\))?", re.IGNORECASE)

def parse_rx_power(s: str) -> list[tuple[str, str, bool]]:
    out = []
    if not s:
        return out
    for line in str(s).splitlines():
        m = CHAN_LINE_RE.search(line)
        if m:
            out.append((m.group(1), m.group(2), bool(m.group(3))))
    return out

# ── Full original build_downlinks ────────────────────────────────────────────
def build_downlinks(wb_out, src_ws, lookup: dict) -> tuple[int, set]:
    ws = wb_out.create_sheet("Downlinks")
    ws.sheet_properties.tabColor = TAB_DOWN

    headers = [
        ("Interface",        HDR_NAVY),
        ("L&R",              HDR_NAVY),
        ("Rack",             HDR_NAVY),
        ("Elevation",        HDR_NAVY),
        ("Source_port",      HDR_SRC),
        ("DMARC1",           HDR_DM1),
        ("DMARC2",           HDR_DM2),
        ("Destination_port", HDR_DEST),
        ("Z Interface",      HDR_DEST),
        ("L&R",              HDR_DEST),
        ("Z Rack",           HDR_DEST),
        ("Z Elevation",      HDR_DEST),
        ("History",          HDR_GREY),
    ]
    write_headers(ws, headers)
    set_widths(ws, [12, 6, 8, 8, 32, 32, 32, 32, 12, 6, 8, 10, 12])

    downlink_set: set = set()
    if src_ws is None or src_ws.max_row < 2:
        return 0, downlink_set

    hmap = {}
    for c in range(1, src_ws.max_column + 1):
        raw = str(src_ws.cell(1, c).value or '').strip().lower()
        hmap[raw] = c
        simplified = raw.replace(" ", "").replace("_", "").replace("-", "")
        hmap[simplified] = c

    def gv(r, *names):
        for n in names:
            key = n.lower().replace(" ", "").replace("_", "").replace("-", "")
            c = hmap.get(key) or hmap.get(n.lower())
            if c:
                v = src_ws.cell(r, c).value
                if v is not None:
                    return str(v)
        return ""

    out_row = 2
    count = 0
    for r in range(2, src_ws.max_row + 1):
        host = gv(r, "Source Device Name")
        port = gv(r, "Source Device Port")
        if not host or not port:
            continue

        downlink_set.add((host.lower(), port.lower()))

        rem_dev  = gv(r, "Remote Device Name")
        rem_port = gv(r, "Remote Device Port")
        matrix   = gv(r, "Patch Panel Matrix")

        info = lookup_or_matrix(lookup, host, port, matrix,
                                remote_dev=rem_dev, remote_port=rem_port)

        z_iface = info["iface_b"] or rem_port
        z_lr    = info["phys_b"] or iface_to_lr(z_iface)
        z_rack  = info["rack_b"]
        z_elev  = info["elev_b"]

        miss = info.get("cutsheet_miss", False)
        row_bg = MISS_BG if miss else WHITE
        history_note = "⚠ Not in cutsheet" if miss else ""

        values = [
            port,
            info["phys_a"] or iface_to_lr(port),
            info["rack_a"],
            info["elev_a"],
            info["src_pp"],
            info["dmarc1"],
            info["dmarc2"],
            info["dest_pp"],
            z_iface,
            z_lr,
            z_rack,
            z_elev,
            history_note,
        ]
        ws.row_dimensions[out_row].height = 15
        for col, v in enumerate(values, start=1):
            bold = (col == 2)
            write_data_cell(ws, out_row, col, v, bold=bold, bg=row_bg, max_col=len(headers))
        out_row += 1
        count += 1
    draw_pair_borders(ws)
    return count, downlink_set

# ── Full original build_optics (per-channel) ─────────────────────────────────
def build_optics(wb_out, src_ws, lookup: dict, downlink_set: set) -> int:
    ws = wb_out.create_sheet("Optics")
    ws.sheet_properties.tabColor = TAB_OPT

    headers = [
        ("Interface",        HDR_NAVY),
        ("L&R",              HDR_NAVY),
        ("Rack",             HDR_NAVY),
        ("Elevation",        HDR_NAVY),
        ("Channel",          HDR_NAVY),
        ("Measured (dBm)",   HDR_NAVY),
        ("Source_port",      HDR_NAVY),
        ("DMARC1",           HDR_NAVY),
        ("DMARC2",           HDR_NAVY),
        ("Destination_port", HDR_NAVY),
        ("Z Interface",      HDR_NAVY),
        ("Z L&R",            HDR_NAVY),
        ("Z Rack",           HDR_NAVY),
        ("Z Elevation",      HDR_NAVY),
        ("DL Flag",          HDR_GREY),
        ("History",          HDR_GREY),
    ]
    write_headers(ws, headers)
    set_widths(ws, [12, 6, 8, 8, 8, 12, 32, 32, 32, 32, 12, 6, 8, 10, 22, 12])

    if src_ws is None or src_ws.max_row < 2:
        return 0

    hmap = {}
    for c in range(1, src_ws.max_column + 1):
        raw = str(src_ws.cell(1, c).value or '').strip().lower()
        hmap[raw] = c
        simplified = raw.replace(" ", "").replace("_", "").replace("-", "")
        hmap[simplified] = c

    def gv(r, *names):
        for n in names:
            key = n.lower().replace(" ", "").replace("_", "").replace("-", "")
            c = hmap.get(key) or hmap.get(n.lower())
            if c:
                v = src_ws.cell(r, c).value
                if v is not None:
                    return str(v)
        return ""

    out_row = 2
    count = 0
    for r in range(2, src_ws.max_row + 1):
        host = gv(r, "Source Device Name")
        port = gv(r, "Source Device Port")
        if not host or not port:
            continue

        rx_raw   = gv(r, "Rx Power")
        rem_dev  = gv(r, "Remote Device Name")
        rem_port = gv(r, "Remote Device Port")
        matrix   = gv(r, "Patch Panel Matrix")
        info     = lookup_or_matrix(lookup, host, port, matrix,
                                    remote_dev=rem_dev, remote_port=rem_port)

        is_dl = (host.lower(), port.lower()) in downlink_set
        miss = info.get("cutsheet_miss", False)

        if is_dl:
            row_bg    = DL_GREY_BG
            lr_bg     = DL_GREY_LR_BG
            text_fg   = DL_GREY_FG
            flag_text = "⬇️ Also Downlink — skip"
            history_note = ""
        elif miss:
            row_bg    = MISS_BG
            lr_bg     = MISS_BG
            text_fg   = "000000"
            flag_text = ""
            history_note = "⚠ Not in cutsheet"
        else:
            row_bg    = WHITE
            lr_bg     = WHITE
            text_fg   = "000000"
            flag_text = ""
            history_note = ""

        z_iface = info["iface_b"] or gv(r, "Remote Device Port")
        z_lr    = info["phys_b"] or iface_to_lr(z_iface)

        channels = [(ch, val) for ch, val, failed in parse_rx_power(rx_raw) if failed]
        if not channels:
            channels = [("", rx_raw or "")]

        for ch, val in channels:
            row_values = [
                (port,                                       row_bg,  False),
                (info["phys_a"] or iface_to_lr(port),        lr_bg,   True),
                (info["rack_a"],                             row_bg,  False),
                (info["elev_a"],                             row_bg,  False),
                (ch,                                         row_bg,  False),
                (val,                                        row_bg,  False),
                (info["src_pp"],                             row_bg,  False),
                (info["dmarc1"],                             row_bg,  False),
                (info["dmarc2"],                             row_bg,  False),
                (info["dest_pp"],                            row_bg,  False),
                (z_iface,                                    row_bg,  False),
                (z_lr,                                       lr_bg,   True),
                (info["rack_b"],                             row_bg,  False),
                (info["elev_b"],                             row_bg,  False),
                (flag_text,                                  row_bg,  is_dl),
                (history_note,                               row_bg,  False),
            ]
            ws.row_dimensions[out_row].height = 15
            for col, (v, bg, bold) in enumerate(row_values, start=1):
                fg = DL_FLAG_FG if (col == 15 and is_dl) else text_fg
                write_data_cell(ws, out_row, col, v, fg=fg, bold=bold, bg=bg,
                                max_col=len(headers))
            out_row += 1
            count += 1
    draw_pair_borders(ws)
    return count

# ── Full original build_fec ──────────────────────────────────────────────────
RAWBER_LANE_RE = re.compile(r"lane_?(\d+)\s*:\s*([-+]?\d+\.?\d*[eE]?[-+]?\d*)\s*(\(failed\))?")

def classify_ber_severity(ber: str) -> tuple[str, str, str]:
    if not ber:
        return ("", WHITE, "000000")
    try:
        val = float(ber)
    except (TypeError, ValueError):
        return ("", WHITE, "000000")
    if val < 1e-7:
        return ("", WHITE, "000000")
    if val < 1e-6:
        return ("Marginal", "FFF2CC", "7F6000")
    if val < 1e-5:
        return ("Warning",  "FCE4D6", "9C5700")
    return     ("Severe",   "FFC7CE", "9C0006")

def extract_max_failed_ber(raw_ber: str, pre_fec: str) -> tuple[str, str]:
    if pre_fec and pre_fec.lower() != "missing":
        return ("", pre_fec)
    failed = []
    for line in (raw_ber or "").splitlines():
        m = RAWBER_LANE_RE.search(line)
        if m and m.group(3):
            try:
                failed.append((int(m.group(1)), float(m.group(2)), m.group(2)))
            except ValueError:
                pass
    if not failed:
        return ("", "")
    failed.sort(key=lambda x: -x[1])
    worst_lane, _, worst_val = failed[0]
    lock = f"RAW_BER_MAX={worst_val} (channel(s)={worst_lane}) > 1e-07"
    return (lock, worst_val)

def build_fec(wb_out, src_ws, lookup: dict, downlink_set: set) -> int:
    ws = wb_out.create_sheet("FEC Errors")
    ws.sheet_properties.tabColor = TAB_FEC

    headers = [
        ("Interface",        HDR_NAVY),
        ("L&R",              HDR_NAVY),
        ("Rack",             HDR_NAVY),
        ("Elevation",        HDR_NAVY),
        ("Lock Status",      HDR_NAVY),
        ("Pre-FEC BER",      HDR_NAVY),
        ("Severity",         HDR_NAVY),
        ("Source_port",      HDR_NAVY),
        ("DMARC1",           HDR_NAVY),
        ("DMARC2",           HDR_NAVY),
        ("Destination_port", HDR_NAVY),
        ("Z Interface",      HDR_NAVY),
        ("Z L&R",            HDR_NAVY),
        ("Z Rack",           HDR_NAVY),
        ("Z Elevation",      HDR_NAVY),
        ("Remote Interface", HDR_NAVY),
        ("DL Flag",          HDR_GREY),
        ("History",          HDR_GREY),
    ]
    write_headers(ws, headers)
    set_widths(ws, [12, 6, 8, 8, 40, 12, 11, 32, 32, 32, 32, 12, 6, 8, 10, 14, 22, 12])

    if src_ws is None or src_ws.max_row < 2:
        return 0

    hmap = {}
    for c in range(1, src_ws.max_column + 1):
        raw = str(src_ws.cell(1, c).value or '').strip().lower()
        hmap[raw] = c
        simplified = raw.replace(" ", "").replace("_", "").replace("-", "")
        hmap[simplified] = c

    def gv(r, *names):
        for n in names:
            key = n.lower().replace(" ", "").replace("_", "").replace("-", "")
            c = hmap.get(key) or hmap.get(n.lower())
            if c:
                v = src_ws.cell(r, c).value
                if v is not None:
                    return str(v)
        return ""

    out_row = 2
    count = 0
    for r in range(2, src_ws.max_row + 1):
        host = gv(r, "Device Name", "Source Device Name")
        port = gv(r, "Device Port", "Source Device Port")
        if not host or not port:
            continue

        rem_dev  = gv(r, "Remote Device Name")
        rem_port = gv(r, "Remote Device Port")
        pre_fec  = gv(r, "PRE_FEC_BER")
        raw_ber  = gv(r, "Optical RawBer")
        lock_in  = gv(r, "Lock Status")
        matrix   = gv(r, "Patch Panel Matrix")

        info = lookup_or_matrix(lookup, host, port, matrix,
                                remote_dev=rem_dev, remote_port=rem_port)

        lock_derived, ber = extract_max_failed_ber(raw_ber, pre_fec)
        lock_status = lock_in if (lock_in and lock_in.lower() != "missing") else lock_derived

        is_dl = (host.lower(), port.lower()) in downlink_set
        miss  = info.get("cutsheet_miss", False)
        if is_dl:
            row_bg  = DL_GREY_BG
            lr_bg   = DL_GREY_LR_BG
            text_fg = DL_GREY_FG
            flag    = "⬇️ Also Downlink — skip"
            history_note = ""
        elif miss:
            row_bg  = MISS_BG
            lr_bg   = MISS_BG
            text_fg = "000000"
            flag    = ""
            history_note = "⚠ Not in cutsheet"
        else:
            row_bg  = WHITE
            lr_bg   = WHITE
            text_fg = "000000"
            flag    = ""
            history_note = ""

        z_iface = info["iface_b"] or rem_port
        z_lr    = info["phys_b"] or iface_to_lr(z_iface)

        sev_label, sev_bg, sev_fg = classify_ber_severity(ber)
        if is_dl:
            sev_bg, sev_fg = row_bg, text_fg

        row_values = [
            (port,                                       row_bg,  False),
            (info["phys_a"] or iface_to_lr(port),        lr_bg,   True),
            (info["rack_a"],                             row_bg,  False),
            (info["elev_a"],                             row_bg,  False),
            (lock_status,                                row_bg,  False),
            (ber,                                        row_bg,  False),
            (sev_label,                                  sev_bg,  True),
            (info["src_pp"],                             row_bg,  False),
            (info["dmarc1"],                             row_bg,  False),
            (info["dmarc2"],                             row_bg,  False),
            (info["dest_pp"],                            row_bg,  False),
            (z_iface,                                    row_bg,  False),
            (z_lr,                                       lr_bg,   True),
            (info["rack_b"],                             row_bg,  False),
            (info["elev_b"],                             row_bg,  False),
            (rem_port,                                   row_bg,  False),
            (flag,                                       row_bg,  is_dl),
            (history_note,                               row_bg,  False),
        ]
        ws.row_dimensions[out_row].height = 15
        for col, (v, bg, bold) in enumerate(row_values, start=1):
            if col == 7:
                fg = sev_fg
            elif col == 17 and is_dl:
                fg = DL_FLAG_FG
            else:
                fg = text_fg
            write_data_cell(ws, out_row, col, v, fg=fg, bold=bold, bg=bg,
                            max_col=len(headers))
        out_row += 1
        count += 1
    draw_pair_borders(ws)
    return count

# ── Full original build_summary ──────────────────────────────────────────────
def build_summary(wb_out, report_name: str, n_miss: int, n_down: int,
                  n_opt: int, n_fec: int) -> None:
    ws = wb_out.create_sheet("Summary", 0)
    ws.sheet_properties.tabColor = TAB_SUMM

    title = ws.cell(1, 2, "VALIDATION REPORT — SUMMARY")
    style_header(title, "1F4E79", font_size=14)
    ws.row_dimensions[1].height = 28
    ws.merge_cells(start_row=1, end_row=1, start_column=2, end_column=4)

    rpt = ws.cell(2, 2, f"Report: {report_name}")
    rpt.fill = fill("0D7377")
    rpt.font = Font(color=WHITE, name="Arial", size=10, italic=True)
    rpt.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.merge_cells(start_row=2, end_row=2, start_column=2, end_column=4)

    total = n_miss + n_down + n_opt + n_fec
    style_header(ws.cell(4, 2, "TOTAL ISSUES"), "1F4E79")
    ws.merge_cells(start_row=4, end_row=4, start_column=2, end_column=3)
    tot_cell = ws.cell(4, 4, total)
    tot_cell.fill = fill("1F4E79")
    tot_cell.font = Font(color=WHITE, name="Arial", size=14, bold=True)
    tot_cell.alignment = center()
    ws.row_dimensions[4].height = 24

    style_header(ws.cell(6, 2, "ERROR TYPE"),    "1F4E79")
    style_header(ws.cell(6, 3, "COUNT"),         "1F4E79")
    ws.merge_cells(start_row=6, end_row=6, start_column=3, end_column=4)

    breakdown = [
        ("Mispatches", n_miss, "C00000"),
        ("Downlinks",  n_down, "ED7D31"),
        ("Optics",     n_opt,  "833C00"),
        ("FEC Errors", n_fec,  "7030A0"),
    ]
    for i, (label, n, hex_color) in enumerate(breakdown):
        r = 7 + i
        ws.row_dimensions[r].height = 20
        lab = ws.cell(r, 2, label)
        lab.fill = fill(hex_color)
        lab.font = Font(color=WHITE, name="Arial", size=10, bold=True)
        lab.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        cnt = ws.cell(r, 3, n)
        cnt.fill = fill("F2F2F2")
        cnt.font = Font(color="000000", name="Arial", size=11, bold=True)
        cnt.alignment = center()
        ws.merge_cells(start_row=r, end_row=r, start_column=3, end_column=4)

    set_widths(ws, [3, 24, 12, 12])

# ── Streamlit UI ─────────────────────────────────────────────────────────────
st.markdown("Upload cutsheet(s) and the LV Portal report. Multiple cutsheets supported.")

cutsheet_files = st.file_uploader(
    "Cutsheet(s) — multiple allowed (new or legacy layout)",
    type=["xlsx"],
    accept_multiple_files=True,
    key="qfab_full_cutsheets"
)
report_file = st.file_uploader(
    "LV Portal Validation Export (.xlsx)",
    type=["xlsx"],
    key="qfab_full_report"
)

if st.button("🚀 Process Report (Full Fidelity)", type="primary",
             disabled=not (cutsheet_files and report_file)):
    temp_dir = tempfile.mkdtemp()
    log_capture = io.StringIO()
    try:
        with contextlib.redirect_stdout(log_capture):
            print("Loading cutsheets...")
            cutsheet_paths = []
            for i, f in enumerate(cutsheet_files):
                p = os.path.join(temp_dir, f"cutsheet_{i}.xlsx")
                with open(p, "wb") as out:
                    out.write(f.getbuffer())
                cutsheet_paths.append(p)

            lookup = build_lookup(cutsheet_paths)
            print(f"Indexed {len(lookup[0])} forward + {len(lookup[1])} reverse entries")

            report_path = os.path.join(temp_dir, "report.xlsx")
            with open(report_path, "wb") as out:
                out.write(report_file.getbuffer())

            wb_src = load_workbook(report_path, data_only=True)

            def find_sheet(wb, *patterns):
                for name in wb.sheetnames:
                    low = name.lower()
                    for p in patterns:
                        if p.lower() in low:
                            return wb[name]
                return None

            ws_optics = find_sheet(wb_src, "optic")
            ws_fec    = find_sheet(wb_src, "fec")
            ws_iface  = find_sheet(wb_src, "interface down", "interface_down", "downlink")

            wb_out = Workbook()
            wb_out.remove(wb_out.active)

            n_miss = 0  # Mispatches is empty stub for this format
            n_down, downlink_set = build_downlinks(wb_out, ws_iface, lookup)
            n_opt = build_optics(wb_out, ws_optics, lookup, downlink_set)
            n_fec = build_fec(wb_out, ws_fec, lookup, downlink_set)

            build_summary(wb_out, report_file.name, n_miss, n_down, n_opt, n_fec)

            buf = io.BytesIO()
            wb_out.save(buf)
            buf.seek(0)
            output_bytes = buf.getvalue()

        st.success("✅ Processing complete! Output matches original script fidelity.")
        st.download_button(
            "📥 Download Formatted Report",
            data=output_bytes,
            file_name=report_file.name.replace(".xlsx", "_formatted.xlsx"),
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )
        with st.expander("Processing Log"):
            st.text(log_capture.getvalue())
    finally:
        shutil.rmtree(temp_dir, ignore_errors=True)

st.caption("Full fidelity conversion — all original logic from lv_portal_formatter.qfabt0 preserved.")
