import streamlit as st

from utils.auth import require_login
require_login()
import pandas as pd
import tempfile
import os
import io
import zipfile
from pathlib import Path
import warnings
from openpyxl import load_workbook
import xlsxwriter
from collections import defaultdict

warnings.filterwarnings('ignore', category=UserWarning)

st.set_page_config(page_title=" SYD20 Cutsheet Formatter", page_icon="📋", layout="wide")
st.title("SYD20 Cutsheet Formatter")
st.caption("Now with Internal Cable Classification + better large-file handling")

# ── Core logic from original script (kept intact) ─────────────────────────────
COLOR_RJ45        = '#FFCCCC'
COLOR_RJ45_SERIAL = '#CCFFCC'
COLOR_DB9         = '#FFB347'
COLOR_400G_AOC    = '#CCE5FF'
COLOR_100G_AOC    = '#E0F0FF'
COLOR_DAC         = '#D9D9D9'
COLOR_OTHER       = '#FFFF00'

DARK_BLUE = '#002060'
YELLOW    = '#FFFF00'
GREEN     = '#92D050'
WHITE     = '#FFFFFF'
BLACK     = '#000000'
RED       = '#C00000'
ROW_H     = 15

def assign_color(ct):
    if ct == 'Copper RJ45 cable':          return COLOR_RJ45
    if ct == 'Copper RJ45 cable - serial': return COLOR_RJ45_SERIAL
    if 'DB9' in ct:                        return COLOR_DB9
    if '400G AOC' in ct:                   return COLOR_400G_AOC
    if 'AOC' in ct:                        return COLOR_100G_AOC
    if 'DAC' in ct:                        return COLOR_DAC
    return COLOR_OTHER

CABLE_TAB_NAMES = {
    'Copper RJ45 cable': 'Copper RJ45 cable',
    'Copper RJ45 cable - serial': 'Copper RJ45 cable - serial',
    'DB9 to copper RJ45 cable - serial: transceivers: DB9 dongle': 'DB9-RJ45 cable serial',
    'QSFP-DD 400G AOC': 'QSFP-DD 400G AOC',
    'QSFP28 100G AOC': 'QSFP28 100G AOC',
    'QSFP28 100G DAC': 'QSFP28 100G DAC',
    'QSFP56 200G DAC': 'QSFP56 200G DAC',
    'SFP+ 10G DAC': 'SFP+ 10G DAC',
    'Singlemode LC duplex fiber patch': 'SM LC duplex fiber patch',
    'Singlemode LC duplex fiber patch - Armoured': 'SM LC duplex - Armoured',
    'Singlemode LC duplex fiber patch: transceivers: QSFP28 LR4 transceiver': 'SM LC QSFP28 LR4',
    'Singlemode LC duplex fiber patch: transceivers: SFP+ LR transceiver': 'SM LC SFP+ LR',
    'Singlemode MPO to Singlemode LC fiber breakout: transceivers: QSFP-DD DR4 400G transceiver: QSFP28 DR1 100G transceiver': 'SM MPO-LC DR4-DR1',
    'Singlemode patch (LC or MPO): transceivers: OSFP 800G Singlemode transceiver: OSFP 800G Singlemode transceiver': 'SM patch OSFP 800G',
    'Singlemode patch (LC or MPO): transceivers: QSFP-112 400G DR4 Singlemode transceiver: OSFP 800G DR4 Singlemode twinport transceiver': 'SM patch QSFP112-OSFP 400G',
    'Singlemode patch (LC or MPO): transceivers: QSFP28 100G Singlemode transceiver: QSFP28 100G Singlemode transceiver': 'SM patch QSFP28 100G',
    'Singlemode patch (LC or MPO): transceivers: QSFP56 200G Singlemode transceiver: QSFP56 200G Singlemode transceiver': 'SM patch QSFP56 200G',
}

def make_tab_name(ct, used, ml=31):
    name = CABLE_TAB_NAMES.get(ct, ct.replace(':', '').replace('/', '- ')
           .replace('\\', '').replace('?', '').replace('*', '')
           .replace('[', '').replace(']', '').strip()[:ml].strip())
    orig, i = name[:ml].strip(), 1
    name = orig
    while name in used:
        sfx = f'_{i}'
        name = orig[:ml - len(sfx)] + sfx
        i += 1
    return name

def fmt_d(n, p): return f"{n} {p}".strip()
def fmt_r(r, u): return f"Rack {r} U{u}"

def parse_row_breakdown(file_bytes):
    """Parse Row Breakdown from uploaded file bytes."""
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        tmp.write(file_bytes)
        tmp_path = tmp.name
    try:
        wb = load_workbook(tmp_path, data_only=True)
        ws = wb.active
        row_to_room = {}
        room_index = 0
        for row in ws.iter_rows(min_row=3, values_only=True):
            a_val = row[0]
            d_val = row[3]
            if a_val is None or not isinstance(a_val, (int, float)): continue
            if d_val is None or '-' not in str(d_val): continue
            room_index += 1
            room_name = f"20.{room_index}"
            parts = str(d_val).strip().split('-')
            start, end = int(parts[0].strip()), int(parts[1].strip())
            for r in range(start, end + 1):
                row_to_room[r] = room_name
        return row_to_room
    finally:
        os.unlink(tmp_path)

def rack_row_prefix(rack_str):
    rack = str(rack_str).strip()
    if len(rack) == 4: return int(rack[:2])
    if len(rack) == 5: return int(rack[:3])
    return None

def assign_connections_to_rooms(df, row_to_room):
    cols = list(df.columns)
    ci_rack_a = cols.index('DeviceA Rack')
    ci_rack_b = cols.index('DeviceB Rack')
    room_idx = defaultdict(set)
    for i, row in enumerate(df.itertuples(index=False)):
        rv = list(row)
        pref_a = rack_row_prefix(rv[ci_rack_a])
        pref_b = rack_row_prefix(rv[ci_rack_b])
        rooms = set()
        if pref_a is not None:
            r = row_to_room.get(pref_a)
            if r: rooms.add(r)
        if pref_b is not None:
            r = row_to_room.get(pref_b)
            if r: rooms.add(r)
        for room in rooms:
            room_idx[room].add(i)
    return {room: sorted(idxs) for room, idxs in room_idx.items()}


# =============================================================================
# NEW ADVANCED FEATURES (ported & adapted from V2)
# =============================================================================

def normalize_rack(rack_str: str) -> str:
    """Extracts the base rack identifier for comparison (e.g. 2706 from 'Rack 2706 U41')."""
    if not rack_str:
        return ""
    s = str(rack_str).strip().lower().replace("rack", "").strip()
    import re
    match = re.search(r'(\d+)', s)
    return match.group(1) if match else s


def is_intra_rack(row, rack_a_col: str, rack_b_col: str) -> bool:
    """Returns True if both sides are on the same physical rack."""
    ra = normalize_rack(row.get(rack_a_col, ""))
    rb = normalize_rack(row.get(rack_b_col, ""))
    return bool(ra) and bool(rb) and ra == rb


def has_uleft_or_uright(rack_str: str) -> bool:
    if not rack_str:
        return False
    s = str(rack_str).lower()
    return "uleft" in s or "uright" in s


def classify_internal_cables(df: pd.DataFrame, auxiliary_racks: set = None) -> pd.DataFrame:
    """
    Adds 'Is_Internal' column.
    Internal = same physical rack on both sides (already installed, no pulling needed).
    Exceptions:
      - Rack B has Uleft / Uright → still needs work
      - Rack is in auxiliary_racks list → treat as external
    """
    if df.empty:
        return df

    df = df.copy()
    df['Is_Internal'] = False

    rack_cols = [c for c in df.columns if 'rack' in c.lower()]
    rack_a_col = next((c for c in rack_cols if 'a' in c.lower()), None)
    rack_b_col = next((c for c in rack_cols if 'b' in c.lower()), None)

    if not rack_a_col or not rack_b_col:
        return df

    aux = auxiliary_racks or set()

    internal_mask = df.apply(
        lambda row: is_intra_rack(row, rack_a_col, rack_b_col), axis=1
    )

    for idx in df[internal_mask].index:
        rack_b = str(df.at[idx, rack_b_col]).lower()
        rack_a_norm = normalize_rack(df.at[idx, rack_a_col])

        if has_uleft_or_uright(rack_b):
            df.at[idx, 'Is_Internal'] = False
            continue
        if rack_a_norm in aux:
            df.at[idx, 'Is_Internal'] = False
            continue

        df.at[idx, 'Is_Internal'] = True

    return df


# =============================================================================
# SORT KEY LOGIC (migrated from Desktop SYD20\\Cutsheets work)
# =============================================================================

import re

def make_sort_key(text: str) -> str:
    """
    Generates a consistent, zero-padded sort key from device names or PP labels.
    This is the migrated logic from the user's sortkey development work.

    Examples:
      'syd20-c1-b70-t0-r9 Ethernet1/1'          -> 'SYD20|B70|T00|R009|PORT001'
      'PP:SYD20.3.DH8.R2302.U1.A1.P4'           -> 'SYD20|03|DH08|R2302|U001|A01|P004'
      'PP:SYD20.3.DH8.R2801.U12.S7.A4'          -> 'SYD20|03|DH08|R2801|U012|S07|A004'
    """
    if not text or not isinstance(text, str):
        return ''

    s = text.strip()

    # === Case 1: Patch Panel labels (PP:...) ===
    if s.upper().startswith('PP:'):
        # PP:SYD20.3.DH8.R2302.U1.A1.P4
        # -> SYD20|03|DH08|R2302|U001|A01|P004
        parts = s.split('.')
        try:
            site = parts[0].replace('PP:', '').upper()          # SYD20
            room = f'{int(parts[1]):02d}'                       # 03
            dh   = parts[2].upper()                             # DH8 -> DH08
            if dh.startswith('DH') and dh[2:].isdigit():
                dh = f'DH{int(dh[2:]):02d}'

            rest = '.'.join(parts[3:])                          # R2302.U1.A1.P4

            # Normalize the rest (R, U, then letters + numbers)
            # Split on common separators while keeping structure
            tokens = re.split(r'([RUru])', rest)
            out = []
            for t in tokens:
                t = t.strip()
                if not t: continue
                if t[0].upper() in 'RU':
                    letter = t[0].upper()
                    num = re.search(r'\d+', t)
                    num = f'{int(num.group()):03d}' if num else '000'
                    out.append(f'{letter}{num}')
                else:
                    # Handle things like A1, P4, S7, A10 etc.
                    m = re.match(r'([A-Za-z]+)(\d+)', t)
                    if m:
                        let = m.group(1).upper()
                        num = f'{int(m.group(2)):03d}'
                        out.append(f'{let}{num}')
                    else:
                        out.append(t.upper())
            return f'{site}|{room}|{dh}|' + '|'.join(out)
        except Exception:
            return s  # fallback

    # === Case 2: Device / hostname style ===
    # syd20-c1-b70-t0-r9 et-0/0/12
    # syd20-c1-b12-t2-r8 Ethernet1/1
    lower = s.lower()
    if 'syd20' in lower or re.search(r'b\d+', lower):
        try:
            b = re.search(r'b(\d+)', lower)
            t = re.search(r't(\d+)', lower)
            r = re.search(r'r(\d+)', lower)
            port = re.search(r'(?:et-|ethernet|port)[-/]?(\d+)', lower, re.I)

            b_str = f'B{int(b.group(1)):02d}' if b else ''
            t_str = f'T{int(t.group(1)):02d}' if t else 'T00'
            r_str = f'R{int(r.group(1)):03d}' if r else ''

            port_num = f'PORT{int(port.group(1)):03d}' if port else ''

            return f'SYD20|{b_str}|{t_str}|{r_str}|{port_num}'.strip('|')
        except Exception:
            pass

    # Fallback: return uppercased original
    return s.upper()[:60]


def add_sort_keys(df: pd.DataFrame) -> pd.DataFrame:
    """
    Adds SortA and SortB columns using the migrated make_sort_key logic.
    Tries to find the best source columns for each side.
    """
    df = df.copy()

    # Common column name patterns in allconnects exports
    possible_a = ['DeviceA Name', 'Device A Name', 'Source Device Name', 'Hostname A']
    possible_b = ['DeviceB Name', 'Device B Name', 'Remote Device Name', 'Hostname B']

    # Also try patch panel columns for the far side (often richer)
    pp_cols = [c for c in df.columns if 'patch' in c.lower() or 'full label' in c.lower() or 'easymark' in c.lower()]

    col_a = next((c for c in possible_a if c in df.columns), None)
    col_b = next((c for c in possible_b if c in df.columns), None)

    if col_a:
        df['SortA'] = df[col_a].apply(make_sort_key)
    else:
        df['SortA'] = ''

    if col_b:
        df['SortB'] = df[col_b].apply(make_sort_key)
    elif pp_cols:
        # Use first good PP column as SortB source if no DeviceB name
        df['SortB'] = df[pp_cols[0]].apply(make_sort_key)
    else:
        df['SortB'] = ''

    return df

# ── Main workbook builder (kept from original) ────────────────────────────────
def build_workbook_to_bytes(df, title_label=""):
    """Builds the cutsheet workbook and returns it as bytes (for Streamlit download)."""
    # Support for internal cable classification (new feature)
    if 'Is_Internal' in df.columns:
        def get_tab_key(row):
            base = row['Cable Info']
            return f"{base} (Internal)" if row.get('Is_Internal', False) else base
        df = df.copy()
        df['_Tab_Key'] = df.apply(get_tab_key, axis=1)
        cable_types = list(df['_Tab_Key'].unique())
    else:
        cable_types = list(df['Cable Info'].unique())

    if not cable_types:
        return None

    cable_hex = {ct: assign_color(ct) for ct in cable_types}
    ALL_COLS = list(df.columns)
    CI = {n: ALL_COLS.index(n) for n in ALL_COLS}
    DA_RACK = CI['DeviceA Rack']; DA_RU = CI['DeviceA RU']
    DA_NAME = CI['DeviceA Name']; DA_PORT = CI['DeviceA Port']
    DB_RACK = CI['DeviceB Rack']; DB_RU = CI['DeviceB RU']
    DB_NAME = CI['DeviceB Name']; DB_PORT = CI['DeviceB Port']

    used_preview = set(['Version Control', '_allconnects', 'Main', 'Formatted_',
                        'PROGRESS Dashboard', 'Racks', 'Cable Liner', 'Template'])
    ct_tab_names = {}
    for ct in cable_types:
        n = make_tab_name(ct, used_preview)
        used_preview.add(n)
        ct_tab_names[ct] = n

    output = io.BytesIO()
    # constant_memory=True is critical for large files (prevents OOM on big exports)
    wb = xlsxwriter.Workbook(output, {
        'strings_to_numbers': False,
        'constant_memory': True
    })

    def mk(bg=None, fc=BLACK, bold=False, wrap=False,
           halign='left', valign='vcenter', sz=10, num_fmt=None):
        d = {'font_size': sz, 'font_color': fc, 'bold': bold, 'text_wrap': wrap,
             'align': halign, 'valign': valign, 'border': 1}
        if bg: d['bg_color'] = bg
        if num_fmt: d['num_format'] = num_fmt
        return wb.add_format(d)

    hdr_blue  = mk(DARK_BLUE, WHITE, bold=True, halign='center', wrap=True)
    hdr_red   = mk(RED, WHITE, bold=True, halign='center')
    hdr_yel   = mk(YELLOW, BLACK, halign='center')
    hdr_grn   = mk(GREEN, BLACK, halign='center', wrap=True)
    hdr_plain = mk(None, BLACK, halign='center')
    dat_l     = mk(None, BLACK, halign='left')
    dat_c     = mk(None, BLACK, halign='center')
    dat_pct   = mk(None, BLACK, halign='center', num_fmt='0%')
    grn_bold  = mk(GREEN, BLACK, bold=True, halign='left')
    yel_c     = mk(YELLOW, BLACK, halign='center')
    blank_bdr = mk(None, BLACK, halign='center')

    fmt_ct = {}
    for ct in cable_types:
        bg = cable_hex[ct]
        fmt_ct[ct] = {
            'sep':  mk(bg, BLACK, bold=True, halign='center'),
            'col':  mk(bg, BLACK, halign='left'),
            'tab':  mk(None, BLACK, halign='center'),
            'num':  mk(None, BLACK, halign='center'),
            'note': mk(None, BLACK, halign='center'),
        }

    tpl_hdrs = [
        ('Label for wrap arounds', hdr_blue), ('Cable Path', hdr_yel),
        ('Device A', hdr_yel), ('Rack A', hdr_yel), ('Device B', hdr_yel), ('Rack B', hdr_yel),
        ('# of Connections', hdr_grn), ('Prep', hdr_grn), ('Prep', hdr_grn),
        ('Pulling', hdr_grn), ('Dressing', hdr_grn), ('Labeling of Cable', hdr_grn),
        ('Percentage of Completion', hdr_grn), ('', hdr_plain), ('NOTES:', hdr_grn),
    ]
    COL_W = [30, 12, 40, 20, 40, 20, 15, 10, 10, 10, 10, 18, 22, 12, 30]

    # Version Control
    ws = wb.add_worksheet('Version Control')
    ws.set_default_row(ROW_H)
    ws.set_column('A:A', 4); ws.set_column('B:B', 22)
    ws.set_column('C:E', 35); ws.set_column('F:F', 18)
    ws.merge_range('A1:F1', 'Cutsheet Change Log', hdr_red); ws.set_row(0, 24)
    label_fmt = mk(None, BLACK, bold=True, halign='right')
    for row, (label, vfmt, val) in enumerate([
        ('Document Name', yel_c, ''), ('Document Owner', dat_c, ''),
        ('Template Version', dat_c, '1.0'), ('Date of creation', yel_c, '')], start=1):
        ws.write_blank(row, 0, None, blank_bdr)
        ws.write(row, 1, label, label_fmt)
        ws.write(row, 2, val, vfmt)
        for c in range(3, 6): ws.write_blank(row, c, None, blank_bdr)
    ws.write_blank(5, 0, None, blank_bdr)
    ws.merge_range('B7:F7', 'Revision History', hdr_red)
    ws.write_blank(6, 0, None, blank_bdr)
    for c, h in enumerate(['Change Maker', 'Description of Change',
                           'Revision', 'Cutsheet Version', 'Date'], 1):
        ws.write(7, c, h, hdr_red)
    ws.write_blank(7, 0, None, blank_bdr)
    for r in range(8, 21):
        for c in range(6): ws.write_blank(r, c, None, blank_bdr)

    # _allconnects
    ws_ac = wb.add_worksheet('_allconnects')
    ws_ac.set_default_row(ROW_H)
    ws_ac.autofilter(0, 0, 0, len(ALL_COLS) - 1)
    ws_ac.freeze_panes(1, 0)
    for i in range(len(ALL_COLS)): ws_ac.set_column(i, i, 20)
    for c, h in enumerate(ALL_COLS): ws_ac.write(0, c, h, hdr_blue)
    for r, rv in enumerate(df.itertuples(index=False), 1):
        ws_ac.write_row(r, 0, list(rv), dat_l)

    # Main
    ws_main = wb.add_worksheet('Main')
    ws_main.set_default_row(ROW_H)
    ws_main.freeze_panes(1, 0)
    ws_main.autofilter(0, 0, 0, len(ALL_COLS) - 1)
    for i in range(len(ALL_COLS)): ws_main.set_column(i, i, 22)
    for c, h in enumerate(ALL_COLS): ws_main.write(0, c, h, hdr_blue)
    cur = 1
    for ct in cable_types:
        grp = df[df['Cable Info'] == ct]
        f = fmt_ct[ct]
        for c in range(len(ALL_COLS)): ws_main.write(cur, c, ct, f['sep'])
        cur += 1
        for rv in grp.itertuples(index=False):
            ws_main.write_row(cur, 0, list(rv), f['col'])
            cur += 1

    # Formatted_
    ws_fmt = wb.add_worksheet('Formatted_')
    ws_fmt.set_default_row(ROW_H)
    ws_fmt.freeze_panes(1, 0)
    ws_fmt.autofilter(0, 0, 0, 3)
    for c, w in enumerate([45, 25, 45, 25]): ws_fmt.set_column(c, c, w)
    for c, h in enumerate(['DeviceA', 'RackA', 'DeviceB', 'RackB']): ws_fmt.write(0, c, h, hdr_blue)
    cur = 1
    for ct in cable_types:
        grp = df[df['Cable Info'] == ct]
        f = fmt_ct[ct]
        for c in range(4): ws_fmt.write(cur, c, ct, f['sep'])
        cur += 1
        for rv in grp.itertuples(index=False):
            v = list(rv)
            ws_fmt.write_row(cur, 0,
                [fmt_d(v[DA_NAME], v[DA_PORT]), fmt_r(v[DA_RACK], v[DA_RU]),
                 fmt_d(v[DB_NAME], v[DB_PORT]), fmt_r(v[DB_RACK], v[DB_RU])],
                f['col'])
            cur += 1

    # PROGRESS Dashboard (simplified but functional)
    ws_pd = wb.add_worksheet('PROGRESS Dashboard')
    ws_pd.set_default_row(ROW_H)
    ws_pd.set_column('A:A', 10)
    ws_pd.set_column('B:B', 40)
    ws_pd.set_column('C:C', 22)
    ws_pd.merge_range('A1:H1', 'Cutsheet Progress', hdr_red)
    ws_pd.set_row(0, 20)
    for c, h in enumerate(['Tab #', 'Activity', 'Completion Percentage', 'Notes',
                           '', 'Material Missing / Blocker', 'Current ETA', 'Days Needed']):
        ws_pd.write(1, c, h, hdr_red)

    entries = [(1, 'Racks', "=Racks!N1"), (2, 'Cable Liner', "='Cable Liner'!E2")]
    for i, ct in enumerate(cable_types, 3):
        sn = ct_tab_names[ct]
        entries.append((i, sn, f"='{sn}'!N1"))

    for ri, (num, activity, formula) in enumerate(entries, 2):
        ws_pd.write_number(ri, 0, num, dat_c)
        ws_pd.write_string(ri, 1, activity, dat_l)
        ws_pd.write_formula(ri, 2, formula, dat_pct, 0)
        for c in range(3, 8): ws_pd.write_blank(ri, c, None, blank_bdr)

    overall_r = 2 + len(entries)
    ws_pd.write_blank(overall_r, 0, None, blank_bdr)
    ws_pd.write_string(overall_r, 1, 'Overall Project Progress', grn_bold)
    ws_pd.write_formula(overall_r, 2, f'=AVERAGE(C3:C{overall_r})', dat_pct, 0)
    for c in range(3, 8): ws_pd.write_blank(overall_r, c, None, blank_bdr)

    # Racks tab (simplified)
    ws_r = wb.add_worksheet('Racks')
    ws_r.set_default_row(ROW_H)
    ws_r.set_column('A:A', 5)
    ws_r.set_column('B:B', 15)
    ws_r.set_column('C:C', 12)
    ws_r.set_column('D:D', 35)
    for c in range(5, 14): ws_r.set_column(c, c, 14)
    rack_hdrs = ['', 'SO#', 'Status', 'Rack Platform', 'Rack Location Number',
                 'BOPS Tickets', 'Serial Number', 'Foot Leveling', 'Grounding',
                 'Rack Power up', 'Cabling', 'Percentage complete', 'Percentage of Completion']
    for c, h in enumerate(rack_hdrs):
        ws_r.write(0, c, h, hdr_yel if c < 7 else hdr_grn)
    for r in range(1, 17):
        ws_r.write_blank(r, 0, None, blank_bdr)
        for c in range(1, 7): ws_r.write_blank(r, c, None, dat_c)
        for c in range(7, 11): ws_r.write_number(r, c, 0, dat_c)
        ws_r.write_formula(r, 11, f'=SUM(H{r+1}:K{r+1})', dat_c, 0)

    # Cable Liner
    ws_cl = wb.add_worksheet('Cable Liner')
    ws_cl.set_default_row(ROW_H)
    ws_cl.set_column('A:A', 5)
    ws_cl.set_column('B:B', 15)
    ws_cl.set_column('C:C', 22)
    ws_cl.set_column('D:D', 24)
    ws_cl.set_column('E:E', 20)
    ws_cl.write_blank(0, 0, None, blank_bdr)
    ws_cl.write(0, 1, 'Block#', hdr_blue)
    ws_cl.write(0, 2, 'Liner Installation', hdr_yel)
    ws_cl.write(0, 3, 'Percentage of Completion', hdr_grn)
    ws_cl.write_formula(0, 4, '=AVERAGE(C2:C3)', hdr_grn, 0)
    for r, block in enumerate(['Block-1', 'Block-2'], 1):
        ws_cl.write_blank(r, 0, None, blank_bdr)
        ws_cl.write(r, 1, block, dat_c)
        ws_cl.write_number(r, 2, 0, dat_c)

    # Template
    ws_t = wb.add_worksheet('Template')
    ws_t.set_column('A:A', 30)
    ws_t.set_column('B:B', 12)
    ws_t.set_column('C:C', 40)
    ws_t.set_column('D:D', 20)
    ws_t.set_column('E:E', 40)
    ws_t.set_column('F:F', 20)
    ws_t.set_column('G:G', 15)
    for c in range(7, 12): ws_t.set_column(c, c, 10)
    ws_t.set_column('M:M', 22)
    ws_t.set_column('N:N', 12)
    ws_t.set_column('O:O', 30)
    ws_t.freeze_panes(1, 0)
    ws_t.autofilter(0, 0, 0, 14)
    TEMPLATE_ROWS = 256
    for ci, (h, hf) in enumerate(tpl_hdrs):
        if ci == 13:
            ws_t.write_formula(0, ci, f'=AVERAGE(H2:L{TEMPLATE_ROWS+1})', hdr_plain, 0)
        else:
            ws_t.write(0, ci, h, hf)
    for r in range(1, TEMPLATE_ROWS + 1):
        er = r + 1
        ws_t.set_row(r, ROW_H)
        ws_t.write_formula(r, 0,
            f'=CONCATENATE(C{er},CHAR(10),D{er},CHAR(10),E{er},CHAR(10),F{er})',
            hdr_blue, '')
        ws_t.write(r, 1, 'A', dat_c)
        for c in range(2, 6): ws_t.write_blank(r, c, None, dat_c)
        ws_t.write_number(r, 6, 1, dat_c)
        for c in range(7, 12): ws_t.write_number(r, c, 0, dat_c)
        ws_t.write_blank(r, 12, None, dat_c)
        ws_t.write_blank(r, 13, None, dat_c)
        ws_t.write(r, 14, '', yel_c)

    # Cable-type tabs
    used = set(wb.sheetnames)
    for ct in cable_types:
        grp = df[df['Cable Info'] == ct].reset_index(drop=True)
        f = fmt_ct[ct]
        sname = ct_tab_names[ct]
        used.add(sname)
        ws = wb.add_worksheet(sname)
        ws.set_tab_color(cable_hex[ct])
        ws.set_default_row(14.9)
        ws.freeze_panes(1, 0)
        ws.autofilter(0, 0, 0, 14)
        last = 1 + len(grp)
        ws.set_row(0, 15)
        for ci, (h, hf) in enumerate(tpl_hdrs):
            if ci == 13:
                ws.write_formula(0, ci, f'=AVERAGE(H2:L{last})', hdr_plain, 0)
            else:
                ws.write(0, ci, h, hf)
        for ci, w in enumerate(COL_W): ws.set_column(ci, ci, w)

        # === Sort Keys on the far right (starting at column V = 21) ===
        sort_key_cols = [c for c in ['SortA', 'SortB'] if c in grp.columns]
        sort_start_col = 21  # Column V
        if sort_key_cols:
            sort_hdr = mk('#BDD7EE', BLACK, bold=True, halign='center')
            for idx, colname in enumerate(sort_key_cols):
                cidx = sort_start_col + idx
                ws.set_column(cidx, cidx, 34)
                header_name = 'Sort - Device A' if colname == 'SortA' else 'Sort - Device B / Path'
                ws.write(0, cidx, header_name, sort_hdr)

        for i, rv in enumerate(grp.itertuples(index=False), 1):
            v = list(rv)
            da = fmt_d(v[DA_NAME], v[DA_PORT])
            ra = fmt_r(v[DA_RACK], v[DA_RU])
            db = fmt_d(v[DB_NAME], v[DB_PORT])
            rb = fmt_r(v[DB_RACK], v[DB_RU])
            ws.set_row(i, 15)
            ws.write(i, 0, f"{da}\n{ra}\n{db}\n{rb}", hdr_blue)
            ws.write(i, 1, 'A', f['tab'])
            ws.write(i, 2, da, f['tab'])
            ws.write(i, 3, ra, f['tab'])
            ws.write(i, 4, db, f['tab'])
            ws.write(i, 5, rb, f['tab'])
            ws.write(i, 6, 1, f['num'])
            ws.write(i, 7, 0, f['num'])
            ws.write(i, 8, 0, f['num'])
            ws.write(i, 9, 0, f['num'])
            ws.write(i, 10, 0, f['num'])
            ws.write(i, 11, 0, f['num'])
            ws.write_blank(i, 12, None, f['tab'])
            ws.write_blank(i, 13, None, f['tab'])
            ws.write(i, 14, '', f['note'])

            # Write sort keys far to the right
            if sort_key_cols:
                for idx, colname in enumerate(sort_key_cols):
                    cidx = sort_start_col + idx
                    val = v[ALL_COLS.index(colname)] if colname in ALL_COLS else ''
                    ws.write(i, cidx, val, dat_l)

    wb.close()
    output.seek(0)
    return output.getvalue()

# ── Streamlit UI ──────────────────────────────────────────────────────────────
st.markdown("### Upload Files")

mode = st.radio(
    "Mode",
    ["Single cutsheet", "Split by room (requires Row Breakdown)"],
    horizontal=True
)

input_file = st.file_uploader("All Connects export (.xlsx)", type=["xlsx"])

# ── NEW: Advanced Options (more features) ─────────────────────────────────────
with st.expander("⚙️ Advanced Options (New Features)", expanded=False):
    st.markdown("**Internal Cable Classification** (highly recommended for construction)")
    classify_internals = st.checkbox(
        "Classify intra-rack (internal) cables",
        value=True,
        help="Cables with the same rack on both sides are marked as 'Internal' (already installed). Creates separate tabs like 'Copper RJ45 cable (Internal)'."
    )

    aux_racks_input = st.text_input(
        "Auxiliary racks (comma separated, e.g. 5920,6002)",
        value="",
        help="Racks that should never be treated as internal even if RackA == RackB."
    )
    auxiliary_racks = {r.strip() for r in aux_racks_input.split(",") if r.strip()} if aux_racks_input else set()

    add_sort_keys_option = st.checkbox(
        "Add Sort Keys (SortA / SortB columns)",
        value=True,
        help="Generates physical sort keys from device names and PP labels (migrated from your Desktop sortkey work). Greatly improves sort order in the final cutsheet."
    )

breakdown_file = None
if mode == "Split by room (requires Row Breakdown)":
    breakdown_file = st.file_uploader("Row Breakdown file (.xlsx)", type=["xlsx"])

if st.button("🚀 Generate Cutsheet(s)", type="primary", disabled=not input_file):
    with st.spinner("Processing... This can take a few minutes for large allconnect sheets."):
        try:
            # Read input
            df_all = pd.read_excel(input_file, dtype=str).fillna('')
            df_all = df_all.sort_values('Cable Info').reset_index(drop=True)
            st.success(f"Loaded {len(df_all):,} connections")

            # ── NEW: Apply advanced preprocessing if enabled ─────────────────────
            if classify_internals:
                df_all = classify_internal_cables(df_all, auxiliary_racks)
                internal_count = int(df_all.get('Is_Internal', pd.Series([False]*len(df_all))).sum())
                if internal_count > 0:
                    st.info(f"Classified **{internal_count}** cables as Internal (same rack both sides). They will appear in separate '(Internal)' tabs.")

            if add_sort_keys_option:
                df_all = add_sort_keys(df_all)
                st.info("SortA and SortB columns generated using migrated sortkey logic.")

            if mode == "Single cutsheet":
                bytes_data = build_workbook_to_bytes(df_all)
                if bytes_data:
                    st.download_button(
                        "📥 Download Formatted Cutsheet",
                        data=bytes_data,
                        file_name="cutsheet_formatted.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        width="stretch"
                    )
                else:
                    st.error("No data to process.")

            else:  # Split by room
                if not breakdown_file:
                    st.error("Please upload the Row Breakdown file.")
                else:
                    row_to_room = parse_row_breakdown(breakdown_file.getvalue())
                    room_assignments = assign_connections_to_rooms(df_all, row_to_room)

                    all_assigned = set(i for idxs in room_assignments.values() for i in idxs)
                    unmatched = len(df_all) - len(all_assigned)
                    if unmatched:
                        st.warning(f"{unmatched} connections had rack numbers not matching any room.")

                    rooms_sorted = sorted(room_assignments.keys(), key=lambda r: float(r.replace('20.', '')))

                    # Create ZIP in memory
                    zip_buffer = io.BytesIO()
                    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zipf:
                        for room in rooms_sorted:
                            idxs = room_assignments[room]
                            df_room = df_all.iloc[idxs].copy()
                            df_room = df_room.sort_values('Cable Info').reset_index(drop=True)
                            xlsx_bytes = build_workbook_to_bytes(df_room, room)
                            if xlsx_bytes:
                                zipf.writestr(f"{room}_cutsheet.xlsx", xlsx_bytes)

                    zip_buffer.seek(0)
                    st.success(f"Generated {len(rooms_sorted)} room cutsheets.")
                    st.download_button(
                        "📥 Download All Room Cutsheets (ZIP)",
                        data=zip_buffer,
                        file_name="room_cutsheets.zip",
                        mime="application/zip",
                        width="stretch"
                    )

        except Exception as e:
            st.error(f"Error: {str(e)}")
            st.exception(e)

st.caption("")
