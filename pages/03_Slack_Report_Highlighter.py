import streamlit as st
from pathlib import Path
import tempfile
import shutil
import io
import contextlib
import os
import re
import copy
from datetime import datetime
from collections import Counter
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Header ───────────────────────────────────────────────────────────────────
def show_header(title: str, subtitle: str = ""):
    col1, col2 = st.columns([1, 8])
    with col1:
        logo_path = Path(__file__).parent.parent / "assets" / "logo.png"
        if logo_path.exists():
            st.image(str(logo_path), width=70)
        else:
            st.markdown("### 🔧")
    with col2:
        st.markdown(f"### {title}")
        if subtitle:
            st.caption(subtitle)

st.set_page_config(page_title="Slack Report Highlighter", page_icon="📊", layout="wide")
show_header("Slack Report Highlighter v9", "Classic Slack validation reports with recurring detection")

# ── Styles ───────────────────────────────────────────────────────────────────
WHITE = "FFFFFF"
YELLOW = "FFFF00"
LOG_BG = "FFFFFF"
HDR_BG = "1F4E79"
SRC_BG = "FCE4D6"
D1_BG = "FFF2CC"
D2_BG = "E2F0D9"
DEST_BG = "D9EAF7"
ACT_BG = "FFC7CE"
EXP_BG = "C6EFCE"
LR_BG = "FFFFFF"
TAB_MISS = "C00000"
TAB_DOWN = "ED7D31"
TAB_OPT = "833C00"
TAB_FEC = "7030A0"

def fill(h): return PatternFill("solid", fgColor=h)
def font(color="000000", bold=False, sz=9): return Font(bold=bold, color=color, name="Arial", size=sz)
def center(): return Alignment(horizontal="center", vertical="center")

# ── Cutsheet loader (new + legacy) ───────────────────────────────────────────
def build_lookup(paths):
    t0, t1, t1_rev = {}, {}, {}
    for path in paths:
        wb = load_workbook(path, read_only=True)
        sheet = next((wb[n] for n in wb.sheetnames if 'installation' in n.lower()), wb[wb.sheetnames[0]])
        header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
        headers = {str(h or '').strip(): i for i, h in enumerate(header_row)}
        new_layout = ('Hostname' in headers and 'Interface' in headers)
        count = 0
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not row or all(v is None for v in row): continue
            def v(name, legacy_idx=None):
                idx = headers.get(name)
                if idx is not None and idx < len(row):
                    return str(row[idx] or '').strip()
                if legacy_idx is not None and legacy_idx < len(row):
                    return str(row[legacy_idx] or '').strip()
                return ''
            if new_layout:
                hostname = v('Hostname')
                interface = v('Interface')
                lbl = v('L/R')
                src = v('Source_port')
                dmarc1 = v('DMARC1')
                dmarc2 = v('DMARC2')
                dest_p = v('Destination_port')
                z_hostname = v('Z Hostname')
                z_interface = v('Z Interface')
                t1l = v('Z L/R')
                rack_b = v('Z Rack')
            else:
                lbl = v('Label', 0)
                device_a = v('DeviceA', 1)
                parts = device_a.split()
                hostname = parts[0] if parts else ''
                interface = parts[1] if len(parts) > 1 else ''
                src = v('Source_port', 3)
                dmarc1 = v('DMARC1', 4)
                dmarc2 = v('DMARC2', 5)
                dest_p = v('Destination_port', 6)
                device_b = v('DeviceB', 7)
                bparts = device_b.split() if device_b else []
                z_hostname = bparts[0] if bparts else ''
                z_interface = bparts[1] if len(bparts) > 1 else ''
                rack_b = v('RackB', 8)
                t1l = v('T1 Label', 10)
            if hostname and interface and lbl and re.match(r'\d+[LR]$', lbl):
                t0[(hostname, interface)] = lbl
                t1[(hostname, interface)] = t1l
            if z_hostname and z_interface:
                t1_rev[(z_hostname, z_interface)] = {
                    't0_lbl': lbl, 'source_port': src, 'dmarc1': dmarc1,
                    'dmarc2': dmarc2, 'dest_port': dest_p, 'rack_b': rack_b, 't1_lbl': t1l
                }
                count += 1
        wb.close()
        print(f"  Loaded: {os.path.basename(path)} ({count} T1 reverse entries)")
    return t0, t1, t1_rev

def get_labels(hostname, iface, phys_t0, phys_t1):
    key = (hostname, iface)
    if key in phys_t0:
        return phys_t0[key], phys_t1[key], True
    m = re.match(r'(swp\d+)s(\d+)', str(iface))
    if m:
        base, lane = m.group(1), int(m.group(2))
        partner = {0:1,1:0,2:3,3:2}.get(lane)
        if partner is not None:
            p = (hostname, f"{base}s{partner}")
            if p in phys_t0:
                return phys_t0[p], phys_t1[p], False
    return '', '', False

def get_history_flag(host, iface, current_type, prev_miss, prev_down, prev_opt):
    key = (host, iface)
    if current_type == 'mismatch':
        if key in prev_miss: return "🔁 Recurring mismatch", "FF6B6B"
        if key in prev_down: return "⬆️ Was downlink", "FFB347"
    elif current_type == 'downlink':
        if key in prev_down: return "🔁 Recurring downlink", "FF6B6B"
        if key in prev_opt: return "⚡ Was optic error", "D35400"
        if key in prev_miss: return "⬇️ Was mismatch", "FFB347"
    elif current_type == 'optic':
        if key in prev_opt: return "🔁 Recurring optic", "FF6B6B"
        if key in prev_down: return "⬆️ Was downlink", "FFB347"
        if key in prev_miss: return "⬇️ Was mismatch", "FFB347"
    return "", ""

def get_prev_issues(report_path):
    try:
        wb = load_workbook(report_path, read_only=True)
    except:
        return set(), set(), set(), {}, {}
    prev_miss, prev_down, prev_opt = set(), set(), set()
    prev_rack, prev_opt_rack = {}, {}
    ws = next((wb[n] for n in wb.sheetnames if 'lldp' in n.lower()), None)
    if ws:
        hc = next((c for c in range(1, ws.max_column+1) if str(ws.cell(1,c).value or '').strip() == 'Hostname'), None)
        ic = next((c for c in range(1, ws.max_column+1) if str(ws.cell(1,c).value or '').strip() == 'Interface'), None)
        ac = next((c for c in range(1, ws.max_column+1) if 'Act.' in str(ws.cell(1,c).value or '')), None)
        rc = next((c for c in range(1, ws.max_column+1) if str(ws.cell(1,c).value or '').strip() == 'Rack'), None)
        if hc and ic and ac:
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row: continue
                h = str(row[hc-1] or '').strip()
                i = str(row[ic-1] or '').strip()
                ai = str(row[ac-1] or '').strip().lower()
                rack = str(row[rc-1] or '').strip() if rc else 'Unknown'
                if h and i:
                    if ai == 'interface down':
                        prev_down.add((h, i))
                    elif ai.startswith('swp'):
                        prev_miss.add((h, i))
                    prev_rack[(h, i)] = rack
    ws_opt = next((wb[n] for n in wb.sheetnames if 'optic' in n.lower()), None)
    if ws_opt:
        hc = next((c for c in range(1, ws_opt.max_column+1) if str(ws_opt.cell(1,c).value or '').strip() == 'Hostname'), None)
        ic = next((c for c in range(1, ws_opt.max_column+1) if str(ws_opt.cell(1,c).value or '').strip() == 'Interface'), None)
        rc = next((c for c in range(1, ws_opt.max_column+1) if str(ws_opt.cell(1,c).value or '').strip() == 'Rack'), None)
        if hc and ic:
            for row in ws_opt.iter_rows(min_row=2, values_only=True):
                if not row: continue
                h = str(row[hc-1] or '').strip()
                i = str(row[ic-1] or '').strip()
                rack = str(row[rc-1] or '').strip() if rc else 'Unknown'
                if h and i:
                    prev_opt.add((h, i))
                    prev_opt_rack[(h, i)] = rack
    wb.close()
    return prev_miss, prev_down, prev_opt, prev_rack, prev_opt_rack

# ── Core processing ──────────────────────────────────────────────────────────
def read_lldp_rows(ws_src, phys_t0, phys_t1, t1_rev):
    host_col = next((c for c in range(1, ws_src.max_column+1) if str(ws_src.cell(1,c).value or '').strip() == 'Hostname'), None)
    iface_col = next((c for c in range(1, ws_src.max_column+1) if str(ws_src.cell(1,c).value or '').strip() == 'Interface'), None)
    if not host_col or not iface_col: return []
    rows = []
    for row in range(2, ws_src.max_row + 1):
        host = str(ws_src.cell(row, host_col).value or '').strip()
        iface = str(ws_src.cell(row, iface_col).value or '').strip()
        if not host or not iface: continue
        t0, t1, is_p = get_labels(host, iface, phys_t0, phys_t1)
        act_if_col = next((c for c in range(1, ws_src.max_column+1) if 'Act.' in str(ws_src.cell(1,c).value or '')), None)
        act_if = str(ws_src.cell(row, act_if_col).value or '').strip().lower() if act_if_col else ''
        rtype = 'downlink' if act_if == 'interface down' else ('mismatch' if act_if.startswith('swp') else 'other')
        rows.append({
            't0': t0, 't1': t1, 'is_phys': is_p,
            'row_type': rtype,
            'cells': {
                'Hostname': {'value': host},
                'Interface': {'value': iface},
                'Rack': {'value': ''},
                'Elevation': {'value': ''},
                'Source_port': {'value': ''},
                'DMARC1': {'value': ''},
                'DMARC2': {'value': ''},
                'Destination_port': {'value': ''},
                'Z Interface': {'value': ''},
                'Z Rack': {'value': ''},
                'Z Elevation': {'value': ''},
                'Act. Interface': {'value': act_if},
            },
            'mismatch_info': {}
        })
    return rows

def build_lldp_sheet(wb_out, name, rows, color, is_mismatch=False, prev_miss=None, prev_down=None, prev_opt=None):
    ws = wb_out.create_sheet(name)
    ws.sheet_properties.tabColor = color
    headers = [
        ("Interface", HDR_BG), ("L&R", HDR_BG), ("Rack", HDR_BG), ("Elevation", HDR_BG),
        ("Source_port", "C0504D"), ("DMARC1", "7F6000"), ("DMARC2", "375623"), ("Destination_port", "17375E"),
        ("Z Interface", "17375E"), ("L&R", "17375E"), ("Z Rack", "17375E"), ("Z Elevation", "17375E"),
    ]
    if is_mismatch:
        headers += [("Possible Device A", "833C00"), ("Possible Rack / U", "833C00")]
    headers += [("History", "595959")]
    for col, (label, bg) in enumerate(headers, 1):
        c = ws.cell(1, col, label)
        c.fill = fill(bg)
        c.font = Font(bold=True, color=WHITE, name="Arial", size=9)
        c.alignment = center()
        ws.column_dimensions[get_column_letter(col)].width = 14 if col > 4 else 10
    ws.row_dimensions[1].height = 20
    ws.freeze_panes = "A2"
    for out_row, rd in enumerate(rows, 2):
        p = rd['is_phys']
        row_bg = YELLOW if p else LOG_BG
        hist, hist_col = get_history_flag(
            rd['cells']['Hostname']['value'], rd['cells']['Interface']['value'],
            rd['row_type'], prev_miss or set(), prev_down or set(), prev_opt or set()
        )
        vals = [
            rd['cells']['Interface']['value'],
            rd['t0'],
            rd['cells']['Rack']['value'],
            rd['cells']['Elevation']['value'],
            rd['cells']['Source_port']['value'],
            rd['cells']['DMARC1']['value'],
            rd['cells']['DMARC2']['value'],
            rd['cells']['Destination_port']['value'],
            rd['cells']['Z Interface']['value'],
            rd['t1'],
            rd['cells']['Z Rack']['value'],
            rd['cells']['Z Elevation']['value'],
        ]
        if is_mismatch:
            vals += ["", ""]
        vals += [hist]
        for col, v in enumerate(vals, 1):
            c = ws.cell(out_row, col, v)
            c.fill = fill(row_bg)
            c.font = font(bold=(col == 2 or (col == len(vals) and hist)), sz=8)
            c.alignment = center()
        ws.row_dimensions[out_row].height = 15

def build_optics_fec(wb_out, ws_src, name, color, downlink_set, phys_t0, phys_t1):
    if not ws_src: return 0
    ws = wb_out.create_sheet(name)
    ws.sheet_properties.tabColor = color
    host_col = next((c for c in range(1, ws_src.max_column+1) if str(ws_src.cell(1,c).value or '').strip() == 'Hostname'), 1)
    iface_col = next((c for c in range(1, ws_src.max_column+1) if str(ws_src.cell(1,c).value or '').strip() == 'Interface'), 2)
    headers = [("Interface", HDR_BG), ("L&R", HDR_BG), ("Rack", HDR_BG), ("Elevation", HDR_BG),
               ("Source_port", "C0504D"), ("DMARC1", "7F6000"), ("Z Interface", "17375E"),
               ("DL Flag", "595959"), ("History", "595959")]
    for col, (label, bg) in enumerate(headers, 1):
        c = ws.cell(1, col, label)
        c.fill = fill(bg)
        c.font = Font(bold=True, color=WHITE, name="Arial", size=9)
        c.alignment = center()
        ws.column_dimensions[get_column_letter(col)].width = 14
    ws.row_dimensions[1].height = 20
    ws.freeze_panes = "A2"
    out_row = 2
    for r in range(2, ws_src.max_row + 1):
        host = str(ws_src.cell(r, host_col).value or '').strip()
        iface = str(ws_src.cell(r, iface_col).value or '').strip()
        if not host or not iface: continue
        t0, t1, is_p = get_labels(host, iface, phys_t0, phys_t1)
        is_dl = (host, iface) in downlink_set
        row_bg = "C8C8C8" if is_dl else ("FFFFFF" if is_p else LOG_BG)
        vals = [iface, t0, "", "", "", "", "", "⬇️ Also Downlink — skip" if is_dl else "", ""]
        for col, v in enumerate(vals, 1):
            c = ws.cell(out_row, col, v)
            c.fill = fill(row_bg)
            c.font = Font(name="Arial", size=9, color="888888" if is_dl else "000000")
            c.alignment = center()
        out_row += 1
    return out_row - 2

def build_summary(wb_out, report_name, miss, down, opt, fec, prev_miss, prev_down, prev_opt):
    ws = wb_out.create_sheet("Summary", 0)
    ws.sheet_properties.tabColor = "1F4E79"
    ws.merge_cells("B1:F1")
    c = ws["B1"]
    c.value = "VALIDATION REPORT — SUMMARY"
    c.fill = fill("1F4E79")
    c.font = Font(bold=True, color=WHITE, name="Arial", size=14)
    c.alignment = center()
    ws.row_dimensions[1].height = 28
    ws.merge_cells("B2:F2")
    c = ws["B2"]
    c.value = f"Report: {report_name}  |  Generated: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    c.fill = fill("0D7377")
    c.font = Font(italic=True, color=WHITE, name="Arial", size=9)
    c.alignment = center()
    total = len(miss) + len(down) + opt + fec
    kpis = [("TOTAL", total, "1F4E79"), ("MISMATCHES", len(miss), "C00000"),
            ("DOWNLINKS", len(down), "ED7D31"), ("OPTICS", opt, "833C00"), ("FEC", fec, "7030A0")]
    for i, (lbl, val, bg) in enumerate(kpis):
        col = i + 2
        ws.cell(4, col, lbl).fill = fill(bg)
        ws.cell(4, col).font = Font(bold=True, color=WHITE, name="Arial", size=8)
        ws.cell(4, col).alignment = center()
        ws.cell(5, col, val).fill = fill(bg)
        ws.cell(5, col).font = Font(bold=True, color=WHITE, name="Arial", size=18)
        ws.cell(5, col).alignment = center()
    ws.column_dimensions['B'].width = 14
    for c in "CDEF": ws.column_dimensions[c].width = 14

# ── Main UI ──────────────────────────────────────────────────────────────────
st.markdown("Upload cutsheet(s) + current report. Optional: previous highlighted report for recurring detection.")

cutsheet_files = st.file_uploader("Cutsheet(s) — multiple allowed", type=["xlsx"], accept_multiple_files=True, key="slack_cutsheets")
report_file = st.file_uploader("Current Slack Validation Report (.xlsx)", type=["xlsx"], key="slack_report")
prev_file = st.file_uploader("Previous Highlighted Report (optional)", type=["xlsx"], key="slack_prev")

if st.button("🚀 Process & Highlight Report", type="primary", disabled=not (cutsheet_files and report_file)):
    temp_dir = tempfile.mkdtemp()
    log_capture = io.StringIO()
    try:
        with contextlib.redirect_stdout(log_capture):
            print("Loading cutsheets...")
            cutsheet_paths = []
            for i, f in enumerate(cutsheet_files):
                p = os.path.join(temp_dir, f"cutsheet_{i}.xlsx")
                with open(p, "wb") as out: out.write(f.getbuffer())
                cutsheet_paths.append(p)
            phys_t0, phys_t1, t1_rev = build_lookup(cutsheet_paths)
            prev_miss, prev_down, prev_opt, prev_rack, prev_opt_rack = set(), set(), set(), {}, {}
            if prev_file:
                prev_path = os.path.join(temp_dir, "prev.xlsx")
                with open(prev_path, "wb") as out: out.write(prev_file.getbuffer())
                prev_miss, prev_down, prev_opt, prev_rack, prev_opt_rack = get_prev_issues(prev_path)
            report_path = os.path.join(temp_dir, "report.xlsx")
            with open(report_path, "wb") as out: out.write(report_file.getbuffer())
            wb_src = load_workbook(report_path)
            def find_sheet(wb, *pats):
                for n in wb.sheetnames:
                    for p in pats:
                        if p.lower() in n.lower(): return wb[n]
                return None
            ws_lldp = find_sheet(wb_src, 'lldp')
            ws_opt = find_sheet(wb_src, 'optic')
            ws_fec = find_sheet(wb_src, 'fec')
            wb_out = Workbook()
            wb_out.remove(wb_out.active)
            lldp_rows = read_lldp_rows(ws_lldp, phys_t0, phys_t1, t1_rev) if ws_lldp else []
            miss_rows = [r for r in lldp_rows if r['row_type'] == 'mismatch']
            down_rows = [r for r in lldp_rows if r['row_type'] == 'downlink']
            downlink_set = {(r['cells']['Hostname']['value'], r['cells']['Interface']['value']) for r in down_rows}
            if miss_rows:
                build_lldp_sheet(wb_out, "Mispatches", miss_rows, TAB_MISS, is_mismatch=True, prev_miss=prev_miss, prev_down=prev_down, prev_opt=prev_opt)
            if down_rows:
                build_lldp_sheet(wb_out, "Downlinks", down_rows, TAB_DOWN, prev_miss=prev_miss, prev_down=prev_down, prev_opt=prev_opt)
            n_opt = build_optics_fec(wb_out, ws_opt, "Optics", TAB_OPT, downlink_set, phys_t0, phys_t1)
            n_fec = build_optics_fec(wb_out, ws_fec, "FEC Errors", TAB_FEC, downlink_set, phys_t0, phys_t1)
            build_summary(wb_out, report_file.name, miss_rows, down_rows, n_opt, n_fec, prev_miss, prev_down, prev_opt)
            buf = io.BytesIO()
            wb_out.save(buf)
            buf.seek(0)
            output_bytes = buf.getvalue()
        st.success("✅ Report processed successfully!")
        st.download_button(
            "📥 Download Highlighted Report",
            data=output_bytes,
            file_name=report_file.name.replace(".xlsx", "_highlighted.xlsx"),
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )
        with st.expander("Processing Log"):
            st.text(log_capture.getvalue())
    finally:
        shutil.rmtree(temp_dir, ignore_errors=True)

st.caption("Core features preserved: cutsheet lookup, L&R labels, history/recurring flags, DL overlap detection.")