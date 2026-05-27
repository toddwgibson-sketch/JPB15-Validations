import streamlit as st
from pathlib import Path
import tempfile
import shutil
import io
import contextlib
import os
import re
from datetime import datetime
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Consistent header ────────────────────────────────────────────────────────
def show_header(title: str, subtitle: str = ""):
    col1, col2 = st.columns([1, 8])
    with col1:
        logo_path = Path(__file__).parent.parent / "assets" / "logo.png"
        if logo_path.exists():
            st.image(str(logo_path), width=70)
        else:
            st.markdown("### 🔧")
    with col2:
        st.markdown(f"### {title}")
        if subtitle:
            st.caption(subtitle)

st.set_page_config(page_title="QFAB / T1→T0 Validator", page_icon="🔗", layout="wide")
show_header("QFAB / T1→T0 Validator", "New LV Portal format → DG19-style report")

# ── Style constants ──────────────────────────────────────────────────────────
TAB_SUMM = "1F4E79"
TAB_DOWN = "ED7D31"
TAB_OPT  = "833C00"
TAB_FEC  = "7030A0"
HDR_NAVY = "1F4E79"
WHITE = "FFFFFF"
MISS_BG = "FFF2CC"
DL_GREY_BG = "C8C8C8"

def fill(hex_color): return PatternFill("solid", fgColor=hex_color)
def center(): return Alignment(horizontal="center", vertical="center")

RACK_RE = re.compile(r"Rack\s+(\S+)\s+U\s*(\S+)", re.IGNORECASE)
IFACE_RE = re.compile(r"swp(\d+)s(\d+)", re.IGNORECASE)

def parse_rack_u(s):
    if not s: return "", ""
    m = RACK_RE.match(str(s).strip())
    return (m.group(1), m.group(2)) if m else ("", "")

def iface_to_lr(iface):
    if not iface: return ""
    m = IFACE_RE.match(str(iface).strip())
    if not m: return ""
    port = int(m.group(1))
    lane = int(m.group(2))
    even_port = (port % 2 == 0)
    low_lane = (lane <= 1)
    side = "L" if even_port == low_lane else "R"
    return f"{port}{side}"

def pair_search_order(iface):
    if not iface: return []
    m = IFACE_RE.match(str(iface).strip())
    if not m: return [iface]
    port = int(m.group(1))
    lane = int(m.group(2))
    orderings = {0: [0, 1], 1: [1, 2, 0], 2: [2, 1, 3], 3: [3, 2]}
    return [f"swp{port}s{n}" for n in orderings.get(lane, [lane])]

# ── Cutsheet loader ──────────────────────────────────────────────────────────
def build_lookup(cutsheet_paths):
    forward = {}
    reverse = {}
    for path in cutsheet_paths:
        try:
            wb = load_workbook(path, data_only=True)
        except Exception as e:
            st.warning(f"Could not load {os.path.basename(path)}: {e}")
            continue
        for ws in wb.worksheets:
            if ws.max_row < 2: continue
            headers = {}
            for col in range(1, ws.max_column + 1):
                h = ws.cell(1, col).value
                if isinstance(h, str):
                    headers[h.strip().lower()] = col
            new_layout = ('Hostname' in headers and 'Interface' in headers)
            def read_str(row, name, legacy_idx=None):
                idx = headers.get(name)
                if idx and idx <= len(row):
                    v = row[idx-1]
                    return str(v).strip() if v is not None else ""
                if legacy_idx and legacy_idx < len(row):
                    v = row[legacy_idx-1]
                    return str(v).strip() if v is not None else ""
                return ""
            for row_vals in ws.iter_rows(min_row=2, values_only=True):
                if not row_vals or all(v is None for v in row_vals): continue
                if new_layout:
                    hostname = read_str(row_vals, 'hostname')
                    interface = read_str(row_vals, 'interface')
                    rack_a, elev_a = parse_rack_u(read_str(row_vals, 'rack'))
                    src = read_str(row_vals, 'source_port')
                    dm1 = read_str(row_vals, 'dmarc1')
                    dm2 = read_str(row_vals, 'dmarc2')
                    dest = read_str(row_vals, 'destination_port')
                    z_host = read_str(row_vals, 'z hostname')
                    z_iface = read_str(row_vals, 'z interface')
                    z_rack, z_elev = parse_rack_u(read_str(row_vals, 'z rack'))
                else:
                    dev_a = read_str(row_vals, 'devicea', 2)
                    parts = dev_a.split()
                    hostname = parts[0] if parts else ""
                    interface = parts[1] if len(parts) > 1 else ""
                    rack_a, elev_a = parse_rack_u(read_str(row_vals, 'racka', 3))
                    src = read_str(row_vals, 'source_port', 4)
                    dm1 = read_str(row_vals, 'dmarc1', 5)
                    dm2 = read_str(row_vals, 'dmarc2', 6)
                    dest = read_str(row_vals, 'destination_port', 7)
                    dev_b = read_str(row_vals, 'deviceb', 8)
                    bparts = dev_b.split()
                    z_host = bparts[0] if bparts else ""
                    z_iface = bparts[1] if len(bparts) > 1 else ""
                    z_rack, z_elev = parse_rack_u(read_str(row_vals, 'rackb', 9))
                if not hostname or not interface: continue
                forward[(hostname.lower(), interface.lower())] = {
                    "rack_a": rack_a, "elev_a": elev_a,
                    "src_pp": src, "dmarc1": dm1, "dmarc2": dm2, "dest_pp": dest,
                    "dev_b": z_host, "iface_b": z_iface,
                    "rack_b": z_rack, "elev_b": z_elev,
                }
                if z_host and z_iface:
                    reverse[(z_host.lower(), z_iface.lower())] = {"rack": z_rack, "elev": z_elev}
        wb.close()
    return forward, reverse

def lookup_or_matrix(lookup, hostname, iface, matrix, remote_dev="", remote_port=""):
    forward, reverse = lookup
    host_lower = hostname.lower().strip()
    iface_lower = iface.lower().strip()
    if (host_lower, iface_lower) in forward:
        info = dict(forward[(host_lower, iface_lower)])
        info["cutsheet_miss"] = False
        return info
    for cand in pair_search_order(iface_lower)[1:]:
        if (host_lower, cand) in forward:
            pair = forward[(host_lower, cand)]
            z_dev = remote_dev.strip()
            z_iface = remote_port.strip()
            z_rack, z_elev = "", ""
            if z_dev and z_iface:
                for zc in pair_search_order(z_iface.lower()):
                    r = reverse.get((z_dev.lower(), zc))
                    if r:
                        z_rack, z_elev = r["rack"], r["elev"]
                        break
            return {
                "rack_a": pair["rack_a"], "elev_a": pair["elev_a"],
                "src_pp": pair["src_pp"], "dmarc1": pair["dmarc1"],
                "dmarc2": pair["dmarc2"], "dest_pp": pair["dest_pp"],
                "dev_b": z_dev, "iface_b": z_iface,
                "rack_b": z_rack, "elev_b": z_elev,
                "cutsheet_miss": False,
            }
    return {
        "rack_a": "", "elev_a": "", "src_pp": "", "dmarc1": "", "dmarc2": "",
        "dest_pp": "", "dev_b": "", "iface_b": "", "rack_b": "", "elev_b": "",
        "cutsheet_miss": True,
    }

# ── Sheet builders ───────────────────────────────────────────────────────────
def write_header(ws, headers, widths):
    for col, (text, bg) in enumerate(headers, 1):
        c = ws.cell(1, col, text)
        c.fill = fill(bg)
        c.font = Font(bold=True, color=WHITE, name="Arial", size=9)
        c.alignment = center()
        ws.column_dimensions[get_column_letter(col)].width = widths[col-1]
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"

def build_downlinks(wb_out, src_ws, lookup):
    ws = wb_out.create_sheet("Downlinks")
    ws.sheet_properties.tabColor = TAB_DOWN
    headers = [
        ("Interface", HDR_NAVY), ("L&R", HDR_NAVY), ("Rack", HDR_NAVY), ("Elevation", HDR_NAVY),
        ("Source_port", "C0504D"), ("DMARC1", "7F6000"), ("DMARC2", "375623"), ("Destination_port", "17375E"),
        ("Z Interface", "17375E"), ("L&R", "17375E"), ("Z Rack", "17375E"), ("Z Elevation", "17375E"),
        ("History", "595959"),
    ]
    write_header(ws, headers, [12,6,8,8,32,32,32,32,12,6,8,10,14])
    if not src_ws or src_ws.max_row < 2: return 0, set()
    hmap = {str(src_ws.cell(1,c).value or '').strip().lower(): c for c in range(1, src_ws.max_column+1)}
    def gv(r, *names):
        for n in names:
            c = hmap.get(n.lower())
            if c: return str(src_ws.cell(r, c).value or '').strip()
        return ""
    downlink_set = set()
    out_row = 2
    for r in range(2, src_ws.max_row + 1):
        host = gv(r, "Source Device Name")
        port = gv(r, "Source Device Port")
        if not host or not port: continue
        downlink_set.add((host.lower(), port.lower()))
        rem_dev = gv(r, "Remote Device Name")
        rem_port = gv(r, "Remote Device Port")
        matrix = gv(r, "Patch Panel Matrix")
        info = lookup_or_matrix(lookup, host, port, matrix, rem_dev, rem_port)
        miss = info.get("cutsheet_miss", False)
        row_bg = MISS_BG if miss else "FFFFFF"
        vals = [
            port, iface_to_lr(port), info.get("rack_a",""), info.get("elev_a",""),
            info.get("src_pp",""), info.get("dmarc1",""), info.get("dmarc2",""), info.get("dest_pp",""),
            info.get("iface_b") or rem_port, iface_to_lr(rem_port or ""),
            info.get("rack_b",""), info.get("elev_b",""),
            "⚠ Not in cutsheet" if miss else "",
        ]
        for col, v in enumerate(vals, 1):
            c = ws.cell(out_row, col, v)
            c.fill = fill(row_bg)
            c.font = Font(name="Arial", size=9, bold=(col==2))
            c.alignment = center()
        out_row += 1
    return out_row-2, downlink_set

def build_optics(wb_out, src_ws, lookup, downlink_set):
    ws = wb_out.create_sheet("Optics")
    ws.sheet_properties.tabColor = TAB_OPT
    headers = [
        ("Interface", HDR_NAVY), ("L&R", HDR_NAVY), ("Rack", HDR_NAVY), ("Elevation", HDR_NAVY),
        ("Channel", HDR_NAVY), ("Measured (dBm)", HDR_NAVY),
        ("Source_port", HDR_NAVY), ("DMARC1", HDR_NAVY), ("DMARC2", HDR_NAVY), ("Destination_port", HDR_NAVY),
        ("Z Interface", HDR_NAVY), ("Z L&R", HDR_NAVY), ("Z Rack", HDR_NAVY), ("Z Elevation", HDR_NAVY),
        ("DL Flag", "595959"), ("History", "595959"),
    ]
    write_header(ws, headers, [12,6,8,8,8,12,32,32,32,32,12,6,8,10,22,14])
    if not src_ws or src_ws.max_row < 2: return 0
    out_row = 2
    for r in range(2, src_ws.max_row + 1):
        host = str(src_ws.cell(r, 3).value or '').strip()
        port = str(src_ws.cell(r, 4).value or '').strip()
        if not host or not port: continue
        is_dl = (host.lower(), port.lower()) in downlink_set
        row_bg = DL_GREY_BG if is_dl else "FFFFFF"
        for ch in ["1","2","3","4"]:
            vals = [port, "", "", "", ch, "", "", "", "", "", "", "", "", "", "⬇️ Also Downlink" if is_dl else "", ""]
            for col, v in enumerate(vals, 1):
                c = ws.cell(out_row, col, v)
                c.fill = fill(row_bg)
                c.font = Font(name="Arial", size=9, color="888888" if is_dl else "000000")
                c.alignment = center()
            out_row += 1
    return out_row-2

def build_fec(wb_out, src_ws, lookup, downlink_set):
    ws = wb_out.create_sheet("FEC Errors")
    ws.sheet_properties.tabColor = TAB_FEC
    headers = [("Interface", HDR_NAVY), ("L&R", HDR_NAVY), ("Rack", HDR_NAVY), ("Elevation", HDR_NAVY),
               ("Pre-FEC BER", HDR_NAVY), ("Source_port", HDR_NAVY), ("DMARC1", HDR_NAVY),
               ("Z Interface", HDR_NAVY), ("DL Flag", "595959"), ("History", "595959")]
    write_header(ws, headers, [12,6,8,8,14,32,32,12,22,14])
    if not src_ws or src_ws.max_row < 2: return 0
    out_row = 2
    for r in range(2, src_ws.max_row + 1):
        host = str(src_ws.cell(r, 1).value or '').strip()
        port = str(src_ws.cell(r, 2).value or '').strip()
        if not host or not port: continue
        is_dl = (host.lower(), port.lower()) in downlink_set
        row_bg = DL_GREY_BG if is_dl else "FFFFFF"
        vals = [port, "", "", "", "", "", "", "", "⬇️ Also Downlink" if is_dl else "", ""]
        for col, v in enumerate(vals, 1):
            c = ws.cell(out_row, col, v)
            c.fill = fill(row_bg)
            c.font = Font(name="Arial", size=9, color="888888" if is_dl else "000000")
            c.alignment = center()
        out_row += 1
    return out_row-2

def build_summary(wb_out, report_name, n_down, n_opt, n_fec):
    ws = wb_out.create_sheet("Summary", 0)
    ws.sheet_properties.tabColor = TAB_SUMM
    ws.merge_cells("B1:E1")
    c = ws["B1"]
    c.value = "VALIDATION REPORT — SUMMARY"
    c.fill = fill("1F4E79")
    c.font = Font(bold=True, color=WHITE, name="Arial", size=14)
    c.alignment = center()
    ws.row_dimensions[1].height = 28
    ws.merge_cells("B2:E2")
    c = ws["B2"]
    c.value = f"Report: {report_name}  |  Generated: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    c.fill = fill("0D7377")
    c.font = Font(italic=True, color=WHITE, name="Arial", size=9)
    c.alignment = center()
    total = n_down + n_opt + n_fec
    for i, (lbl, val, bg) in enumerate([
        ("TOTAL ISSUES", total, "1F4E79"),
        ("DOWNLINKS", n_down, "ED7D31"),
        ("OPTICS (channels)", n_opt, "833C00"),
        ("FEC ERRORS", n_fec, "7030A0"),
    ]):
        col = i+2
        ws.cell(4, col, lbl).fill = fill(bg)
        ws.cell(4, col).font = Font(bold=True, color=WHITE, name="Arial", size=9)
        ws.cell(4, col).alignment = center()
        ws.cell(5, col, val).fill = fill(bg)
        ws.cell(5, col).font = Font(bold=True, color=WHITE, name="Arial", size=18)
        ws.cell(5, col).alignment = center()
    ws.column_dimensions['B'].width = 22
    for c in "CDE": ws.column_dimensions[c].width = 18

# ── Main processing UI ───────────────────────────────────────────────────────
st.markdown("Upload cutsheet(s) and the LV Portal report below.")

cutsheet_files = st.file_uploader("Cutsheet(s) — multiple allowed (new or legacy layout)", type=["xlsx"], accept_multiple_files=True, key="qfab_cutsheets")
report_file = st.file_uploader("LV Portal Validation Export (.xlsx)", type=["xlsx"], key="qfab_report")

if st.button("🚀 Process Report", type="primary", disabled=not (cutsheet_files and report_file), key="qfab_process"):
    temp_dir = tempfile.mkdtemp()
    log_capture = io.StringIO()
    try:
        with contextlib.redirect_stdout(log_capture):
            print("Loading cutsheets...")
            cutsheet_paths = []
            for i, f in enumerate(cutsheet_files):
                p = os.path.join(temp_dir, f"cutsheet_{i}.xlsx")
                with open(p, "wb") as out: out.write(f.getbuffer())
                cutsheet_paths.append(p)
            lookup = build_lookup(cutsheet_paths)
            print(f"Indexed {len(lookup[0])} forward entries")
            report_path = os.path.join(temp_dir, "report.xlsx")
            with open(report_path, "wb") as out: out.write(report_file.getbuffer())
            wb_src = load_workbook(report_path, data_only=True)
            def find_sheet(wb, *pats):
                for n in wb.sheetnames:
                    for p in pats:
                        if p.lower() in n.lower(): return wb[n]
                return None
            ws_down = find_sheet(wb_src, "interface down", "downlink")
            ws_opt = find_sheet(wb_src, "optic")
            ws_fec = find_sheet(wb_src, "fec")
            wb_out = Workbook()
            wb_out.remove(wb_out.active)
            n_down, dl_set = build_downlinks(wb_out, ws_down, lookup)
            n_opt = build_optics(wb_out, ws_opt, lookup, dl_set)
            n_fec = build_fec(wb_out, ws_fec, lookup, dl_set)
            build_summary(wb_out, report_file.name, n_down, n_opt, n_fec)
            buf = io.BytesIO()
            wb_out.save(buf)
            buf.seek(0)
            output_bytes = buf.getvalue()
        st.success("✅ Processing complete!")
        st.download_button(
            "📥 Download Formatted Report",
            data=output_bytes,
            file_name=report_file.name.replace(".xlsx", "_formatted.xlsx"),
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )
        with st.expander("Processing Log"):
            st.text(log_capture.getvalue())
    finally:
        shutil.rmtree(temp_dir, ignore_errors=True)

st.caption("This version preserves the core cutsheet fallback and logical-pair logic from the original script.")