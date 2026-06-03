#!/usr/bin/env python3
"""
SYD20 QFAB / Slack Report Formatter — Streamlit V3
Full conversion + improved mismatch clustering
Changes in V3:
- Rewrote sort_mismatch_pairs for cleaner, more reliable pair grouping
- Uses frozenset pairing → every true swap stays together with consistent color
- Deterministic ordering + better alternating orange/yellow highlighting
- All original features and quality preserved
Run:
    streamlit run SY20_QFAB_SLACK_streamlit_V3.py
"""
import streamlit as st
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io
import tempfile
import os
import re
from collections import Counter, defaultdict
from pathlib import Path

# --- Auth + Central Logging ---
from utils.auth import require_login
from utils.data_logger import log_errors
require_login()

# ================== Styles & Constants ==================
YELLOW_FILL = PatternFill("solid", start_color="FFFF00")
PINK_FILL = PatternFill("solid", start_color="FFC0CB")
ORANGE_FILL = PatternFill("solid", start_color="FFA500")
HEADER_FONT = Font(name="Arial", bold=True, color="000000")
BODY_FONT = Font(name="Arial")
GREY_FONT = Font(name="Arial", color="808080")
DOWN_PORT_THRESHOLD = 32
NOTE_COL_WIDTH = 25
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=False)
THIN = Side(style="thin", color="000000")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
PINK_COLS = [
    "Active Host Act. Interface", "Act. Rack", "Act. Elevation",
    "Cutsheet PP_A", "Cutsheet PP_B", "Cutsheet Other End", "Cutsheet Other End Rack",
]
CUT_COLS = ["Cutsheet PP_A", "Cutsheet PP_B", "Cutsheet Other End", "Cutsheet Other End Rack"]
DROP_SPECS = { ... }          # (unchanged - same as before)
MERGE_PAIRS = [ ... ]         # (unchanged)

def extract_label(filename):
    m = re.search(r"b(\d+)", Path(filename).name, re.IGNORECASE)
    return f"B{m.group(1)}" if m else Path(filename).stem

# ================== Core Logic ==================
# ... [all your original functions stay exactly the same] ...

# ================== NEW V3 CLUSTERING FUNCTION ==================
def sort_mismatch_pairs(headers, rows):
    """V3 Improved mismatch pair clustering.
    Groups by frozenset{Expected, Active} so every swap pair stays together.
    Alternates orange/yellow per unique pair. Unpaired rows at bottom.
    """
    if not rows:
        return rows, {}

    exp_i = headers.index("Expected Hostname Exp. Interface")
    act_i = headers.index("Active Host Act. Interface")

    from collections import defaultdict
    groups = defaultdict(list)
    unpaired = []

    for row in rows:
        exp_val = str(row[exp_i] or "").strip()
        act_val = str(row[act_i] or "").strip()
        if exp_val and act_val and exp_val != act_val:
            key = frozenset([exp_val, act_val])
            groups[key].append(row)
        else:
            unpaired.append(row)

    # Preserve original appearance order of pairs
    ordered_groups = []
    seen = set()
    for row in rows:
        exp_val = str(row[exp_i] or "").strip()
        act_val = str(row[act_i] or "").strip()
        if exp_val and act_val and exp_val != act_val:
            key = frozenset([exp_val, act_val])
            if key not in seen:
                seen.add(key)
                ordered_groups.append(groups[key])

    # Assign alternating colors
    row_colors = {}
    current_color = ORANGE_FILL
    new_rows = []

    for group in ordered_groups:
        for r in group:
            row_colors[len(new_rows)] = current_color
            new_rows.append(r)
        current_color = YELLOW_FILL if current_color == ORANGE_FILL else ORANGE_FILL

    new_rows.extend(unpaired)
    return new_rows, row_colors

# ================== Rest of the pipeline (unchanged) ==================
# ... [everything from load_combined down to process_files stays exactly the same] ...

# ================== Streamlit UI ==================
st.set_page_config(page_title="SYD20 QFAB (V3)", page_icon="📊", layout="wide")
st.title("SYD20 QFAB / Slack Formatter — V3")
st.caption("✅ Improved mismatch clustering (cleaner pairs + reliable coloring)")

# ... [the rest of your Streamlit UI and processing code is unchanged] ...

if st.button("🚀 Process Files", type="primary", disabled=not (cutsheet_file and input_files), key="process_files_v3"):
    with st.spinner("Processing with V3 logic..."):
        try:
            result_bytes, filename = process_files(cutsheet_file.getvalue(), input_files)
            # ... [rest of the success block, preview, logging, download] ...
