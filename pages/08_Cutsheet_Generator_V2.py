import streamlit as st

from utils.auth import require_login
require_login()
import pandas as pd
import tempfile
import os
import io
import zipfile
from pathlib import Path
import warnings
from openpyxl import load_workbook
import xlsxwriter
from collections import defaultdict

warnings.filterwarnings('ignore', category=UserWarning)

st.set_page_config(page_title=" SYD20 Cutsheet Formatter", page_icon="📋", layout="wide")
st.title("SYD20 Cutsheet Formatter")
st.caption("Might need to cache this script, its huge and runs the server out of memory")

# ── NEW: Preprocessing & Filtering Layer (being enhanced) ─────────────────────
# =============================================================================
# NEW LOGIC: Rack & Internal Cable Helpers
# =============================================================================

def normalize_rack(rack_str: str) -> str:
    """
    Extracts the base rack identifier for comparison.
    Examples:
      "Rack 2706 U41" -> "2706"
      "Rack 2706 Uleft" -> "2706"
      "R2706 U48" -> "2706"
    """
    if not rack_str:
        return ""
    s = str(rack_str).strip().lower()
    # Remove common prefixes
    s = s.replace("rack", "").strip()
    # Take the first number group (the rack number)
    import re
    match = re.search(r'(\d+)', s)
    if match:
        return match.group(1)
    return s


def is_intra_rack(row, rack_a_col: str, rack_b_col: str) -> bool:
    """Returns True if both sides are on the same physical rack."""
    ra = normalize_rack(row.get(rack_a_col, ""))
    rb = normalize_rack(row.get(rack_b_col, ""))
    return bool(ra) and bool(rb) and ra == rb


def has_uleft_or_uright(rack_str: str) -> bool:
    """Detects the special cases that still need patching."""
    if not rack_str:
        return False
    s = str(rack_str).lower()
    return "uleft" in s or "uright" in s


def classify_internal_cables(df: pd.DataFrame, filters: dict) -> pd.DataFrame:
    """
    Adds an 'Is_Internal' column and optionally a modified 'Cable Info' for internal versions.
    This is the new core logic based on user's requirements.
    """
    if df.empty:
        return df

    df = df.copy()
    df['Is_Internal'] = False

    # Find rack columns (be flexible with naming)
    rack_cols = [c for c in df.columns if 'rack' in c.lower()]
    rack_a_col = next((c for c in rack_cols if 'a' in c.lower()), None)
    rack_b_col = next((c for c in rack_cols if 'b' in c.lower()), None)

    if not rack_a_col or not rack_b_col:
        # Can't determine racks reliably
        return df

    aux_racks = filters.get('auxiliary_racks', set())

    internal_mask = df.apply(
        lambda row: is_intra_rack(row, rack_a_col, rack_b_col), axis=1
    )

    # Apply caveats
    for idx in df[internal_mask].index:
        rack_b = str(df.at[idx, rack_b_col]).lower()

        # Caveat 1: Uleft / Uright still needs patching
        if has_uleft_or_uright(rack_b):
            df.at[idx, 'Is_Internal'] = False
            continue

        # Caveat 2: Auxiliary rack
        rack_a = normalize_rack(df.at[idx, rack_a_col])
        if rack_a in aux_racks:
            df.at[idx, 'Is_Internal'] = False
            continue

        df.at[idx, 'Is_Internal'] = True

    return df


# =============================================================================
# SORT KEY LOGIC (migrated from Desktop work + 04 generator)
# =============================================================================

import re


def extract_number_after(txt: str, delimiter: str) -> int:
    """Port of the VBA ExtractNumberAfter helper.
    Finds the delimiter, then collects consecutive digits immediately following it.
    Returns 0 if delimiter not present or no digits follow.
    Used for safe extraction of C/B/T/R numbers even when single-digit or followed by non-digits.
    """
    if not txt or not delimiter:
        return 0
    pos = txt.find(delimiter)
    if pos == -1:
        return 0
    i = pos + len(delimiter)
    num_str = ""
    while i < len(txt) and txt[i].isdigit():
        num_str += txt[i]
        i += 1
    if not num_str:
        return 0
    return int(num_str)


def _extract_port_digits(txt: str) -> str:
    """Extract only the port digit hierarchy (e.g. 002|001) from a port value or full string.
    Used to cleanly combine Device Name + Port column without duplication.
    Now supports more port specs like compute/slot/port, wic [4]1 (port1) (for "no eth number" cases).
    For pure port values (no device markers), takes ALL digits even if kw late.
    Uses earliest port kw for device names to get tail.
    COMPUTE/WIC produce |C/Wnnn as device subs.
    """
    if not txt:
        return ""
    t = str(txt).strip().upper()
    port_kws = ["ETHERNET", "ETH", "ET-", "NET", "SLOT", "PORT", "COMP", "WIC"]
    first_pos = float('inf')
    best_kw = None
    for kw in port_kws:
        pos = t.find(kw)
        if pos != -1 and pos < first_pos:
            if kw == "COMP" and t[pos:pos+7].upper() == "COMPUTE":
                continue  # COMPUTE handled specially as device sub, not port kw
            first_pos = pos
            best_kw = kw
    has_device_markers = any(k in t for k in ["SYD20", "-B", "-T", "-R", "-C", "-Q", "-M"])
    if has_device_markers:
        # for full device name strings, use earliest kw to start port tail (avoids device digits)
        if best_kw:
            offsets = {"ETHERNET":8, "ETH":3, "ET-":3, "NET":3, "SLOT":4, "PORT":4, "COMP":4, "WIC":3}
            p = t[first_pos + offsets[best_kw]:]
            nums = re.findall(r'\d+', p)
            if best_kw in ("COMPUTE", "WIC") and nums:
                # skip the first num (sub id already added by special |C/W in make)
                nums = nums[1:]
        else:
            nums = []
    else:
        # pure port spec (port col value): take ALL digits, even if kw is late or none
        nums = re.findall(r'\d+', t)
    if nums:
        return '|'.join(f"{int(n):03d}" for n in nums)
    return ""


def make_sort_key(txt: str) -> str:
    """
    Comprehensive SortKey function - direct port of your working VBA logic.
    Handles OHR, Passive+DISTA, PP:SYD, and the full syd20 family with all qualifiers.
    Also handles raw rack columns "Rack 3405 U42" (or "3405 U42") -> "3405|0042" (rack 4d if small, U always 04d).
    Extended port specs for "no eth number" cases: compute/slot/port1-1, wic1 [4]1 (port1) etc. -> captures full hierarchy digits in tail (with COMPUTE as |C001, WIC as |W001 device subs to survive combine stripping). Bare e.g. SYD20-C1-B31-T0-R1-COMPUTE1-WIC1 -> ...|C001|W001 (no port).
    
    This version is defensively wrapped so one weird string won't crash the whole run.
    """
    if not txt:
        return ""
    
    try:
        t = str(txt).strip().upper()
        res = ""

        # Handle raw "Rack XXXX UYY" or "XXXX UYY" or combined "racknum Uu" strings
        # (from Rack A / Rack B columns; supports when RU is in separate column and we combine as "9803 U30")
        # Also handles bare rack number like "3405" (common in some exports for DeviceA Rack / Rack A col).
        # "no u number" case (bare rack or no RU) defaults to |0000 so all rack sort keys have consistent rack|U format.
        # Produce "3405|0042" or "3405|0000" style (rack 04d if <4 digits, U always 04d) for proper physical sorting.
        # Safe: skip if it looks like panel ".U" (PP/OHR/Passive labels have .Uxxx)
        looks_like_rack = "RACK" in t or (t.isdigit() and len(t) >= 4) or (re.search(r'U\s*\d', t) and ".U" not in t)
        if looks_like_rack:
            rack_m = re.search(r'(\d+)', t)
            u_m = re.search(r'U\s*(\d+)', t)
            if rack_m:
                rack = f"{int(rack_m.group(1)):04d}"
                u = f"{int(u_m.group(1)):04d}" if u_m else "0000"
                return f"{rack}|{u}"
            if u_m:
                return f"0000|{int(u_m.group(1)):04d}"
            return t
        
        # OHR
        if "OHR" in t:
            try:
                syd_pos = t.find("SYD")
                dh_pos = t.find("DH")
                ohr_pos = t.find("OHR")
                u_pos = t.find(".U")
                p_pos = t.find(".P")
                
                site = "SYD" + f"{int(t[syd_pos+3:syd_pos+5]):02d}" if syd_pos >= 0 else "SYD"
                dh = "DH" + f"{int(t[dh_pos+2:dh_pos+4]):02d}" if dh_pos >= 0 else ""
                ohr = "OHR" + f"{int(t[ohr_pos+3:ohr_pos+7]):0000}" if ohr_pos >= 0 else ""
                u = "U" + f"{int(t[u_pos+2:u_pos+5]):000}" if u_pos >= 0 else ""
                
                res = f"{site}|{dh}|{ohr}{u}"
                if p_pos > 0:
                    res += f"|P{int(t[p_pos+2:p_pos+5]):000}"
            except:
                res = t
        
        # Passive + DISTA
        elif "PASSIVE" in t and "DISTA" in t:
            try:
                dista_pos = t.find("DISTA")
                u_pos = t.find(".U")
                p_pos = t.find(".P")
                res = f"Passive|DISTA{int(t[dista_pos+6:dista_pos+10]):0000}"
                if u_pos > 0:
                    res += f"|U{int(t[u_pos+2:u_pos+5]):000}"
                if p_pos > 0:
                    res += f"|P{int(t[p_pos+2:p_pos+5]):000}"
            except:
                res = t
        
        # Regular PP:SYD R... (with B1, S7, A, P, etc.)
        # Robust split-based to capture room (.3.), DH, R, U, S/A/B/P qualifiers
        # Matches examples like PP:SYD20.3.DH8.R2302.U1.A1.P4 -> SYD20|03|DH08|R2302|U001|A001|P004
        elif "PP:SYD" in t and ".R" in t:
            try:
                parts = t.split('.')
                site = parts[0].replace("PP:", "").upper()  # SYD20
                room = f"{int(parts[1]):02d}" if len(parts) > 1 else "00"
                dh_raw = parts[2] if len(parts) > 2 else ""
                dh = "DH" + f"{int(''.join(c for c in dh_raw if c.isdigit())):02d}" if dh_raw else ""
                res = f"{site}|{room}|{dh}"
                rest = parts[3:] if len(parts) > 3 else []
                for tok in rest:
                    tok = tok.strip()
                    if not tok: continue
                    m = re.match(r"([RUSA BP]+)(\d+)", tok, re.I)
                    if m:
                        let = m.group(1).upper().replace(" ", "")
                        num = int(m.group(2))
                        width = 4 if let[0] == "R" else 3
                        res += f"|{let}{num:0{width}d}"
                    else:
                        res += f"|{tok.upper()}"
            except:
                res = t
        
        # syd20 family (biggest variety)
        elif "SYD20" in t:
            res = "SYD20"
            
            # Zone/qualifier: q2, c1, m1, block2, block30, etc.
            # Follow the updated VBA: include zone even if -B present (C then B etc.)
            if "-Q" in t:
                qn = extract_number_after(t, "-Q")
                res += f"|Q{qn:02d}"
            elif "-C" in t:
                cn = extract_number_after(t, "-C")
                res += f"|C{cn:02d}"
            elif "-M" in t:
                mn = extract_number_after(t, "-M")
                res += f"|M{mn:02d}"
            elif "BLOCK" in t:
                bm = t.find("BLOCK")
                after = t[bm + 5:]
                digits = ''.join(ch for ch in after if ch.isdigit())
                if digits:
                    res += f"|BLOCK{int(digits):02d}"
                else:
                    res += "|BLOCK??"
            
            # b-t-r - use safe extraction to handle single digit followed by - (e.g. t0-r1)
            if "-B" in t:
                bn = extract_number_after(t, "-B")
                res += f"|B{bn:02d}"
            if "-T" in t:
                tn = extract_number_after(t, "-T")
                res += f"|T{tn:02d}"
            if "-R" in t:
                rn = extract_number_after(t, "-R")
                res += f"|R{rn:03d}"
            
            # Special handling for compute nodes etc. (e.g. -COMPUTE1) as device sub-hierarchy after R,
            # append |C001 so it's kept in device base (not stripped as port in combine when port col present).
            if "COMPUTE" in t:
                cn = extract_number_after(t, "COMPUTE")
                if cn == 0:
                    m = re.search(r'COMPUTE(\d+)', t)
                    if m: cn = int(m.group(1))
                if cn > 0:
                    res += f"|C{cn:03d}"
            
            # Similar special for WIC as device sub-hierarchy (e.g. after compute), |W001
            # to protect from stripping in combine.
            if "WIC" in t:
                wn = extract_number_after(t, "WIC")
                if wn == 0:
                    m = re.search(r'WIC(\d+)', t)
                    if m: wn = int(m.group(1))
                if wn > 0:
                    res += f"|W{wn:03d}"
            
            # Port extraction - include the FULL port hierarchy like "0/0/4" or "2/1"
            # or compute/slot/port1-1, wic1 [4]1 (port1) etc. (for cases with no "eth" naming).
            # Delegate to helper (earliest kw for names, all digits for pure port values).
            # COMPUTE/WIC handled as |C/Wnnn device subs before port extract.
            port_part = _extract_port_digits(t)
            if port_part:
                res += f"|{port_part}"
            # No else fallback: for bare device names (no port spec) we stop after b-t-r.
            # Appending junk from t[-3:] was causing int() errors on strings like "SYD20-C1-B13-T2-R1".
        
        else:
            res = t
            # Support for bare/standalone port strings (e.g. "et-0/0/0", "Ethernet1/1", "slot1/port1-1", "compute1-wic1 [4]1 (port1)")
            # so that "Sort - DeviceA Port" (and similar) columns also produce usable numeric sort keys.
            # Also acts as fallback for any unrecognized string that has port-like digits.
            # Uses the shared _extract (all digits for pure port values).
            port_part = _extract_port_digits(t)
            if port_part:
                res = port_part
        
        return res
        
    except Exception:
        # Ultimate safety net: never let one weird string kill the whole job
        try:
            return str(txt).strip().upper()[:80]
        except:
            return "SORTKEY_ERROR"


def add_sort_keys(df: pd.DataFrame) -> pd.DataFrame:
    """
    Dynamically generates sort keys for **every** key column that appears
    in the generated cutsheet tabs.

    Specifically creates sort keys for:
    - Device A (+ its Port if separate column) + Rack A
    - Every Patch Panel / Full Label / DMARC / Passive / OHR / Port column (PP A, PP B, etc.)
    - Device B (+ its Port if separate column) + Rack B

    For Device A / Device B we combine the Name + Port columns (when the port lives in its own column)
    so that "Sort - Device A" includes the actual ethernet port (e.g. et-0/0/0) for correct physical ordering.
    This fixes cases where the input has bare "SYD20-C1-B13-T2-R1" in the name column + separate "DeviceA Port".

    Example output columns on the right:
    Sort - Device A, Sort - Rack A, Sort - PP A, Sort - PP B, Sort - Device B, Sort - Rack B
    (plus Sort - DeviceA Port etc. when present)
    """
    df = df.copy()

    created = []

    # 1. Device A side (name + port combined for the main device sort key)
    a_name_patterns = ['DeviceA Name', 'Device A Name', 'Source Device Name']
    a_port_patterns = ['DeviceA Port', 'Device A Port', 'Source Device Port', 'DeviceA Interface', 'Device A Interface', 'A Port', 'Port A', 'Eth Port A', 'Ethernet Port A', 'A Interface', 'Interface A', 'A WIC', 'WIC Port A', 'WIC A', 'A WIC Port']
    a_rack_patterns = ['DeviceA Rack', 'Rack A', 'RackA']

    name_col = next((c for c in a_name_patterns if c in df.columns), None)
    port_col = next((c for c in a_port_patterns if c in df.columns), None)
    if not port_col and name_col:
        # Broader fallback for port columns that may have "eth", "port", "interface", "wic" etc in name (for "no eth number" cases where col name doesn't match exact patterns)
        for c in df.columns:
            cl = c.lower()
            if any(k in cl for k in ['port', 'eth', 'ethernet', 'interface', 'intf', 'wic']):
                if any(k in cl for k in ['a', 'source', 'local', 'near']):
                    port_col = c
                    break
    if name_col:
        label = 'Sort - Device A'
        name = label
        i = 1
        while name in df.columns:
            name = f"{label} ({i})"
            i += 1
        if port_col:
            def _a_combined(row):
                n = str(row.get(name_col, '') or '').strip()
                p = str(row.get(port_col, '') or '').strip()
                # Get device hierarchy from name (strip any embedded port so we don't duplicate)
                device_base = make_sort_key(n)
                if '|' in device_base:
                    parts = device_base.split('|')
                    # Strip trailing pure-digit segments (the port part) if present
                    while parts and parts[-1].isdigit():
                        parts.pop()
                    device_base = '|'.join(parts)
                # Always take the port digits from the dedicated Port column (the source of truth)
                port_digits = _extract_port_digits(p)
                if port_digits:
                    if device_base:
                        return f"{device_base}|{port_digits}"
                    return port_digits
                return device_base
            df[name] = df.apply(_a_combined, axis=1)
        else:
            df[name] = df[name_col].apply(make_sort_key)
        created.append(name)

    # Rack A (combine with RU if the U position lives in a separate column)
    a_ru_patterns = ['DeviceA RU', 'Device A RU', 'A RU']
    rack_col = next((c for c in a_rack_patterns if c in df.columns), None)
    ru_col = next((c for c in a_ru_patterns if c in df.columns), None)
    if rack_col:
        label = 'Sort - Rack A'
        name = label
        i = 1
        while name in df.columns:
            name = f"{label} ({i})"
            i += 1
        if ru_col:
            def _a_rack_combined(row):
                r = str(row.get(rack_col, '') or '').strip()
                u = str(row.get(ru_col, '') or '').strip()
                if u and not re.search(r'U\s*\d', r):
                    combined = f"{r} U{u}".strip()
                else:
                    combined = r
                return make_sort_key(combined)
            df[name] = df.apply(_a_rack_combined, axis=1)
        else:
            df[name] = df[rack_col].apply(make_sort_key)
        created.append(name)

    # Device A Port (standalone, now benefits from improved make_sort_key for bare ports)
    if port_col:
        label = 'Sort - DeviceA Port'
        name = label
        i = 1
        while name in df.columns:
            name = f"{label} ({i})"
            i += 1
        df[name] = df[port_col].apply(make_sort_key)
        created.append(name)

    # 2. All Patch Panel / path-related columns (these become PP A, PP B, DMARC, etc.)
    pp_keywords = ['patch', 'full label', 'easymark', 'dmarc', 'passive', 'ohr', 'pp']
    for col in df.columns:
        if any(kw in col.lower() for kw in pp_keywords) and col not in created:
            name = f"Sort - {col}"
            i = 1
            while name in df.columns:
                name = f"Sort - {col} ({i})"
                i += 1
            df[name] = df[col].apply(make_sort_key)
            created.append(name)

    # 3. Device B side (name + port combined)
    b_name_patterns = ['DeviceB Name', 'Device B Name', 'Remote Device Name']
    b_port_patterns = ['DeviceB Port', 'Device B Port', 'Remote Device Port', 'DeviceB Interface', 'Device B Interface', 'B Port', 'Port B', 'Eth Port B', 'Ethernet Port B', 'B Interface', 'Interface B', 'B WIC', 'WIC Port B', 'WIC B', 'B WIC Port']
    b_rack_patterns = ['DeviceB Rack', 'Rack B', 'RackB']

    name_col = next((c for c in b_name_patterns if c in df.columns), None)
    port_col = next((c for c in b_port_patterns if c in df.columns), None)
    if not port_col and name_col:
        # Broader fallback for port columns (for "no eth number" cases)
        for c in df.columns:
            cl = c.lower()
            if any(k in cl for k in ['port', 'eth', 'ethernet', 'interface', 'intf', 'wic']):
                if any(k in cl for k in ['b', 'remote', 'dest', 'far']):
                    port_col = c
                    break
    if name_col:
        label = 'Sort - Device B'
        name = label
        i = 1
        while name in df.columns:
            name = f"{label} ({i})"
            i += 1
        if port_col:
            def _b_combined(row):
                n = str(row.get(name_col, '') or '').strip()
                p = str(row.get(port_col, '') or '').strip()
                # Get device hierarchy from name (strip any embedded port so we don't duplicate)
                device_base = make_sort_key(n)
                if '|' in device_base:
                    parts = device_base.split('|')
                    # Strip trailing pure-digit segments (the port part) if present
                    while parts and parts[-1].isdigit():
                        parts.pop()
                    device_base = '|'.join(parts)
                # Always take the port digits from the dedicated Port column (the source of truth)
                port_digits = _extract_port_digits(p)
                if port_digits:
                    if device_base:
                        return f"{device_base}|{port_digits}"
                    return port_digits
                return device_base
            df[name] = df.apply(_b_combined, axis=1)
        else:
            df[name] = df[name_col].apply(make_sort_key)
        created.append(name)

    # Rack B (combine with RU if the U position lives in a separate column)
    b_ru_patterns = ['DeviceB RU', 'Device B RU', 'B RU']
    rack_col = next((c for c in b_rack_patterns if c in df.columns), None)
    ru_col = next((c for c in b_ru_patterns if c in df.columns), None)
    if rack_col:
        label = 'Sort - Rack B'
        name = label
        i = 1
        while name in df.columns:
            name = f"{label} ({i})"
            i += 1
        if ru_col:
            def _b_rack_combined(row):
                r = str(row.get(rack_col, '') or '').strip()
                u = str(row.get(ru_col, '') or '').strip()
                if u and not re.search(r'U\s*\d', r):
                    combined = f"{r} U{u}".strip()
                else:
                    combined = r
                return make_sort_key(combined)
            df[name] = df.apply(_b_rack_combined, axis=1)
        else:
            df[name] = df[rack_col].apply(make_sort_key)
        created.append(name)

    # Device B Port standalone
    if port_col:
        label = 'Sort - DeviceB Port'
        name = label
        i = 1
        while name in df.columns:
            name = f"{label} ({i})"
            i += 1
        df[name] = df[port_col].apply(make_sort_key)
        created.append(name)

    return df


def get_sort_key_columns(df: pd.DataFrame):
    """Returns all columns that start with 'Sort -' in a consistent order."""
    return sorted([c for c in df.columns if c.startswith('Sort -')])


def preprocess_and_filter_connections(df: pd.DataFrame, filters: dict = None) -> pd.DataFrame:
    """
    Central place for all business logic that decides what should (and should not)
    appear in the generated cutsheet(s).

    This is the main place we will grow "smart" filtering over time.
    """
    if filters is None:
        filters = {}

    original_count = len(df)
    df = df.copy()
    df = df.fillna('')

    removed_reasons = []

    # ============================================================
    # NEW RULE: Classify intra-rack (internal) cables
    # ============================================================
    # Based on user's requirement:
    # - If RackA and RackB are the same physical rack → internal (already installed)
    # - Exception 1: RackB contains Uleft or Uright → still needs patching
    # - Exception 2: Auxiliary racks (user will provide list)
    if filters.get('classify_internals', True):   # Default on for now
        df = classify_internal_cables(df, filters)

        internal_count = int(df['Is_Internal'].sum()) if 'Is_Internal' in df.columns else 0
        if internal_count > 0:
            removed_reasons.append(f"Classified {internal_count} cables as Internal (same rack on both sides)")

    # ============================================================
    # RULE: Legacy exclusion (kept for backward compatibility)
    # ============================================================
    if filters.get('exclude_already_connected', False):
        # Strategy: Look for the most relevant column(s) that likely indicate work status
        status_keywords = ['status', 'state', 'work', 'install', 'complete', 'done', 'cable status', 'connection status']
        internal_keywords = ['internal', 'intra', 'in-rack', 'existing', 'already']

        candidate_cols = [c for c in df.columns if any(kw in c.lower() for kw in status_keywords + internal_keywords)]

        if candidate_cols:
            col = candidate_cols[0]
            before = len(df)
            exclude_pattern = '|'.join([
                'installed', 'connected', 'complete', 'done', 'internal',
                'existing', 'in place', 'already', 'intra', 'in-rack'
            ])
            mask = ~df[col].astype(str).str.lower().str.contains(exclude_pattern, na=False, regex=True)
            df = df[mask]
            after = len(df)
            if before != after:
                removed_reasons.append(f"Removed {before - after} rows using column '{col}'")

    final_count = len(df)

    if removed_reasons:
        st.warning("**Filtering / Classification applied:**\n" + "\n".join(f"• {r}" for r in removed_reasons))

    return df


# ── Core logic from original script (kept intact) ─────────────────────────────
COLOR_RJ45        = '#FFCCCC'  # bg for RJ45 tabs (pastel); actual tab color is red (see tab_hex in build)
COLOR_RJ45_SERIAL = '#CCFFCC'  # bg for serial; tab color is green
COLOR_DB9         = '#FFB347'
COLOR_400G_AOC    = '#CCE5FF'
COLOR_100G_AOC    = '#E0F0FF'
COLOR_DAC         = '#D9D9D9'
COLOR_OTHER       = '#FFFF00'

DARK_BLUE = '#002060'
YELLOW    = '#FFFF00'
GREEN     = '#92D050'
WHITE     = '#FFFFFF'
BLACK     = '#000000'
RED       = '#C00000'
ROW_H     = 15

def assign_color(ct):
    base = ct.replace(' (Internal)', '')
    if 'Copper RJ45 cable - serial' in base:
        return COLOR_RJ45_SERIAL
    elif 'Copper RJ45 cable' in base:
        return COLOR_RJ45
    if 'DB9' in base:                        return COLOR_DB9
    if '400G AOC' in base:                   return COLOR_400G_AOC
    if 'AOC' in base:                        return COLOR_100G_AOC
    if 'DAC' in base:                        return COLOR_DAC
    return COLOR_OTHER

CABLE_TAB_NAMES = {
    'Copper RJ45 cable': 'Copper RJ45 cable',
    'Copper RJ45 cable - serial': 'Copper RJ45 cable - serial',
    'DB9 to copper RJ45 cable - serial: transceivers: DB9 dongle': 'DB9-RJ45 cable serial',
    'QSFP-DD 400G AOC': 'QSFP-DD 400G AOC',
    'QSFP28 100G AOC': 'QSFP28 100G AOC',
    'QSFP28 100G DAC': 'QSFP28 100G DAC',
    'QSFP56 200G DAC': 'QSFP56 200G DAC',
    'SFP+ 10G DAC': 'SFP+ 10G DAC',
    'Singlemode LC duplex fiber patch': 'SM LC duplex fiber patch',
    'Singlemode LC duplex fiber patch - Armoured': 'SM LC duplex - Armoured',
    'Singlemode LC duplex fiber patch: transceivers: QSFP28 LR4 transceiver': 'SM LC QSFP28 LR4',
    'Singlemode LC duplex fiber patch: transceivers: SFP+ LR transceiver': 'SM LC SFP+ LR',
    'Singlemode MPO to Singlemode LC fiber breakout: transceivers: QSFP-DD DR4 400G transceiver: QSFP28 DR1 100G transceiver': 'SM MPO-LC DR4-DR1',
    'Singlemode patch (LC or MPO): transceivers: OSFP 800G Singlemode transceiver: OSFP 800G Singlemode transceiver': 'SM patch OSFP 800G',
    'Singlemode patch (LC or MPO): transceivers: QSFP-112 400G DR4 Singlemode transceiver: OSFP 800G DR4 Singlemode twinport transceiver': 'SM patch QSFP112-OSFP 400G',
    'Singlemode patch (LC or MPO): transceivers: QSFP28 100G Singlemode transceiver: QSFP28 100G Singlemode transceiver': 'SM patch QSFP28 100G',
    'Singlemode patch (LC or MPO): transceivers: QSFP56 200G Singlemode transceiver: QSFP56 200G Singlemode transceiver': 'SM patch QSFP56 200G',
}

def make_tab_name(ct, used, ml=31):
    name = CABLE_TAB_NAMES.get(ct, ct.replace(':', '').replace('/', '- ')
           .replace('\\', '').replace('?', '').replace('*', '')
           .replace('[', '').replace(']', '').strip()[:ml].strip())
    orig, i = name[:ml].strip(), 1
    name = orig
    while name in used:
        sfx = f'_{i}'
        name = orig[:ml - len(sfx)] + sfx
        i += 1
    return name

def fmt_d(n, p): return f"{n} {p}".strip()
def fmt_r(r, u): return f"Rack {r} U{u}"

def parse_row_breakdown(file_bytes):
    """Parse Row Breakdown from uploaded file bytes."""
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        tmp.write(file_bytes)
        tmp_path = tmp.name
    try:
        wb = load_workbook(tmp_path, data_only=True)
        ws = wb.active
        row_to_room = {}
        room_index = 0
        for row in ws.iter_rows(min_row=3, values_only=True):
            a_val = row[0]
            d_val = row[3]
            if a_val is None or not isinstance(a_val, (int, float)): continue
            if d_val is None or '-' not in str(d_val): continue
            room_index += 1
            room_name = f"20.{room_index}"
            parts = str(d_val).strip().split('-')
            start, end = int(parts[0].strip()), int(parts[1].strip())
            for r in range(start, end + 1):
                row_to_room[r] = room_name
        return row_to_room
    finally:
        os.unlink(tmp_path)

def rack_row_prefix(rack_str):
    rack = str(rack_str).strip()
    if len(rack) == 4: return int(rack[:2])
    if len(rack) == 5: return int(rack[:3])
    return None

def assign_connections_to_rooms(df, row_to_room):
    cols = list(df.columns)
    ci_rack_a = cols.index('DeviceA Rack')
    ci_rack_b = cols.index('DeviceB Rack')
    room_idx = defaultdict(set)
    for i, row in enumerate(df.itertuples(index=False)):
        rv = list(row)
        pref_a = rack_row_prefix(rv[ci_rack_a])
        pref_b = rack_row_prefix(rv[ci_rack_b])
        rooms = set()
        if pref_a is not None:
            r = row_to_room.get(pref_a)
            if r: rooms.add(r)
        if pref_b is not None:
            r = row_to_room.get(pref_b)
            if r: rooms.add(r)
        for room in rooms:
            room_idx[room].add(i)
    return {room: sorted(idxs) for room, idxs in room_idx.items()}

# ── Main workbook builder (kept from original) ────────────────────────────────
def build_workbook_to_bytes(df, title_label="", skip_heavy_progress=False):
    """
    Builds the cutsheet workbook.

    IMPORTANT:
    The tabs created here (besides the fixed ones like _allconnects, Main, Template, etc.)
    are determined entirely by the unique values in df['Cable Info'].

    This column comes directly from the export file the user uploaded.
    There is no other "tab list" in the export.
    """
    # --- Support for Internal cable classification ---
    # If the dataframe has an 'Is_Internal' column, we create separate tabs for internal cables
    # Example: "Copper RJ45 cable - mgmt" → "Copper RJ45 cable - mgmt (Internal)"
    if 'Is_Internal' in df.columns:
        def get_tab_key(row):
            base = row['Cable Info']
            if row.get('Is_Internal', False):
                return f"{base} (Internal)"
            return base

        df['_Tab_Key'] = df.apply(get_tab_key, axis=1)
        cable_types = list(df['_Tab_Key'].unique())
    else:
        cable_types = list(df['Cable Info'].unique())

    if not cable_types:
        return None

    cable_hex = {ct: assign_color(ct) for ct in cable_types}

    # Tab colors (sheet tabs in Excel) for copper RJ45 = red, serial = green (as requested)
    # bg colors remain the soft pastels from COLOR_RJ45* for readability in cells.
    # Other tabs use the same hex as their cell bg color.
    tab_hex = {}
    for ct in cable_types:
        base = ct.replace(' (Internal)', '')
        if 'Copper RJ45 cable - serial' in base:
            tab_hex[ct] = GREEN
        elif 'Copper RJ45 cable' in base:
            tab_hex[ct] = RED
        else:
            tab_hex[ct] = cable_hex[ct]

    ALL_COLS = list(df.columns)
    CI = {n: ALL_COLS.index(n) for n in ALL_COLS}
    DA_RACK = CI['DeviceA Rack']; DA_RU = CI['DeviceA RU']
    DA_NAME = CI['DeviceA Name']; DA_PORT = CI['DeviceA Port']
    DB_RACK = CI['DeviceB Rack']; DB_RU = CI['DeviceB RU']
    DB_NAME = CI['DeviceB Name']; DB_PORT = CI['DeviceB Port']

    used_preview = set(['Version Control', '_allconnects', 'Main', 'Formatted_',
                        'PROGRESS Dashboard', 'Racks', 'Cable Liner', 'Template'])
    ct_tab_names = {}
    for ct in cable_types:
        n = make_tab_name(ct, used_preview)
        used_preview.add(n)
        ct_tab_names[ct] = n

    output = io.BytesIO()
    # constant_memory=True is critical for large files to avoid running out of memory
    wb = xlsxwriter.Workbook(output, {
        'strings_to_numbers': False,
        'constant_memory': True
    })

    def mk(bg=None, fc=BLACK, bold=False, wrap=False,
           halign='left', valign='vcenter', sz=10, num_fmt=None):
        d = {'font_size': sz, 'font_color': fc, 'bold': bold, 'text_wrap': wrap,
             'align': halign, 'valign': valign, 'border': 1}
        if bg: d['bg_color'] = bg
        if num_fmt: d['num_format'] = num_fmt
        return wb.add_format(d)

    hdr_blue  = mk(DARK_BLUE, WHITE, bold=True, halign='center', wrap=True)
    hdr_red   = mk(RED, WHITE, bold=True, halign='center')
    hdr_yel   = mk(YELLOW, BLACK, halign='center')
    hdr_grn   = mk(GREEN, BLACK, halign='center', wrap=True)
    hdr_plain = mk(None, BLACK, halign='center')
    dat_l     = mk(None, BLACK, halign='left')
    dat_c     = mk(None, BLACK, halign='center')
    dat_pct   = mk(None, BLACK, halign='center', num_fmt='0%')
    grn_bold  = mk(GREEN, BLACK, bold=True, halign='left')
    yel_c     = mk(YELLOW, BLACK, halign='center')
    blank_bdr = mk(None, BLACK, halign='center')

    fmt_ct = {}
    for ct in cable_types:
        bg = cable_hex[ct]
        fmt_ct[ct] = {
            'sep':  mk(bg, BLACK, bold=True, halign='center'),
            'col':  mk(bg, BLACK, halign='left'),
            'tab':  mk(None, BLACK, halign='center'),
            'num':  mk(None, BLACK, halign='center'),
            'note': mk(None, BLACK, halign='center'),
        }

    tpl_hdrs = [
        ('Label for wrap arounds', hdr_blue), ('Cable Path', hdr_yel),
        ('Device A', hdr_yel), ('Rack A', hdr_yel), ('Device B', hdr_yel), ('Rack B', hdr_yel),
        ('# of Connections', hdr_grn), ('Prep', hdr_grn), ('Prep', hdr_grn),
        ('Pulling', hdr_grn), ('Dressing', hdr_grn), ('Labeling of Cable', hdr_grn),
        ('Percentage of Completion', hdr_grn), ('', hdr_plain), ('NOTES:', hdr_grn),
    ]
    COL_W = [30, 12, 40, 20, 40, 20, 15, 10, 10, 10, 10, 18, 22, 12, 30]

    # Version Control
    ws = wb.add_worksheet('Version Control')
    ws.set_default_row(ROW_H)
    ws.set_column('A:A', 4); ws.set_column('B:B', 22)
    ws.set_column('C:E', 35); ws.set_column('F:F', 18)
    ws.merge_range('A1:F1', 'Cutsheet Change Log', hdr_red); ws.set_row(0, 24)
    label_fmt = mk(None, BLACK, bold=True, halign='right')
    for row, (label, vfmt, val) in enumerate([
        ('Document Name', yel_c, ''), ('Document Owner', dat_c, ''),
        ('Template Version', dat_c, '1.0'), ('Date of creation', yel_c, '')], start=1):
        ws.write_blank(row, 0, None, blank_bdr)
        ws.write(row, 1, label, label_fmt)
        ws.write(row, 2, val, vfmt)
        for c in range(3, 6): ws.write_blank(row, c, None, blank_bdr)
    ws.write_blank(5, 0, None, blank_bdr)
    ws.merge_range('B7:F7', 'Revision History', hdr_red)
    ws.write_blank(6, 0, None, blank_bdr)
    for c, h in enumerate(['Change Maker', 'Description of Change',
                           'Revision', 'Cutsheet Version', 'Date'], 1):
        ws.write(7, c, h, hdr_red)
    ws.write_blank(7, 0, None, blank_bdr)
    for r in range(8, 21):
        for c in range(6): ws.write_blank(r, c, None, blank_bdr)

    # _allconnects
    ws_ac = wb.add_worksheet('_allconnects')
    ws_ac.set_default_row(ROW_H)
    ws_ac.autofilter(0, 0, 0, len(ALL_COLS) - 1)
    ws_ac.freeze_panes(1, 0)
    for i in range(len(ALL_COLS)): ws_ac.set_column(i, i, 20)
    for c, h in enumerate(ALL_COLS): ws_ac.write(0, c, h, hdr_blue)
    for r, rv in enumerate(df.itertuples(index=False), 1):
        ws_ac.write_row(r, 0, list(rv), dat_l)

    # Main
    ws_main = wb.add_worksheet('Main')
    ws_main.set_default_row(ROW_H)
    ws_main.freeze_panes(1, 0)
    ws_main.autofilter(0, 0, 0, len(ALL_COLS) - 1)
    for i in range(len(ALL_COLS)): ws_main.set_column(i, i, 22)
    for c, h in enumerate(ALL_COLS): ws_main.write(0, c, h, hdr_blue)
    cur = 1
    for ct in cable_types:
        if '_Tab_Key' in df.columns:
            grp = df[df['_Tab_Key'] == ct]
        else:
            grp = df[df['Cable Info'] == ct]

        f = fmt_ct[ct]
        for c in range(len(ALL_COLS)): ws_main.write(cur, c, ct, f['sep'])
        cur += 1
        for rv in grp.itertuples(index=False):
            ws_main.write_row(cur, 0, list(rv), f['col'])
            cur += 1

    # Formatted_
    ws_fmt = wb.add_worksheet('Formatted_')
    ws_fmt.set_default_row(ROW_H)
    ws_fmt.freeze_panes(1, 0)
    ws_fmt.autofilter(0, 0, 0, 3)
    for c, w in enumerate([45, 25, 45, 25]): ws_fmt.set_column(c, c, w)
    for c, h in enumerate(['DeviceA', 'RackA', 'DeviceB', 'RackB']): ws_fmt.write(0, c, h, hdr_blue)
    cur = 1
    for ct in cable_types:
        if '_Tab_Key' in df.columns:
            grp = df[df['_Tab_Key'] == ct]
        else:
            grp = df[df['Cable Info'] == ct]

        f = fmt_ct[ct]
        for c in range(4): ws_fmt.write(cur, c, ct, f['sep'])
        cur += 1
        for rv in grp.itertuples(index=False):
            v = list(rv)
            ws_fmt.write_row(cur, 0,
                [fmt_d(v[DA_NAME], v[DA_PORT]), fmt_r(v[DA_RACK], v[DA_RU]),
                 fmt_d(v[DB_NAME], v[DB_PORT]), fmt_r(v[DB_RACK], v[DB_RU])],
                f['col'])
            cur += 1

    # PROGRESS Dashboard
    if skip_heavy_progress:
        # Lightweight version
        ws_pd = wb.add_worksheet('PROGRESS Dashboard')
        ws_pd.set_default_row(ROW_H)
        ws_pd.write(0, 0, "Lite mode - progress tracking disabled for speed", hdr_red)
    else:
        ws_pd = wb.add_worksheet('PROGRESS Dashboard')
        ws_pd.set_default_row(ROW_H)
        ws_pd.set_column('A:A', 10)
        ws_pd.set_column('B:B', 40)
        ws_pd.set_column('C:C', 22)
        ws_pd.merge_range('A1:H1', 'Cutsheet Progress', hdr_red)
        ws_pd.set_row(0, 20)
        for c, h in enumerate(['Tab #', 'Activity', 'Completion Percentage', 'Notes',
                               '', 'Material Missing / Blocker', 'Current ETA', 'Days Needed']):
            ws_pd.write(1, c, h, hdr_red)

        entries = [(1, 'Racks', "=Racks!N1"), (2, 'Cable Liner', "='Cable Liner'!E2")]
        for i, ct in enumerate(cable_types, 3):
            sn = ct_tab_names.get(ct, make_tab_name(ct, set()))
            entries.append((i, sn, f"='{sn}'!N1"))

        for ri, (num, activity, formula) in enumerate(entries, 2):
            ws_pd.write_number(ri, 0, num, dat_c)
            ws_pd.write_string(ri, 1, activity, dat_l)
            ws_pd.write_formula(ri, 2, formula, dat_pct, 0)
            for c in range(3, 8): ws_pd.write_blank(ri, c, None, blank_bdr)

        overall_r = 2 + len(entries)
        ws_pd.write_blank(overall_r, 0, None, blank_bdr)
        ws_pd.write_string(overall_r, 1, 'Overall Project Progress', grn_bold)
        ws_pd.write_formula(overall_r, 2, f'=AVERAGE(C3:C{overall_r})', dat_pct, 0)
        for c in range(3, 8): ws_pd.write_blank(overall_r, c, None, blank_bdr)

    # Racks tab (simplified)
    ws_r = wb.add_worksheet('Racks')
    ws_r.set_default_row(ROW_H)
    ws_r.set_column('A:A', 5)
    ws_r.set_column('B:B', 15)
    ws_r.set_column('C:C', 12)
    ws_r.set_column('D:D', 35)
    for c in range(5, 14): ws_r.set_column(c, c, 14)
    rack_hdrs = ['', 'SO#', 'Status', 'Rack Platform', 'Rack Location Number',
                 'BOPS Tickets', 'Serial Number', 'Foot Leveling', 'Grounding',
                 'Rack Power up', 'Cabling', 'Percentage complete', 'Percentage of Completion']
    for c, h in enumerate(rack_hdrs):
        ws_r.write(0, c, h, hdr_yel if c < 7 else hdr_grn)
    for r in range(1, 17):
        ws_r.write_blank(r, 0, None, blank_bdr)
        for c in range(1, 7): ws_r.write_blank(r, c, None, dat_c)
        for c in range(7, 11): ws_r.write_number(r, c, 0, dat_c)
        ws_r.write_formula(r, 11, f'=SUM(H{r+1}:K{r+1})', dat_c, 0)

    # Cable Liner
    ws_cl = wb.add_worksheet('Cable Liner')
    ws_cl.set_default_row(ROW_H)
    ws_cl.set_column('A:A', 5)
    ws_cl.set_column('B:B', 15)
    ws_cl.set_column('C:C', 22)
    ws_cl.set_column('D:D', 24)
    ws_cl.set_column('E:E', 20)
    ws_cl.write_blank(0, 0, None, blank_bdr)
    ws_cl.write(0, 1, 'Block#', hdr_blue)
    ws_cl.write(0, 2, 'Liner Installation', hdr_yel)
    ws_cl.write(0, 3, 'Percentage of Completion', hdr_grn)
    ws_cl.write_formula(0, 4, '=AVERAGE(C2:C3)', hdr_grn, 0)
    for r, block in enumerate(['Block-1', 'Block-2'], 1):
        ws_cl.write_blank(r, 0, None, blank_bdr)
        ws_cl.write(r, 1, block, dat_c)
        ws_cl.write_number(r, 2, 0, dat_c)

    # Template (heavy sheet)
    if not skip_heavy_progress:
        ws_t = wb.add_worksheet('Template')
        ws_t.set_column('A:A', 30)
        ws_t.set_column('B:B', 12)
        ws_t.set_column('C:C', 40)
        ws_t.set_column('D:D', 20)
        ws_t.set_column('E:E', 40)
        ws_t.set_column('F:F', 20)
        ws_t.set_column('G:G', 15)
        for c in range(7, 12): ws_t.set_column(c, c, 10)
        ws_t.set_column('M:M', 22)
        ws_t.set_column('N:N', 12)
        ws_t.set_column('O:O', 30)
        ws_t.freeze_panes(1, 0)
        ws_t.autofilter(0, 0, 0, 14)

        # Make the master Template sheet size dynamic
        total_rows = len(df)
        TEMPLATE_ROWS = min(max(100, total_rows), 500)

        for ci, (h, hf) in enumerate(tpl_hdrs):
            if ci == 13:
                ws_t.write_formula(0, ci, f'=AVERAGE(H2:L{TEMPLATE_ROWS+1})', hdr_plain, 0)
            else:
                ws_t.write(0, ci, h, hf)
        for r in range(1, TEMPLATE_ROWS + 1):
            er = r + 1
            ws_t.set_row(r, ROW_H)
            ws_t.write_formula(r, 0,
                f'=CONCATENATE(C{er},CHAR(10),D{er},CHAR(10),E{er},CHAR(10),F{er})',
                hdr_blue, '')
            ws_t.write(r, 1, 'A', dat_c)
            for c in range(2, 6): ws_t.write_blank(r, c, None, dat_c)
            ws_t.write_number(r, 6, 1, dat_c)
            for c in range(7, 12): ws_t.write_number(r, c, 0, dat_c)
            ws_t.write_blank(r, 12, None, dat_c)
            ws_t.write_blank(r, 13, None, dat_c)
            ws_t.write(r, 14, '', yel_c)

    # Cable-type tabs
    used = set(wb.sheetnames)
    for ct in cable_types:
        if '_Tab_Key' in df.columns:
            grp = df[df['_Tab_Key'] == ct].reset_index(drop=True)
        else:
            grp = df[df['Cable Info'] == ct].reset_index(drop=True)

        f = fmt_ct[ct]
        sname = ct_tab_names.get(ct, make_tab_name(ct, used))
        used.add(sname)
        ws = wb.add_worksheet(sname)
        ws.set_tab_color(tab_hex[ct])
        ws.set_default_row(14.9)
        ws.freeze_panes(1, 0)
        last = 1 + len(grp)

        is_internal_tab = "(Internal)" in sname

        # Compute sort keys only for non-internal tabs
        sort_key_cols = []
        sort_start_col = 21
        if not is_internal_tab:
            sort_key_cols = get_sort_key_columns(grp)

        # Remove column A (label) and B (cable type) - they are not useful on the tabs
        new_tpl_hdrs = tpl_hdrs[2:]
        new_col_w = COL_W[2:]
        main_data_cols = len(new_tpl_hdrs)  # should be 13

        ws.autofilter(0, 0, 0, main_data_cols - 1)

        if skip_heavy_progress:
            # Lite version: much simpler and faster to write
            ws.set_row(0, 15)
            # Cable type put back as column A (label removed as redundant)
            lite_headers = ['Cable Type', 'Device A', 'Rack A', 'Device B', 'Rack B', 'Notes']
            for ci, h in enumerate(lite_headers):
                ws.write(0, ci, h, hdr_blue)
            ws.set_column(0, 0, 25)  # cable type
            ws.set_column(1, 1, 45)
            ws.set_column(2, 2, 22)
            ws.set_column(3, 3, 45)
            ws.set_column(4, 4, 22)
            ws.set_column(5, 5, 30)

            # Sort keys on the far right - skip for internal tabs
            if sort_key_cols:
                sort_hdr = mk('#BDD7EE', BLACK, bold=True, halign='center')
                for idx, colname in enumerate(sort_key_cols):
                    cidx = sort_start_col + idx
                    ws.set_column(cidx, cidx, 34)
                    ws.write(0, cidx, colname, sort_hdr)

            for i, rv in enumerate(grp.itertuples(index=False), 1):
                v = list(rv)
                da = fmt_d(v[DA_NAME], v[DA_PORT])
                ra = fmt_r(v[DA_RACK], v[DA_RU])
                db = fmt_d(v[DB_NAME], v[DB_PORT])
                rb = fmt_r(v[DB_RACK], v[DB_RU])
                ws.set_row(i, 15)
                # col 0: cable type back, then the rest shifted (no label)
                ws.write(i, 0, ct, f['tab'])
                ws.write(i, 1, da, f['tab'])
                ws.write(i, 2, ra, f['tab'])
                ws.write(i, 3, db, f['tab'])
                ws.write(i, 4, rb, f['tab'])
                ws.write(i, 5, '', f['note'])

                # Write sort keys on the right (skip for internal)
                if sort_key_cols:
                    for idx, colname in enumerate(sort_key_cols):
                        cidx = sort_start_col + idx
                        val = v[ALL_COLS.index(colname)] if colname in ALL_COLS else ''
                        ws.write(i, cidx, val, dat_l)
        else:
            # Full complex version
            ws.set_row(0, 15)
            # Cable type back as first column (label removed)
            effective_headers = ['Cable Type'] + [h for h, hf in new_tpl_hdrs]
            for ci, h in enumerate(effective_headers):
                if ci == 12:  # the blank col (now 12 after adding cable type)
                    # average range: preps etc now shifted to cols 6-10 (G-K)
                    ws.write_formula(0, ci, f'=AVERAGE(G2:K{last})', hdr_plain, 0)
                elif ci == 0:
                    ws.write(0, ci, h, hdr_blue)  # cable type header
                else:
                    # find matching format from original
                    orig_hf = next((hf for hh, hf in new_tpl_hdrs if hh == h), hdr_blue)
                    ws.write(0, ci, h, orig_hf)
            # widths: cable type + the new ones
            ws.set_column(0, 0, 25)  # cable type
            for ci, w in enumerate(new_col_w):
                ws.set_column(ci + 1, ci + 1, w)

            # Sort key headers on the right - skip for internal tabs
            if sort_key_cols:
                sort_hdr = mk('#BDD7EE', BLACK, bold=True, halign='center')
                for idx, colname in enumerate(sort_key_cols):
                    cidx = sort_start_col + idx
                    ws.set_column(cidx, cidx, 34)
                    ws.write(0, cidx, colname, sort_hdr)

            for i, rv in enumerate(grp.itertuples(index=False), 1):
                v = list(rv)
                da = fmt_d(v[DA_NAME], v[DA_PORT])
                ra = fmt_r(v[DA_RACK], v[DA_RU])
                db = fmt_d(v[DB_NAME], v[DB_PORT])
                rb = fmt_r(v[DB_RACK], v[DB_RU])
                ws.set_row(i, 15)
                # col 0: cable type back, then data shifted (skipped the label col)
                ws.write(i, 0, ct, f['tab'])
                ws.write(i, 1, da, f['tab'])
                ws.write(i, 2, ra, f['tab'])
                ws.write(i, 3, db, f['tab'])
                ws.write(i, 4, rb, f['tab'])
                ws.write(i, 5, 1, f['num'])
                ws.write(i, 6, 0, f['num'])
                ws.write(i, 7, 0, f['num'])
                ws.write(i, 8, 0, f['num'])
                ws.write(i, 9, 0, f['num'])
                ws.write(i, 10, 0, f['num'])
                ws.write_blank(i, 11, None, f['tab'])
                ws.write_blank(i, 12, None, f['tab'])
                ws.write(i, 13, '', f['note'])

                # === Sort Keys on the far right - skip for internal tabs
                if sort_key_cols:
                    for idx, colname in enumerate(sort_key_cols):
                        cidx = sort_start_col + idx
                        val = v[ALL_COLS.index(colname)] if colname in ALL_COLS else ''
                        ws.write(i, cidx, val, dat_l)

    wb.close()
    output.seek(0)
    return output.getvalue()

# ── Streamlit UI ──────────────────────────────────────────────────────────────
st.markdown("### Upload Files")

mode = st.radio(
    "Mode",
    ["Single cutsheet", "Split by room (requires Row Breakdown)"],
    horizontal=True
)

input_file = st.file_uploader("All Connects export (.xlsx)", type=["xlsx"])

breakdown_file = None
if mode == "Split by room (requires Row Breakdown)":
    breakdown_file = st.file_uploader("Row Breakdown file (.xlsx)", type=["xlsx"])

# ── Filtering options moved OUTSIDE the button using a form (fixes the looping bug) ──
with st.form("options_form"):
    st.markdown("#### Filtering & Internal Cable Logic")

    classify_internals = st.checkbox(
        "Classify intra-rack cables as Internal (recommended)",
        value=True,
        help="Cables where RackA and RackB are the same rack will be marked Internal and can get their own tab"
    )

    aux_racks_input = st.text_input(
        "Auxiliary rack numbers (comma separated)",
        value="",
        help="e.g. 2701,2702,2710 — cables to these racks will NOT be treated as internal even if RackA == RackB"
    )
    aux_racks = {r.strip() for r in aux_racks_input.split(",") if r.strip()}

    add_sort_keys_option = st.checkbox(
        "Add Sort Keys on the right side (for all path columns)",
        value=True,
        help="Uses your comprehensive SortKey logic (OHR, Passive, PP:SYD, full syd20 family). Creates the right sort keys for whatever columns exist on each tab and places them starting at column V."
    )

    exclude_done = st.checkbox(
        "Also try to exclude rows using status/keyword columns (legacy)",
        value=False,
        help="Uses older heuristic based on column names containing 'status', 'installed', etc."
    )

    st.markdown("#### Performance Options (for speed)")
    skip_heavy_progress = st.checkbox(
        "Skip heavy progress sheets (much faster)",
        value=True,
        help="Skips the big 'Template' sheet and some progress tracking. Recommended for large files."
    )

    submitted = st.form_submit_button("🚀 Generate Cutsheet(s)", type="primary", disabled=not input_file)

if submitted:
    with st.spinner("Processing... This can take a few minutes for large allconnect sheets."):
        try:
            # Read input
            df_all = pd.read_excel(input_file, dtype=str).fillna('')
            st.success(f"Loaded {len(df_all):,} raw connections")

            # Show columns for transparency
            with st.expander("Input columns detected (for building better filters)"):
                st.write(list(df_all.columns))

            # Show what tabs would be created
            cable_types_preview = sorted(df_all['Cable Info'].dropna().unique().tolist())
            with st.expander(f"📋 Tabs that will be created from this export ({len(cable_types_preview)} cable types)"):
                st.dataframe(
                    pd.DataFrame({"Cable Info Value → Will become this tab": cable_types_preview}),
                    use_container_width=True,
                    hide_index=True
                )

            filters = {
                'classify_internals': classify_internals,
                'auxiliary_racks': aux_racks,
                'exclude_already_connected': exclude_done,
                'skip_heavy_progress': skip_heavy_progress,
                'add_sort_keys': add_sort_keys_option,
            }

            # === Filtering ===
            df_filtered = preprocess_and_filter_connections(df_all, filters)

            # Show preview of what was filtered
            if len(df_filtered) < len(df_all):
                with st.expander("🔍 Preview: Sample of rows that were filtered / reclassified"):
                    removed_mask = ~df_all.index.isin(df_filtered.index)
                    removed_sample = df_all[removed_mask].head(10)
                    st.dataframe(removed_sample, use_container_width=True)
                    st.caption(f"Showing up to 10 examples of affected rows (out of {len(df_all) - len(df_filtered)} total affected)")

            st.info(f"**Final connections** that will go into the generated cutsheet(s): **{len(df_filtered):,}**")

            df_all = df_filtered.sort_values('Cable Info').reset_index(drop=True)

            # --- NEW: Apply sort keys if requested ---
            if filters.get('add_sort_keys'):
                df_all = add_sort_keys(df_all)
                st.info("Sort keys generated and will be placed on the far right of each tab (starting column V).")

            # Give the user visible progress so it doesn't feel stuck
            if filters.get('skip_heavy_progress'):
                st.write("Generating **lite version** (fewer formulas, simpler tabs)...")
            else:
                st.write("Generating **full version** with all progress tracking...")

            st.write("Preparing data for per-cable tabs...")

            df_all = df_all.sort_values('Cable Info').reset_index(drop=True)

            if mode == "Single cutsheet":
                bytes_data = build_workbook_to_bytes(df_all, skip_heavy_progress=filters.get('skip_heavy_progress', False))
                if bytes_data:
                    st.download_button(
                        "📥 Download Formatted Cutsheet",
                        data=bytes_data,
                        file_name="cutsheet_formatted.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True
                    )
                else:
                    st.error("No data to process.")

            else:  # Split by room
                if not breakdown_file:
                    st.error("Please upload the Row Breakdown file.")
                else:
                    row_to_room = parse_row_breakdown(breakdown_file.getvalue())
                    room_assignments = assign_connections_to_rooms(df_all, row_to_room)

                    all_assigned = set(i for idxs in room_assignments.values() for i in idxs)
                    unmatched = len(df_all) - len(all_assigned)
                    if unmatched:
                        st.warning(f"{unmatched} connections had rack numbers not matching any room.")

                    rooms_sorted = sorted(room_assignments.keys(), key=lambda r: float(r.replace('20.', '')))

                    # Create ZIP in memory
                    zip_buffer = io.BytesIO()
                    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zipf:
                        for room in rooms_sorted:
                            idxs = room_assignments[room]
                            df_room = df_all.iloc[idxs].copy()
                            # Re-apply any additional per-room logic here if needed in future
                            xlsx_bytes = build_workbook_to_bytes(df_room, room, skip_heavy_progress=filters.get('skip_heavy_progress', False))
                            if xlsx_bytes:
                                zipf.writestr(f"{room}_cutsheet.xlsx", xlsx_bytes)

                    zip_buffer.seek(0)
                    st.success(f"Generated {len(rooms_sorted)} room cutsheets.")
                    st.download_button(
                        "📥 Download All Room Cutsheets (ZIP)",
                        data=zip_buffer,
                        file_name="room_cutsheets.zip",
                        mime="application/zip",
                        use_container_width=True
                    )

        except Exception as e:
            st.error(f"Error: {str(e)}")
            st.exception(e)

st.caption("")

# ── Sort Key Tester (always available) ────────────────────────────────────────
st.divider()
with st.expander("🧪 Sort Key Tester — Test the new comprehensive logic", expanded=False):
    st.markdown("""
    Paste any strings here (one per line) — PP labels, device names, OHR, Passive, etc. — 
    to see exactly what sort keys the migrated VBA logic produces.
    """)
    
    test_input = st.text_area(
        "Paste test strings",
        height=140,
        placeholder="PP:SYD20.3.DH8.R2302.U1.A1.P4\nsyd20-c1-b70-t0-r9 Ethernet1/1\nsyd20-c1-b12-t2-r8 et-0/0/12\nSYD20-C1-B13-T2-R1\nsyd20-c1-b13-t2-r1 et-0/0/0\nRack 2708 U23\nRack 3405 U42\n3405\nsyd20-q2-b31-t0-r1 Ethernet2/1\nSYD20-Q2-B31-T0-R1\nsyd20-c1-b31-t0-r1-compute1 slot1/port1-1\nSYD20-C1-B31-T0-R1-COMPUTE1\nSYD20-C1-B31-T0-R1\nsyd20-c1-b31-t0-r1-compute1-wic1 [4]1 (port1)\nSYD20-C1-B31-T0-R1-COMPUTE1-WIC1"
    )
    
    if st.button("Test Sort Keys", key="sortkey_tester"):
        if test_input.strip():
            lines = [line.strip() for line in test_input.split("\n") if line.strip()]
            results = []
            for line in lines:
                try:
                    key = make_sort_key(line)
                    results.append({"Original": line, "Sort Key": key})
                except Exception as ex:
                    results.append({"Original": line, "Sort Key": f"ERROR: {ex}"})
            
            if results:
                st.dataframe(results, use_container_width=True, hide_index=True)
                st.caption(f"Tested {len(results)} strings using the full ported VBA logic.")
        else:
            st.warning("Please paste at least one string to test.")

