import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Font
from openpyxl.utils import get_column_letter
from copy import copy
from pathlib import Path
import io
import zipfile
import tempfile
import os
from collections import defaultdict

st.set_page_config(page_title="SYD20 CFAB Validation Formatter", page_icon="🗄️", layout="wide")
st.title("SYD20 CFAB Validation Formatter")
st.caption("CFAB)")

# ---------------------------------------------------------------------------
# Style constants (from original)
# ---------------------------------------------------------------------------
YELLOW = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
PINK   = PatternFill(start_color="FFC0CB", end_color="FFC0CB", fill_type="solid")
THIN   = Side(style="thin", color="000000")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
LIGHT_GREY = "A6A6A6"

# ---------------------------------------------------------------------------
# Cutsheet lookup (unchanged from original)
# ---------------------------------------------------------------------------
def build_cutsheet_lookup(path: str):
    df = pd.read_excel(path, sheet_name=0, header=None)
    lookup = {}
    device_rack_lookup = {}

    def add(key, device_rack, pp_list, peer_device, peer_rack):
        if not key or key in lookup:
            return
        lookup[key] = {
            "device_rack": device_rack,
            "pp1": pp_list[0] if len(pp_list) > 0 else None,
            "pp2": pp_list[1] if len(pp_list) > 1 else None,
            "pp3": pp_list[2] if len(pp_list) > 2 else None,
            "pp4": pp_list[3] if len(pp_list) > 3 else None,
            "peer_device": peer_device,
            "peer_rack": peer_rack,
        }
        dev_name = key.split(" ", 1)[0] if key else None
        if dev_name and dev_name not in device_rack_lookup and pd.notna(device_rack):
            device_rack_lookup[dev_name] = device_rack

    for _, row in df.iterrows():
        is_long = pd.notna(row[4])
        if is_long:
            has_4pp = pd.notna(row[5]) and isinstance(row[5], str) and row[5].strip().startswith("PP")
            if has_4pp:
                a_key = str(row[0]) if pd.notna(row[0]) else None
                b_key = str(row[6]) if pd.notna(row[6]) else None
                add(a_key, row[1], [row[2], row[3], row[4], row[5]], row[6], row[7])
                add(b_key, row[7], [row[5], row[4], row[3], row[2]], row[0], row[1])
            else:
                a_key = str(row[0]) if pd.notna(row[0]) else None
                b_key = str(row[5]) if pd.notna(row[5]) else None
                add(a_key, row[1], [row[2], row[3], row[4]], row[5], row[6])
                add(b_key, row[6], [row[4], row[3], row[2]], row[0], row[1])
        else:
            a_key = str(row[0]) if pd.notna(row[0]) else None
            b_key = str(row[2]) if pd.notna(row[2]) else None
            add(a_key, row[1], [], row[2], row[3])
            add(b_key, row[3], [], row[0], row[1])

    return lookup, device_rack_lookup

def split_device_port(s):
    if not s or not isinstance(s, str):
        return None, None
    parts = s.strip().split(" ", 1)
    return (parts[0], parts[1]) if len(parts) == 2 else (parts[0], None)

def style_cell(cell, ref):
    cell.font = copy(ref.font)
    cell.fill = copy(ref.fill)
    cell.border = copy(ref.border)
    cell.alignment = copy(ref.alignment)
    cell.number_format = ref.number_format

def header_indices(ws):
    return {c.value: i + 1 for i, c in enumerate(ws[1])}

def delete_columns_by_name(ws, names):
    for name in names:
        hdrs = [c.value for c in ws[1]]
        if name in hdrs:
            ws.delete_cols(hdrs.index(name) + 1)

def rack_number_from(ws):
    hdrs = header_indices(ws)
    if "Device A Rack" not in hdrs:
        return "output"
    col = hdrs["Device A Rack"]
    for r in range(2, ws.max_row + 1):
        v = ws.cell(row=r, column=col).value
        if v:
            return str(v).split(":")[0].strip()
    return "output"

def autofit_columns(ws, min_w=12, max_w=40):
    for col_idx in range(1, ws.max_column + 1):
        max_len = 0
        for r in range(1, ws.max_row + 1):
            v = ws.cell(row=r, column=col_idx).value
            if v is not None:
                max_len = max(max_len, len(str(v)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, min_w), max_w)

# ---------------------------------------------------------------------------
# All processing steps (kept from original)
# ---------------------------------------------------------------------------
def split_lldp_sheet(wb):
    src_name = "LLDP Mismatch + Link Down"
    if src_name not in wb.sheetnames:
        return
    src = wb[src_name]
    hdrs = [c.value for c in src[1]]
    status_idx = hdrs.index("LLDP Status")

    header_styles = [{
        "font": copy(c.font), "fill": copy(c.fill),
        "border": copy(c.border), "alignment": copy(c.alignment),
        "number_format": c.number_format,
    } for c in src[1]]
    col_widths = {k: v.width for k, v in src.column_dimensions.items()}

    down, mismatch = [], []
    for row in src.iter_rows(min_row=2, values_only=True):
        if row[status_idx] == "DOWN":
            down.append(row)
        elif row[status_idx] == "MISMATCH":
            mismatch.append(row)

    def build(name, rows):
        if name in wb.sheetnames:
            del wb[name]
        ws = wb.create_sheet(name)
        ws.append(hdrs)
        for i, c in enumerate(ws[1], start=1):
            s = header_styles[i - 1]
            c.font = s["font"]; c.fill = s["fill"]; c.border = s["border"]
            c.alignment = s["alignment"]; c.number_format = s["number_format"]
        for r in rows:
            ws.append(r)
        for k, w in col_widths.items():
            if w:
                ws.column_dimensions[k].width = w
        ws.freeze_panes = "A2"

    build("Downlink", down)
    build("Mismatch", mismatch)
    del wb[src_name]

def clean_columns(wb):
    if "Downlink" in wb.sheetnames:
        delete_columns_by_name(wb["Downlink"], [
            "Device A Rack", "Expected Device B Rack", "Expected Device B Name",
            "Expected Device B Port", "Device B Rack", "Device B Name",
            "Device B Port", "LLDP Status", "Patch Panel Matrix",
        ])

    if "Mismatch" in wb.sheetnames:
        delete_columns_by_name(wb["Mismatch"], [
            "Device A Rack", "Expected Device B Rack", "Expected Device B Name",
            "Expected Device B Port", "Device B Rack",
            "LLDP Status", "Patch Panel Matrix",
        ])

    if "Optic Errors" in wb.sheetnames:
        ws = wb["Optic Errors"]
        delete_columns_by_name(ws, [
            "Remote Device Name", "Remote Device Port", "Patch Panel Matrix",
        ])
        hdrs = [c.value for c in ws[1]]
        if "Rx Power" in hdrs and hdrs.index("Rx Power") != 0:
            rx_idx = hdrs.index("Rx Power")
            new_order = [rx_idx] + [i for i in range(len(hdrs)) if i != rx_idx]
            data = [[c.value for c in row] for row in ws.iter_rows()]
            styles = [[{
                "font": copy(c.font), "fill": copy(c.fill),
                "border": copy(c.border), "alignment": copy(c.alignment),
                "number_format": c.number_format,
            } for c in row] for row in ws.iter_rows()]
            old_widths = [ws.column_dimensions[get_column_letter(i + 1)].width
                          for i in range(len(hdrs))]
            ws.delete_rows(1, ws.max_row)
            for r_i, (row_vals, row_styles) in enumerate(zip(data, styles), start=1):
                for new_c, old_c in enumerate(new_order, start=1):
                    cell = ws.cell(row=r_i, column=new_c, value=row_vals[old_c])
                    st = row_styles[old_c]
                    cell.font = st["font"]; cell.fill = st["fill"]
                    cell.border = st["border"]; cell.alignment = st["alignment"]
                    cell.number_format = st["number_format"]
            for new_c, old_c in enumerate(new_order, start=1):
                w = old_widths[old_c]
                if w:
                    ws.column_dimensions[get_column_letter(new_c)].width = w

    if "Interface Down Errors" in wb.sheetnames:
        delete_columns_by_name(wb["Interface Down Errors"], [
            "Source Device Location", "Remote Device Name", "Remote Device Port",
            "Issue", "Patch Panel Matrix",
        ])

HOP_HEADERS = ["Device Rack U", "PP 1", "PP 2", "PP 3", "PP 4", "Peer Device", "Peer Port", "Peer Rack"]

def enrich(ws, name_col_idx, port_col_idx, insert_after_idx, lookup):
    header_ref = ws.cell(row=1, column=1)
    data_ref = ws.cell(row=2, column=1) if ws.max_row >= 2 else header_ref

    new_data = []
    for r in range(2, ws.max_row + 1):
        name = ws.cell(row=r, column=name_col_idx).value
        port = ws.cell(row=r, column=port_col_idx).value
        key = f"{name} {port}" if name and port else None
        info = lookup.get(key) if key else None
        if info:
            peer_name, peer_port = split_device_port(info["peer_device"])
            if info["pp1"] is not None:
                new_data.append([
                    info["device_rack"],
                    info["pp1"], info["pp2"], info["pp3"], info["pp4"],
                    peer_name, peer_port, info["peer_rack"],
                ])
            else:
                new_data.append([
                    info["device_rack"],
                    None, None, None, None,
                    peer_name, peer_port, info["peer_rack"],
                ])
        else:
            new_data.append([None] * 8)

    for i in range(8):
        ws.insert_cols(insert_after_idx + 1 + i)
    for i, h in enumerate(HOP_HEADERS):
        c = ws.cell(row=1, column=insert_after_idx + 1 + i, value=h)
        style_cell(c, header_ref)
    for r, vals in enumerate(new_data, start=2):
        for i, v in enumerate(vals):
            c = ws.cell(row=r, column=insert_after_idx + 1 + i)
            c.value = v
            style_cell(c, data_ref)

def enrich_all(wb, lookup):
    if "Downlink" in wb.sheetnames:
        enrich(wb["Downlink"], 1, 2, 2, lookup)
    if "Optic Errors" in wb.sheetnames:
        enrich(wb["Optic Errors"], 2, 3, 3, lookup)
    if "Mismatch" in wb.sheetnames:
        enrich(wb["Mismatch"], 1, 2, 2, lookup)
    if "Interface Down Errors" in wb.sheetnames:
        enrich(wb["Interface Down Errors"], 1, 2, 2, lookup)

SPLIT_KEY_HEADERS = {
    "Downlink":              ("Device A Name",      "Device A Port"),
    "Mismatch":              ("Device A Name",      "Device A Port"),
    "Optic Errors":          ("Source Device Name", "Source Device Port"),
    "Interface Down Errors": ("Source Device Name", "Source Device Port"),
}

def split_long_short(wb, lookup):
    for src_name in ["Downlink", "Mismatch", "Optic Errors", "Interface Down Errors"]:
        if src_name not in wb.sheetnames:
            continue
        src = wb[src_name]
        headers = [c.value for c in src[1]]
        name_hdr, port_hdr = SPLIT_KEY_HEADERS[src_name]
        name_col = headers.index(name_hdr) + 1
        port_col = headers.index(port_hdr) + 1
        header_styles = [{
            "font": copy(c.font), "fill": copy(c.fill),
            "border": copy(c.border), "alignment": copy(c.alignment),
            "number_format": c.number_format,
        } for c in src[1]]
        data_ref = src.cell(row=2, column=1) if src.max_row >= 2 else src.cell(row=1, column=1)
        col_widths = {k: v.width for k, v in src.column_dimensions.items()}

        long_rows, short_rows = [], []
        for r in range(2, src.max_row + 1):
            row_vals = [src.cell(row=r, column=c).value for c in range(1, src.max_column + 1)]
            name = row_vals[name_col - 1]
            port = row_vals[port_col - 1]
            key = f"{name} {port}" if name and port else None
            info = lookup.get(key) if key else None
            if info and info.get("pp1") is not None:
                long_rows.append(row_vals)
            else:
                short_rows.append(row_vals)

        def build(name, rows):
            if name in wb.sheetnames:
                del wb[name]
            ws = wb.create_sheet(name)
            ws.append(headers)
            for i, c in enumerate(ws[1], start=1):
                s = header_styles[i - 1]
                c.font = s["font"]; c.fill = s["fill"]; c.border = s["border"]
                c.alignment = s["alignment"]; c.number_format = s["number_format"]
            for row_vals in rows:
                ws.append(row_vals)
                for c in ws[ws.max_row]:
                    style_cell(c, data_ref)
            for k, w in col_widths.items():
                if w:
                    ws.column_dimensions[k].width = w
            ws.freeze_panes = "A2"

        build(f"T3-T2 {src_name}", long_rows)
        build(f"T2-T1-T0 {src_name}", short_rows)
        del wb[src_name]

def trim_short_tabs(wb):
    for sn in ["T2-T1-T0 Downlink", "T2-T1-T0 Mismatch", "T2-T1-T0 Optic Errors",
               "T2-T1-T0 Interface Down Errors"]:
        if sn in wb.sheetnames:
            delete_columns_by_name(wb[sn], ["PP 1", "PP 2", "PP 3", "PP 4"])

def enrich_mismatch_b_side(wb, lookup, device_rack_lookup):
    for sn, b_headers, expect_long in [
        ("T3-T2 Mismatch",
         ["Act. Rack U", "Cut. PP 1", "Cut. PP 2", "Cut. PP 3", "Cut. PP 4",
          "Cut. Other End", "Cut. Other End Port", "Cut. Other End Rack"],
         True),
        ("T2-T1-T0 Mismatch",
         ["Act. Rack U", "Cut. Other End", "Cut. Other End Port", "Cut. Other End Rack"],
         False),
    ]:
        if sn not in wb.sheetnames:
            continue
        ws = wb[sn]
        header_ref = ws.cell(row=1, column=1)
        data_ref = ws.cell(row=2, column=1) if ws.max_row >= 2 else header_ref

        rename_map = {
            "Device Rack U": "A Rack U",
            "PP 1": "A PP 1", "PP 2": "A PP 2",
            "PP 3": "A PP 3", "PP 4": "A PP 4",
            "Peer Device": "Exp. Device",
            "Peer Port":   "Exp. Port",
            "Peer Rack":   "Exp. Rack",
        }
        for c in ws[1]:
            if c.value in rename_map:
                c.value = rename_map[c.value]

        start_col = ws.max_column + 1
        for i, h in enumerate(b_headers):
            c = ws.cell(row=1, column=start_col + i, value=h)
            style_cell(c, header_ref)

        hdrs = [c.value for c in ws[1]]
        bname_col = hdrs.index("Device B Name") + 1
        bport_col = hdrs.index("Device B Port") + 1
        slots = len(b_headers)

        for r in range(2, ws.max_row + 1):
            name = ws.cell(row=r, column=bname_col).value
            port = ws.cell(row=r, column=bport_col).value
            key = f"{name} {port}" if name and port else None
            info = lookup.get(key) if key else None

            if info:
                peer_name, peer_port = split_device_port(info["peer_device"])
                if expect_long:
                    vals = [info["device_rack"], info["pp1"], info["pp2"], info["pp3"], info["pp4"],
                            peer_name, peer_port, info["peer_rack"]]
                else:
                    vals = [info["device_rack"], peer_name, peer_port, info["peer_rack"]]
            else:
                fallback_rack = device_rack_lookup.get(str(name).strip()) if name else None
                if expect_long:
                    vals = [fallback_rack, None, None, None, None, None, None, None]
                else:
                    vals = [fallback_rack, None, None, None]

            for i, v in enumerate(vals):
                c = ws.cell(row=r, column=start_col + i)
                c.value = v
                style_cell(c, data_ref)

        hdrs = [c.value for c in ws[1]]
        start_pink = hdrs.index("Device B Name") + 1
        rename_b = {"Device B Name": "Act. Device", "Device B Port": "Act. Port"}
        for col in range(start_pink, ws.max_column + 1):
            cur_h = ws.cell(row=1, column=col).value
            if cur_h in rename_b:
                ws.cell(row=1, column=col).value = rename_b[cur_h]
            for r in range(1, ws.max_row + 1):
                ws.cell(row=r, column=col).fill = PINK

def grey_optics_matching_downlink(wb):
    keys = set()
    for sn in ["T3-T2 Downlink", "T2-T1-T0 Downlink"]:
        if sn not in wb.sheetnames:
            continue
        ws = wb[sn]
        hdrs = header_indices(ws)
        n_col, p_col, r_col = hdrs.get("Device A Name"), hdrs.get("Device A Port"), hdrs.get("Device Rack U")
        if not (n_col and p_col):
            continue
        for r in range(2, ws.max_row + 1):
            n = ws.cell(row=r, column=n_col).value
            p = ws.cell(row=r, column=p_col).value
            ru = ws.cell(row=r, column=r_col).value if r_col else None
            if n and p:
                keys.add((str(n).strip(), str(p).strip(), str(ru).strip() if ru else None))

    for sn in ["T3-T2 Optic Errors", "T2-T1-T0 Optic Errors"]:
        if sn not in wb.sheetnames:
            continue
        ws = wb[sn]
        hdrs = header_indices(ws)
        n_col = hdrs.get("Source Device Name")
        p_col = hdrs.get("Source Device Port")
        r_col = hdrs.get("Device Rack U")
        if not (n_col and p_col):
            continue
        for r in range(2, ws.max_row + 1):
            n = ws.cell(row=r, column=n_col).value
            p = ws.cell(row=r, column=p_col).value
            ru = ws.cell(row=r, column=r_col).value if r_col else None
            key = (str(n).strip() if n else None, str(p).strip() if p else None,
                   str(ru).strip() if ru else None)
            if n and p and key in keys:
                for c in range(1, ws.max_column + 1):
                    cell = ws.cell(row=r, column=c)
                    f = copy(cell.font)
                    f.color = LIGHT_GREY
                    cell.font = f

def finalise_column_names(wb):
    DOWNLINK_MAP = {"Peer Device": "Exp. Device", "Peer Port": "Exp. Port", "Peer Rack": "Exp. Rack"}
    OPTIC_MAP = {"Peer Device": "Cut. Other End", "Peer Port": "Cut. Other End Port", "Peer Rack": "Cut. Other End Rack"}
    for sn in wb.sheetnames:
        if sn == "Summary" or "Mismatch" in sn:
            continue
        rename = OPTIC_MAP if "Optic" in sn else DOWNLINK_MAP
        for c in wb[sn][1]:
            if c.value in rename:
                c.value = rename[c.value]

def fill_empty_pps(wb):
    for sn in wb.sheetnames:
        if sn == "Summary":
            continue
        ws = wb[sn]
        pp_cols = [i + 1 for i, c in enumerate(ws[1]) if c.value and "PP" in str(c.value)]
        for r in range(2, ws.max_row + 1):
            for col in pp_cols:
                cell = ws.cell(row=r, column=col)
                if cell.value is None or cell.value == "":
                    cell.value = "<=>"

def add_note_column(wb):
    for sn in wb.sheetnames:
        if sn == "Summary":
            continue
        ws = wb[sn]
        ref = ws.cell(row=1, column=1)
        col = ws.max_column + 1
        c = ws.cell(row=1, column=col, value="Note")
        f = copy(ref.font); f.bold = True; c.font = f
        c.alignment = copy(ref.alignment)
        c.border = copy(ref.border)
        c.fill = YELLOW
        ws.column_dimensions[get_column_letter(col)].width = 30

def finalize_styling(wb):
    for sn in wb.sheetnames:
        ws = wb[sn]
        for c in ws[1]:
            f = copy(c.font); f.bold = True; f.size = 11; c.font = f
            c.fill = YELLOW
        if sn in ("T3-T2 Optic Errors", "T2-T1-T0 Optic Errors"):
            ws.freeze_panes = "B2"
        else:
            ws.freeze_panes = "A2"
        for r in range(1, ws.max_row + 1):
            for c in range(1, ws.max_column + 1):
                ws.cell(row=r, column=c).border = BORDER
        if sn != "Summary":
            last_col = get_column_letter(ws.max_column)
            ws.auto_filter.ref = f"A1:{last_col}{ws.max_row}"

def rebuild_summary(wb):
    if "Summary" not in wb.sheetnames:
        s = wb.create_sheet("Summary", 0)
        s.cell(row=1, column=1, value="Error Category")
        s.cell(row=1, column=2, value="Error Count")
    s = wb["Summary"]

    for r in range(s.max_row, 1, -1):
        s.delete_rows(r)

    rows = [
        ("T3-T2 Downlink",                 "=COUNTA('T3-T2 Downlink'!A:A)-1"),
        ("T2-T1-T0 Downlink",              "=COUNTA('T2-T1-T0 Downlink'!A:A)-1"),
        ("T3-T2 Mismatch",                 "=COUNTA('T3-T2 Mismatch'!A:A)-1"),
        ("T2-T1-T0 Mismatch",              "=COUNTA('T2-T1-T0 Mismatch'!A:A)-1"),
        ("T3-T2 Optic Errors",             "=COUNTA('T3-T2 Optic Errors'!A:A)-1"),
        ("T2-T1-T0 Optic Errors",          "=COUNTA('T2-T1-T0 Optic Errors'!A:A)-1"),
        ("T3-T2 Interface Down Errors",    "=COUNTA('T3-T2 Interface Down Errors'!A:A)-1"),
        ("T2-T1-T0 Interface Down Errors", "=COUNTA('T2-T1-T0 Interface Down Errors'!A:A)-1"),
    ]
    rows = [(label, formula) for (label, formula) in rows if label in wb.sheetnames]

    data_ref_cell = s.cell(row=1, column=1)
    no_fill = PatternFill(fill_type=None)

    for i, (label, formula) in enumerate(rows, start=2):
        c1 = s.cell(row=i, column=1, value=label)
        c2 = s.cell(row=i, column=2, value=formula)
        style_cell(c1, data_ref_cell)
        style_cell(c2, data_ref_cell)
        for c in (c1, c2):
            f = copy(c.font); f.bold = False; c.font = f
            c.fill = no_fill

    total_row = len(rows) + 2
    c1 = s.cell(row=total_row, column=1, value="Total")
    c2 = s.cell(row=total_row, column=2, value=f"=SUM(B2:B{len(rows) + 1})")
    style_cell(c1, data_ref_cell); style_cell(c2, data_ref_cell)
    for c in (c1, c2):
        f = copy(c.font); f.bold = True; c.font = f
        c.fill = no_fill

    s.column_dimensions["A"].width = 24
    s.column_dimensions["B"].width = 12

def reorder_tabs(wb):
    desired = [
        "Summary",
        "T3-T2 Downlink", "T2-T1-T0 Downlink",
        "T3-T2 Mismatch", "T2-T1-T0 Mismatch",
        "T3-T2 Optic Errors", "T2-T1-T0 Optic Errors",
        "T3-T2 Interface Down Errors", "T2-T1-T0 Interface Down Errors",
    ]
    wb._sheets = [wb[n] for n in desired if n in wb.sheetnames] + \
                 [wb[n] for n in wb.sheetnames if n not in desired]

# ---------------------------------------------------------------------------
# Main processing function (returns bytes)
# ---------------------------------------------------------------------------
def process_file_to_bytes(input_bytes, lookup, device_rack_lookup):
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp_in:
        tmp_in.write(input_bytes)
        tmp_in_path = tmp_in.name

    try:
     rack = "output"
if "LLDP Mismatch + Link Down" in wb.sheetnames:
    rack = rack_number_from(wb["LLDP Mismatch + Link Down"])

        split_lldp_sheet(wb)
        clean_columns(wb)
        enrich_all(wb, lookup)
        split_long_short(wb, lookup)
        trim_short_tabs(wb)
        enrich_mismatch_b_side(wb, lookup, device_rack_lookup)
        grey_optics_matching_downlink(wb)
        finalise_column_names(wb)
        fill_empty_pps(wb)
        add_note_column(wb)
        finalize_styling(wb)
        rebuild_summary(wb)
        reorder_tabs(wb)

        for sn in wb.sheetnames:
            if sn != "Summary":
                autofit_columns(wb[sn])

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue(), rack
    finally:
        os.unlink(tmp_in_path)

# ---------------------------------------------------------------------------
# Streamlit UI
# ---------------------------------------------------------------------------
st.markdown("### Upload Files")

cutsheet_file = st.file_uploader("CFAB Cutsheet (.xlsx)", type=["xlsx"], key="cutsheet")
rack_files = st.file_uploader("Rack Validation file(s) — you can select multiple", 
                              type=["xlsx"], accept_multiple_files=True, key="rack_files")

if st.button("🚀 Process Files", type="primary", disabled=not (cutsheet_file and rack_files)):
    with st.spinner("Processing... This can take some time for large files."):
        try:
            # Build cutsheet lookup
            with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
                tmp.write(cutsheet_file.getvalue())
                cutsheet_path = tmp.name
            lookup, device_rack_lookup = build_cutsheet_lookup(cutsheet_path)
            os.unlink(cutsheet_path)

            st.success(f"Cutsheet loaded — {len(lookup)} lookup entries")

            # Process each rack file
            results = []
            progress = st.progress(0)
            for idx, f in enumerate(rack_files):
                bytes_data, rack_name = process_file_to_bytes(f.getvalue(), lookup, device_rack_lookup)
                results.append((f"{rack_name}.xlsx", bytes_data))
                progress.progress((idx + 1) / len(rack_files))

            st.success(f"Processed {len(results)} file(s) successfully!")

            # Create ZIP for download
            zip_buffer = io.BytesIO()
            with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zipf:
                for filename, data in results:
                    zipf.writestr(filename, data)

            zip_buffer.seek(0)
            st.download_button(
                "📥 Download All Formatted Rack Files (ZIP)",
                data=zip_buffer,
                file_name="formatted_rack_validations.zip",
                mime="application/zip",
                use_container_width=True
            )

        except Exception as e:
            st.error(f"Error during processing: {str(e)}")
            st.exception(e)

st.caption("")
