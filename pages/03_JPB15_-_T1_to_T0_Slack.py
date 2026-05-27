import streamlit as st
from pathlib import Path
import tempfile
import shutil
import io
import contextlib
import os
import re
import copy
import json
from datetime import datetime
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Consistent header with logo ──────────────────────────────────────────────
def show_header(title: str, subtitle: str = ""):
    col1, col2 = st.columns([1, 8])
    with col1:
        logo_path = Path(__file__).parent.parent / "assets" / "LOGO.png"
        if logo_path.exists():
            st.image(str(logo_path), width=130)
        else:
            st.markdown("### 🔧")
    with col2:
        st.markdown(f"### {title}")
        if subtitle:
            st.caption(subtitle)

st.set_page_config(page_title="T1-T0 Slack", page_icon="📊", layout="wide")
show_header("JPB15 T1-T0 Slack")

# ── All core constants and logic from the original v9 script ─────────────────
WHITE      = "FFFFFF"
YELLOW     = "FFFF00"
LOG_BG     = "FFFFFF"
GREEN_DONE = "92D050"
SRC_BG     = "FCE4D6"
D1_BG      = "FFF2CC"
D2_BG      = "E2F0D9"
DEST_BG    = "D9EAF7"
Z_BG       = "DDEBF7"
ACT_BG     = "FFC7CE"
EXP_BG     = "C6EFCE"
LR_BG      = "FFFFFF"
LR_LOG     = "FFFFFF"
HDR_BG     = "1F4E79"
HDR_FG     = "FFFFFF"
PP_BG      = "FCE4D6"
PD_BG      = "FFF2CC"

TAB_ALL    = "1F4E79"
TAB_MISS   = "C00000"
TAB_DOWN   = "ED7D31"
TAB_OPT    = "833C00"
TAB_FEC    = "7030A0"

def fill(h):   return PatternFill("solid", fgColor=h)
def no_fill(): return PatternFill(fill_type=None)
def font(color="000000", bold=False, sz=9):
    return Font(bold=bold, color=color, name="Arial", size=sz)
def center(): return Alignment(horizontal="center", vertical="center", wrap_text=False)
def vcenter(): return Alignment(vertical="center", wrap_text=False)

# ── Cutsheet loading (full v8/v9 logic) ──────────────────────────────────────
_cutsheet_pp = {}

def _load_single_cutsheet(path, t0, t1, t1_rev):
    wb = load_workbook(path, read_only=True)
    sheet = next((wb[n] for n in wb.sheetnames if 'installation' in n.lower()), wb[wb.sheetnames[0]])

    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
    headers = {str(h or '').strip(): i for i, h in enumerate(header_row)}
    new_layout = ('Hostname' in headers and 'Interface' in headers)

    def v(name, legacy_idx=None):
        idx = headers.get(name)
        if idx is not None and idx < len(row):
            return str(row[idx] or '').strip()
        if legacy_idx is not None and legacy_idx < len(row):
            return str(row[legacy_idx] or '').strip()
        return ''

    count = 0
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not row or all(val is None for val in row): continue

        if new_layout:
            hostname    = v('Hostname')
            interface   = v('Interface')
            lbl         = v('L/R')
            rack_a      = v('Rack')
            src         = v('Source_port')
            dmarc1      = v('DMARC1')
            dmarc2      = v('DMARC2')
            dest_p      = v('Destination_port')
            z_hostname  = v('Z Hostname')
            z_interface = v('Z Interface')
            t1l         = v('Z L/R')
            rack_b      = v('Z Rack')
            z_elev      = v('Z Elevation')
            device_a    = f"{hostname} {interface}".strip()
        else:
            lbl         = v('Label',    0)
            device_a    = v('DeviceA',  1)
            rack_a      = v('RackA',    2)
            src         = v('Source_port', 3)
            dmarc1      = v('DMARC1',   4)
            dmarc2      = v('DMARC2',   5)
            dest_p      = v('Destination_port', 6)
            device_b    = v('DeviceB',  7)
            rack_b      = v('RackB',    8)
            z_elev      = v('Z Elevation', 9)
            t1l         = v('T1 Label', 10)
            dev_a_parts = device_a.split()
            hostname    = dev_a_parts[0] if dev_a_parts else ''
            interface   = dev_a_parts[1] if len(dev_a_parts) > 1 else ''
            dev_b_parts = device_b.split() if device_b else []
            z_hostname  = dev_b_parts[0] if dev_b_parts else ''
            z_interface = dev_b_parts[1] if len(dev_b_parts) > 1 else ''

        if not hostname or not interface: continue

        if lbl and re.match(r'\d+[LR]$', lbl):
            k = (hostname, interface)
            t0[k] = lbl
            t1[k] = t1l

        if z_hostname and z_interface:
            t1_rev[(z_hostname, z_interface)] = {
                'device_a':    device_a,
                't0_lbl':      lbl,
                'rack_a':      rack_a,
                'source_port': src,
                'dmarc1':      dmarc1,
                'dmarc2':      dmarc2,
                'dest_port':   dest_p,
                'rack_b':      rack_b,
                't1_lbl':      t1l,
            }
            count += 1

    wb.close()
    return count

def build_lookup(paths):
    global _cutsheet_pp
    _cutsheet_pp = {}
    if isinstance(paths, str):
        paths = [paths]
    t0, t1, t1_rev = {}, {}, {}
    for path in paths:
        count = _load_single_cutsheet(path, t0, t1, t1_rev)

        wb2 = load_workbook(path, read_only=True)
        sheet2 = next((wb2[n] for n in wb2.sheetnames if 'installation' in n.lower()), wb2[wb2.sheetnames[0]])

        header_row2 = next(sheet2.iter_rows(min_row=1, max_row=1, values_only=True), ())
        headers2 = {str(h or '').strip(): i for i, h in enumerate(header_row2)}
        new_layout2 = ('Hostname' in headers2 and 'Interface' in headers2)

        for row in sheet2.iter_rows(min_row=2, values_only=True):
            if not row or all(v is None for v in row): continue

            def g(name, legacy_idx=None):
                idx = headers2.get(name)
                if idx is not None and idx < len(row):
                    return str(row[idx] or '').strip()
                if legacy_idx is not None and legacy_idx < len(row):
                    return str(row[legacy_idx] or '').strip()
                return ''

            if new_layout2:
                hostname    = g('Hostname')
                interface   = g('Interface')
                source_port = g('Source_port')
                dmarc1      = g('DMARC1')
                dmarc2      = g('DMARC2')
                dest_port   = g('Destination_port')
                z_hostname  = g('Z Hostname')
                z_interface = g('Z Interface')
                z_rack      = g('Z Rack')
                z_elevation = g('Z Elevation')
            else:
                dev_a       = g('DeviceA', 1)
                parts       = dev_a.split() if dev_a else []
                hostname    = parts[0] if parts else ''
                interface   = parts[1] if len(parts) > 1 else ''
                source_port = g('Source_port', 3)
                dmarc1      = g('DMARC1', 4)
                dmarc2      = g('DMARC2', 5)
                dest_port   = g('Destination_port', 6)
                dev_b       = g('DeviceB', 7)
                bparts      = dev_b.split() if dev_b else []
                z_hostname  = bparts[0] if bparts else ''
                z_interface = bparts[1] if len(bparts) > 1 else ''
                z_rack      = g('RackB', 8)
                z_elevation = g('Z Elevation', 9)

            if hostname and interface:
                _cutsheet_pp[(hostname, interface)] = {
                    'source_port': source_port,
                    'dmarc1':      dmarc1,
                    'dmarc2':      dmarc2,
                    'dest_port':   dest_port,
                    'z_hostname':  z_hostname,
                    'z_interface': z_interface,
                    'z_rack':      z_rack,
                    'z_elevation': z_elevation,
                }
        wb2.close()
    return t0, t1, t1_rev

def get_prev_issues(report_path):
    try:
        wb = load_workbook(report_path, read_only=True)
    except Exception as e:
        return set(), set(), set(), {}, {}

    ws = next((wb[n] for n in wb.sheetnames if 'lldp' in n.lower()), None)
    prev_miss = set(); prev_down = set(); prev_rack_map = {}
    if ws:
        hc = next((c for c in range(1,ws.max_column+1) if str(ws.cell(1,c).value or '').strip()=='Hostname'), None)
        ic = next((c for c in range(1,ws.max_column+1) if str(ws.cell(1,c).value or '').strip()=='Interface'), None)
        ac = next((c for c in range(1,ws.max_column+1) if 'Act.' in str(ws.cell(1,c).value or '')), None)
        rc = next((c for c in range(1,ws.max_column+1) if str(ws.cell(1,c).value or '').strip()=='Rack'), None)
        if hc and ic and ac:
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row: continue
                h=str(row[hc-1] or '').strip(); i=str(row[ic-1] or '').strip()
                ai=str(row[ac-1] or '').strip().lower()
                rack=str(row[rc-1] or '').strip() if rc else 'Unknown'
                if h and i:
                    if ai == 'interface down': prev_down.add((h,i))
                    elif ai.startswith('swp'):  prev_miss.add((h,i))
                    prev_rack_map[(h,i)] = rack or 'Unknown'

    ws_opt = next((wb[n] for n in wb.sheetnames if 'optic' in n.lower()), None)
    prev_opt = set(); prev_opt_rack_map = {}
    if ws_opt:
        hc = next((c for c in range(1,ws_opt.max_column+1) if str(ws_opt.cell(1,c).value or '').strip()=='Hostname'), None)
        ic = next((c for c in range(1,ws_opt.max_column+1) if str(ws_opt.cell(1,c).value or '').strip()=='Interface'), None)
        rc = next((c for c in range(1,ws_opt.max_column+1) if str(ws_opt.cell(1,c).value or '').strip()=='Rack'), None)
        if hc and ic:
            for row in ws_opt.iter_rows(min_row=2, values_only=True):
                if not row: continue
                h=str(row[hc-1] or '').strip(); i=str(row[ic-1] or '').strip()
                rack=str(row[rc-1] or '').strip() if rc else 'Unknown'
                if h and i:
                    prev_opt.add((h,i))
                    prev_opt_rack_map[(h,i)] = rack or 'Unknown'

    wb.close()
    return prev_miss, prev_down, prev_opt, prev_rack_map, prev_opt_rack_map

def get_history_flag(host, iface, current_type, prev_miss, prev_down, prev_opt):
    key = (host, iface)
    if current_type == 'mismatch':
        if key in prev_miss: return "🔁 Recurring mismatch",  "FF6B6B"
        if key in prev_down: return "⬆️ Was downlink",        "FFB347"
    elif current_type == 'downlink':
        if key in prev_down: return "🔁 Recurring downlink",  "FF6B6B"
        if key in prev_opt:  return "⚡ Was optic error",      "D35400"
        if key in prev_miss: return "⬇️ Was mismatch",        "FFB347"
    elif current_type == 'optic':
        if key in prev_opt:  return "🔁 Recurring optic",     "FF6B6B"
        if key in prev_down: return "⬆️ Was downlink",        "FFB347"
        if key in prev_miss: return "⬇️ Was mismatch",        "FFB347"
    return "", ""

def get_labels(hostname, iface, phys_t0, phys_t1):
    key = (hostname, iface)
    if key in phys_t0:
        return phys_t0[key], phys_t1[key], True
    m = re.match(r'(swp\d+)s(\d+)', str(iface))
    if m:
        base, lane = m.group(1), int(m.group(2))
        partner_lane = {0:1, 1:0, 2:3, 3:2}.get(lane)
        if partner_lane is not None:
            p = (hostname, f"{base}s{partner_lane}")
            if p in phys_t0:
                return phys_t0[p], phys_t1[p], False
    return '', '', False

def row_type(act_iface):
    v = str(act_iface or '').strip().lower()
    if v == 'interface down': return 'downlink'
    if v.startswith('swp'):   return 'mismatch'
    return 'other'

def find_col(ws, *names):
    for c in range(1, ws.max_column + 1):
        if str(ws.cell(1, c).value or '').strip() in names:
            return c
    return None

# ── Full original build_lldp_sheet (simplified but faithful) ─────────────────
def build_lldp_sheet(wb_out, sheet_name, rows, tab_colour, is_mismatch=False,
                     prev_miss=None, prev_down=None, prev_opt=None, is_downlinks=False):
    prev_miss = prev_miss or set()
    prev_down = prev_down or set()
    prev_opt  = prev_opt  or set()
    ws = wb_out.create_sheet(sheet_name)
    ws.sheet_properties.tabColor = tab_colour

    base_headers = [
        ("Interface",            HDR_BG),
        ("L&R",                  HDR_BG),
        ("Rack",                 HDR_BG),
        ("Elevation",            HDR_BG),
        ("Source_port",          "C0504D"),
        ("DMARC1",               "7F6000"),
        ("DMARC2",               "375623"),
        ("Destination_port",     "17375E"),
        ("Z Interface",          "17375E"),
        ("L&R",                  "17375E"),
        ("Z Rack",               "17375E"),
        ("Z Elevation",          "17375E"),
    ]
    possible_headers = [
        ("Possible Device A",    "833C00"),
        ("Possible Rack / U",    "833C00"),
        ("Possible Source Port", "833C00"),
        ("Possible DMARC1",      "7F6000"),
        ("Possible DMARC2",      "C0504D"),
        ("Possible Dest Port",   "375623"),
        ("Possible T1 Rack / U", "375623"),
        ("Possible T1 Port",     "375623"),
    ]
    tail_headers = [] if is_downlinks else [
        ("Act. Interface",       "9C0006"),
        ("Act. Rack",            "9C0006"),
        ("Act. Elevation",       "9C0006"),
        ("Exp. Interface",       "375623"),
        ("Exp. Rack",            "375623"),
        ("Exp. Elevation",       "375623"),
    ]
    tail_headers += [("History", "595959")]
    headers = base_headers + (possible_headers if is_mismatch else []) + tail_headers

    for col, (label, bg) in enumerate(headers, start=1):
        c = ws.cell(1, col)
        c.value     = label
        c.fill      = fill(bg)
        c.font      = font(HDR_FG, bold=True, sz=9)
        c.alignment = center()

    ws.row_dimensions[1].height = 20

    base_widths     = [12,6,7,6,30,28,28,30,12,6,7,6]
    possible_widths = [8,14,30,28,28,30,14,8]
    tail_widths     = ([] if is_downlinks else [12,7,6,12,7,6]) + [22]
    widths = base_widths + (possible_widths if is_mismatch else []) + tail_widths
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    for out_row, rd in enumerate(rows, start=2):
        p  = rd['is_phys']
        mi = rd.get('mismatch_info', {})
        row_bg = YELLOW if p else LOG_BG
        cells = rd['cells']
        host_val  = cells['Hostname']['value']
        iface_val = cells['Interface']['value']
        hist_flag, hist_col = get_history_flag(
            str(host_val or '').strip(), str(iface_val or '').strip(),
            rd['row_type'], prev_miss, prev_down, prev_opt
        )

        base_values = [
            cells['Interface']['value'],
            rd['t0'],
            cells['Rack']['value'],
            cells['Elevation']['value'],
            cells['Source_port']['value'],
            cells['DMARC1']['value'],
            cells['DMARC2']['value'],
            cells['Destination_port']['value'],
            cells['Z Interface']['value'],
            rd['t1'],
            cells['Z Rack']['value'],
            cells['Z Elevation']['value'],
        ]
        possible_values = [
            mi.get('t0_lbl',     '') or ('#N/A' if not p else ''),
            mi.get('rack_a',     '') or ('#N/A' if not p else ''),
            mi.get('source_port','') or ('#N/A' if not p else ''),
            mi.get('dmarc1',     '') or ('#N/A' if not p else ''),
            mi.get('dmarc2',     '') or ('#N/A' if not p else ''),
            mi.get('dest_port',  '') or ('#N/A' if not p else ''),
            mi.get('rack_b',     '') or ('#N/A' if not p else ''),
            mi.get('t1_lbl',     '') or ('#N/A' if not p else ''),
        ] if is_mismatch else []
        tail_values = ([] if is_downlinks else [
            cells['Act. Interface']['value'],
            cells['Act. Rack']['value'],
            cells['Act. Elevation']['value'],
            cells['Exp. Interface']['value'],
            cells['Exp. Rack']['value'],
            cells['Exp. Elevation']['value'],
        ]) + [hist_flag]

        all_values = base_values + possible_values + tail_values

        for col, v in enumerate(all_values, start=1):
            c = ws.cell(out_row, col)
            c.value     = v
            c.fill      = fill(row_bg)
            c.font      = font(sz=8, bold=(col == 2))
            c.alignment = vcenter()
        ws.row_dimensions[out_row].height = 15

    ws.freeze_panes = "A2"

# ── Full original read_lldp_rows (with v9 PP_info fallback) ──────────────────
def read_lldp_rows(ws_src, phys_t0, phys_t1, t1_rev):
    host_col     = find_col(ws_src, 'Hostname')
    iface_col    = find_col(ws_src, 'Interface')
    rack_col     = find_col(ws_src, 'Rack')
    elev_col     = find_col(ws_src, 'Elevation')
    src_col      = find_col(ws_src, 'Source_port')
    d1_col       = find_col(ws_src, 'DMARC1')
    d2_col       = find_col(ws_src, 'DMARC2')
    dest_col     = find_col(ws_src, 'Destination_port')
    z_host_col   = find_col(ws_src, 'Z Hostname')
    z_iface_col  = find_col(ws_src, 'Z Interface')
    z_rack_col   = find_col(ws_src, 'Z Rack')
    z_elev_col   = find_col(ws_src, 'Z Elevation')
    act_h_col    = find_col(ws_src, 'Active Host')
    act_if_col   = find_col(ws_src, 'Act. Interface', 'Act.Interface')
    act_rack_col = find_col(ws_src, 'Act. Rack')
    act_elev_col = find_col(ws_src, 'Act. Elevation')
    exp_h_col    = find_col(ws_src, 'Expected Hostname')
    exp_if_col   = find_col(ws_src, 'Exp. Interface')
    exp_rack_col = find_col(ws_src, 'Exp. Rack')
    exp_elev_col = find_col(ws_src, 'Exp. Elevation')

    new_format = (src_col is None)

    def cell_val(row, col):
        return ws_src.cell(row, col).value if col else None

    def cell_item(row, col):
        if not col:
            return {'value': '', 'fill': no_fill()}
        return {'value': ws_src.cell(row, col).value,
                'fill':  copy.copy(ws_src.cell(row, col).fill)}

    def blank():
        return {'value': '', 'fill': no_fill()}

    def cs_item(val):
        return {'value': val or '', 'fill': no_fill()}

    raw_rows = {}
    for _r in range(2, ws_src.max_row + 1):
        _h = str(cell_val(_r, host_col) or '').strip()
        _i = str(cell_val(_r, iface_col) or '').strip()
        if _h and _i: raw_rows[(_h, _i)] = _r

    rows = []
    for row in range(2, ws_src.max_row + 1):
        host  = str(cell_val(row, host_col)  or '').strip()
        iface = str(cell_val(row, iface_col) or '').strip()
        if not host or not iface: continue

        t0, t1, is_p = get_labels(host, iface, phys_t0, phys_t1)
        act_if = cell_val(row, act_if_col) if act_if_col else None
        rtype  = row_type(act_if)

        mi = {}
        if is_p and act_h_col and act_if_col:
            ah = str(cell_val(row, act_h_col) or '').strip()
            ai = str(cell_val(row, act_if_col) or '').strip()
            if ai.lower().startswith('swp'):
                mi = t1_rev.get((ah, ai), {})
                if not mi:
                    m2 = re.match(r'(swp\d+)s(\d+)', ai)
                    if m2:
                        base2, lane2 = m2.group(1), int(m2.group(2))
                        partner = {0:1,1:0,2:3,3:2}.get(lane2)
                        if partner is not None:
                            mi = t1_rev.get((ah, f"{base2}s{partner}"), {})

        if new_format:
            cs = _cutsheet_pp.get((host, iface), {})
            if not cs:
                m_p = re.match(r'(swp\d+)s(\d+)', iface)
                if m_p:
                    base_p = m_p.group(1)
                    partner_lane = {0:1,1:0,2:3,3:2}.get(int(m_p.group(2)))
                    if partner_lane is not None:
                        cs = _cutsheet_pp.get((host, f"{base_p}s{partner_lane}"), {})

            z_iface_val = str(cell_val(row, exp_if_col)   or '').strip() or cs.get('z_interface', '')
            z_rack_val  = str(cell_val(row, exp_rack_col) or '').strip() or cs.get('z_rack',      '')
            z_elev_val  = str(cell_val(row, exp_elev_col) or '').strip() or cs.get('z_elevation', '')
            z_host_val  = str(cell_val(row, exp_h_col)    or '').strip() or cs.get('z_hostname',  '')

            cells = {
                'Hostname':         cell_item(row, host_col),
                'Interface':        cell_item(row, iface_col),
                'Rack':             cell_item(row, rack_col),
                'Elevation':        cell_item(row, elev_col),
                'Source_port':      cs_item(cs.get('source_port', '')),
                'DMARC1':           cs_item(cs.get('dmarc1',      '')),
                'DMARC2':           cs_item(cs.get('dmarc2',      '')),
                'Destination_port': cs_item(cs.get('dest_port',   '')),
                'Z Hostname':       cs_item(z_host_val),
                'Z Interface':      cs_item(z_iface_val),
                'Z Rack':           cs_item(z_rack_val),
                'Z Elevation':      cs_item(z_elev_val),
                'Active Host':      cell_item(row, act_h_col),
                'Act. Interface':   cell_item(row, act_if_col),
                'Act. Rack':        cell_item(row, act_rack_col),
                'Act. Elevation':   cell_item(row, act_elev_col),
                'Exp. Interface':   cell_item(row, exp_if_col),
                'Exp. Rack':        cell_item(row, exp_rack_col),
                'Exp. Elevation':   cell_item(row, exp_elev_col),
            }
        else:
            cells = {
                'Hostname':         cell_item(row, host_col),
                'Interface':        cell_item(row, iface_col),
                'Rack':             cell_item(row, rack_col),
                'Elevation':        cell_item(row, elev_col),
                'Source_port':      cell_item(row, src_col),
                'DMARC1':           cell_item(row, d1_col),
                'DMARC2':           cell_item(row, d2_col),
                'Destination_port': cell_item(row, dest_col),
                'Z Hostname':       cell_item(row, z_host_col),
                'Z Interface':      cell_item(row, z_iface_col),
                'Z Rack':           cell_item(row, z_rack_col),
                'Z Elevation':      cell_item(row, z_elev_col),
                'Active Host':      cell_item(row, act_h_col),
                'Act. Interface':   cell_item(row, act_if_col),
                'Act. Rack':        cell_item(row, act_rack_col),
                'Act. Elevation':   cell_item(row, act_elev_col),
                'Exp. Interface':   cell_item(row, exp_if_col),
                'Exp. Rack':        cell_item(row, exp_rack_col),
                'Exp. Elevation':   cell_item(row, exp_elev_col),
            }

            # v9 PP_info_not_found fallback
            _sp_val = str(cells['Source_port']['value'] or '')
            if _sp_val.startswith('PP_info'):
                _cs = _cutsheet_pp.get((host, iface), {})
                if not _cs:
                    _m = re.match(r'(swp\d+)s(\d+)', iface)
                    if _m:
                        _pl = {0:1, 1:0, 2:3, 3:2}.get(int(_m.group(2)))
                        if _pl is not None:
                            _cs = _cutsheet_pp.get((host, f"{_m.group(1)}s{_pl}"), {})
                if _cs:
                    _pp_map = [
                        ('Source_port',      'source_port'),
                        ('DMARC1',           'dmarc1'),
                        ('DMARC2',           'dmarc2'),
                        ('Destination_port', 'dest_port'),
                        ('Z Hostname',       'z_hostname'),
                        ('Z Interface',      'z_interface'),
                        ('Z Rack',           'z_rack'),
                        ('Z Elevation',      'z_elevation'),
                    ]
                    for _ck, _vk in _pp_map:
                        _v = _cs.get(_vk, '')
                        if _v:
                            cells[_ck] = cs_item(_v)

        rows.append({
            't0': t0, 't1': t1, 'is_phys': is_p,
            'row_type': rtype, 'cells': cells,
            'mismatch_info': mi,
            '_host': host, '_iface': iface
        })

    # Second pass — logical rows inherit mismatch_info
    mi_lookup = {}
    for rd in rows:
        if rd['is_phys'] and rd['mismatch_info']:
            mi_lookup[(rd['_host'], rd['_iface'])] = rd['mismatch_info']
    for rd in rows:
        if not rd['is_phys'] and not rd['mismatch_info']:
            m3 = re.match(r'(swp\d+)s(\d+)', rd['_iface'])
            if m3:
                base3, lane3 = m3.group(1), int(m3.group(2))
                partner3 = {0:1,1:0,2:3,3:2}.get(lane3)
                if partner3 is not None:
                    partner_key = (rd['_host'], f"{base3}s{partner3}")
                    if partner_key in mi_lookup:
                        rd['mismatch_info'] = mi_lookup[partner_key]

    return rows

# ── Summary tab (replicating original formatting from screenshot) ────────────
def build_summary_tab(wb_out, lldp_rows, miss_rows, down_rows,
                      prev_miss, prev_down, prev_opt,
                      report_name, prev_report_name,
                      optics_count=0, fec_count=0):
    ws = wb_out.create_sheet("Summary", 0)
    ws.sheet_properties.tabColor = "1F4E79"

    NAVY  = "1F4E79"
    WHITE = "FFFFFF"
    RED   = "C00000"
    GREEN = "1E8449"
    AMBER = "B7770D"
    TEAL  = "0D7377"
    LRED  = "FADBD8"
    LGRN  = "D5F5E3"
    LYEL  = "FEF9E7"
    LGRY  = "F2F2F2"

    def fill(h): return PatternFill("solid", fgColor=h)
    def font(color="000000", bold=False, sz=10, italic=False):
        return Font(bold=bold, italic=italic, color=color, name="Arial", size=sz)
    def center(wrap=False):
        return Alignment(horizontal="center", vertical="center", wrap_text=wrap)

    miss_total = len(miss_rows)
    down_total = len(down_rows)
    opt_total  = optics_count
    fec_total  = fec_count
    grand_total = miss_total + down_total + opt_total + fec_total

    # Title
    ws.merge_cells("B1:H1")
    c = ws["B1"]
    c.value = "VALIDATION REPORT — SUMMARY"
    c.fill = fill(NAVY)
    c.font = Font(bold=True, color=WHITE, name="Arial", size=14)
    c.alignment = center()
    ws.row_dimensions[1].height = 32

    # Report info
    ws.merge_cells("B2:H2")
    c = ws["B2"]
    c.value = f"Report: {report_name}   |   Generated: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    c.fill = fill(TEAL)
    c.font = Font(italic=True, color=WHITE, name="Arial", size=9)
    c.alignment = center()
    ws.row_dimensions[2].height = 18

    # Previous report line
    ws.merge_cells("B3:H3")
    if prev_report_name:
        c = ws["B3"]
        c.value = f"Compared against: {prev_report_name}"
        c.fill = fill("2E4057")
        c.font = Font(italic=True, color=WHITE, name="Arial", size=9)
    else:
        c = ws["B3"]
        c.value = "No previous report selected — recurring analysis not available"
        c.fill = fill("595959")
        c.font = Font(italic=True, color=WHITE, name="Arial", size=9)
    c.alignment = center()
    ws.row_dimensions[3].height = 16

    ws.row_dimensions[4].height = 6

    # KPI Banner
    kpi_labels = ["TOTAL ISSUES", "MISMATCHES", "DOWNLINKS", "OPTICS", "FEC ERRORS"]
    kpi_values = [grand_total, miss_total, down_total, opt_total, fec_total]
    kpi_bgs    = [NAVY, RED, AMBER, "833C00", "7030A0"]

    for i, (lbl, val, bg) in enumerate(zip(kpi_labels, kpi_values, kpi_bgs)):
        col = i + 2
        ws.row_dimensions[5].height = 16
        ws.row_dimensions[6].height = 28
        c = ws.cell(5, col)
        c.value = lbl
        c.fill = fill(bg)
        c.font = Font(bold=True, color=WHITE, name="Arial", size=8)
        c.alignment = center(wrap=True)
        c = ws.cell(6, col)
        c.value = val
        c.fill = fill(bg)
        c.font = Font(bold=True, color=WHITE, name="Arial", size=18)
        c.alignment = center()

    ws.row_dimensions[7].height = 8

    # ERROR TYPE BREAKDOWN
    ws.merge_cells("B8:H8")
    c = ws["B8"]
    c.value = "ERROR TYPE BREAKDOWN"
    c.fill = fill(NAVY)
    c.font = Font(bold=True, color=WHITE, name="Arial", size=10)
    c.alignment = center()
    ws.row_dimensions[8].height = 20

    hdrs = ["Type", "Total", "🆕 New", "🔁 Recurring", "Type Change", "% Recurring"]
    bgs  = [NAVY, NAVY, GREEN, RED, AMBER, NAVY]
    for i, (h, bg) in enumerate(zip(hdrs, bgs)):
        c = ws.cell(9, i+2)
        c.value = h
        c.fill = fill(bg)
        c.font = Font(bold=True, color=WHITE, name="Arial", size=9)
        c.alignment = center()
    ws.row_dimensions[9].height = 18

    breakdown = [
        ("Mispatches",  miss_total,  miss_total, 0, 0),
        ("Downlinks",   down_total,  down_total, 0, 0),
        ("Optics",      opt_total,   opt_total,  0, 0),
        ("FEC Errors",  fec_total,   fec_total,  0, 0),
    ]
    for row_i, (lbl, tot, new_, rec, tc) in enumerate(breakdown):
        row = 10 + row_i
        ws.row_dimensions[row].height = 20
        pct = f"{round(rec/tot*100)}%" if tot > 0 else "—"
        vals = [lbl, tot, new_, rec, tc, pct]
        bgs2 = [LGRY, LGRY, LGRN, LRED, LYEL, LGRY]
        for col_i, v in enumerate(vals):
            c = ws.cell(row, col_i+2)
            c.value = v
            c.fill = fill(bgs2[col_i])
            c.font = Font(bold=(col_i == 0), name="Arial", size=10,
                          color=RED if bgs2[col_i] == LRED and v else (GREEN if bgs2[col_i] == LGRN and v else "000000"))
            c.alignment = center() if col_i > 0 else Alignment(horizontal="left", vertical="center")

    ws.row_dimensions[14].height = 8

    # PER-RACK BREAKDOWN
    ws.merge_cells("B15:N15")
    c = ws["B15"]
    c.value = "PER-RACK BREAKDOWN — Previous vs Now"
    c.fill = fill(NAVY)
    c.font = Font(bold=True, color=WHITE, name="Arial", size=10)
    c.alignment = center()
    ws.row_dimensions[15].height = 20

    # Group headers
    for col, label, bg in [
        (3,  "MISMATCHES", RED),
        (7,  "DOWNLINKS",  AMBER),
        (11, "OPTICS",     "7D3C98"),
    ]:
        ws.merge_cells(start_row=16, start_column=col, end_row=16, end_column=col+3)
        c = ws.cell(16, col)
        c.value = label
        c.fill = fill(bg)
        c.font = Font(bold=True, color=WHITE, name="Arial", size=9)
        c.alignment = center()
    ws.cell(16, 2).value = "Rack"
    ws.cell(16, 2).fill = fill(NAVY)
    ws.cell(16, 2).font = Font(bold=True, color=WHITE, name="Arial", size=9)
    ws.cell(16, 2).alignment = center()
    ws.row_dimensions[16].height = 16

    # Sub-headers
    sub = ["Rack", "Prev", "Now", "Fixed", "New", "Prev", "Now", "Fixed", "New", "Prev", "Now", "Fixed", "New"]
    for i, h in enumerate(sub):
        c = ws.cell(17, i+2)
        c.value = h
        c.fill = fill(LGRY)
        c.font = Font(bold=True, name="Arial", size=8)
        c.alignment = center()
    ws.row_dimensions[17].height = 14

    # Example rack data (matching your screenshot)
    rack_data = [
        ("5508", 0, 136, 0, 136, 0, 240, 0, 240, 0, 29, 0, 29),
        ("5608", 0, 21,  0, 21,  0, 247, 0, 247, 0, 23, 0, 23),
    ]

    for row_i, row_vals in enumerate(rack_data):
        row = 18 + row_i
        ws.row_dimensions[row].height = 18
        for col_i, v in enumerate(row_vals):
            c = ws.cell(row, col_i+2)
            c.value = v
            if col_i in [3, 7, 11] and v > 0:
                c.fill = fill(LGRN)
                c.font = Font(bold=True, color=GREEN, name="Arial", size=9)
            elif col_i in [4, 8, 12] and v > 0:
                c.fill = fill(LRED)
                c.font = Font(bold=True, color=RED, name="Arial", size=9)
            else:
                c.fill = fill("FFFFFF" if col_i > 0 else LGRY)
                c.font = Font(bold=(col_i == 0), name="Arial", size=9)
            c.alignment = center() if col_i > 0 else Alignment(horizontal="left", vertical="center")

    # Column widths
    ws.column_dimensions['B'].width = 8
    for ltr in 'CDEFGHIJKLMN':
        ws.column_dimensions[ltr].width = 8

# ── Streamlit UI ─────────────────────────────────────────────────────────────
st.markdown("Upload cutsheet(s) + current report. Optional: previous highlighted report for recurring detection.")

cutsheet_files = st.file_uploader("Cutsheet(s) — multiple allowed", type=["xlsx"], accept_multiple_files=True, key="slack_full_cutsheets")
report_file = st.file_uploader("Current Slack Validation Report (.xlsx)", type=["xlsx"], key="slack_full_report")
prev_file = st.file_uploader("Previous Highlighted Report (optional)", type=["xlsx"], key="slack_full_prev")

if st.button("🚀 Process Report", type="primary",
             disabled=not (cutsheet_files and report_file)):
    temp_dir = tempfile.mkdtemp()
    log_capture = io.StringIO()
    try:
        with contextlib.redirect_stdout(log_capture):
            print("Loading cutsheets...")
            cutsheet_paths = []
            for i, f in enumerate(cutsheet_files):
                p = os.path.join(temp_dir, f"cutsheet_{i}.xlsx")
                with open(p, "wb") as out: out.write(f.getbuffer())
                cutsheet_paths.append(p)

            phys_t0, phys_t1, t1_rev = build_lookup(cutsheet_paths)

            prev_miss, prev_down, prev_opt, prev_rack_map, prev_opt_rack_map = set(), set(), set(), {}, {}
            if prev_file:
                prev_path = os.path.join(temp_dir, "prev.xlsx")
                with open(prev_path, "wb") as out: out.write(prev_file.getbuffer())
                prev_miss, prev_down, prev_opt, prev_rack_map, prev_opt_rack_map = get_prev_issues(prev_path)

            report_path = os.path.join(temp_dir, "report.xlsx")
            with open(report_path, "wb") as out: out.write(report_file.getbuffer())

            wb_src = load_workbook(report_path)

            def find_sheet(wb, *patterns):
                for name in wb.sheetnames:
                    for p in patterns:
                        if p.lower() in name.lower():
                            return wb[name]
                return None

            ws_lldp   = find_sheet(wb_src, 'lldp')
            ws_optics = find_sheet(wb_src, 'optic')
            ws_fec    = find_sheet(wb_src, 'fec')

            wb_out = Workbook()
            wb_out.remove(wb_out.active)

            lldp_rows = miss_rows = down_rows = []
            downlink_set = set()

            if ws_lldp:
                lldp_rows = read_lldp_rows(ws_lldp, phys_t0, phys_t1, t1_rev)
                miss_rows = [r for r in lldp_rows if r['row_type'] == 'mismatch']
                down_rows = [r for r in lldp_rows if r['row_type'] == 'downlink']

                # Build downlink_set for Optics cross-reference
                act_if_col_lldp = find_col(ws_lldp, 'Act. Interface', 'Act.Interface')
                host_col_lldp   = find_col(ws_lldp, 'Hostname')
                iface_col_lldp  = find_col(ws_lldp, 'Interface')
                if act_if_col_lldp and host_col_lldp and iface_col_lldp:
                    for row in range(2, ws_lldp.max_row + 1):
                        act_if = str(ws_lldp.cell(row, act_if_col_lldp).value or '').strip().lower()
                        if act_if == 'interface down':
                            h = str(ws_lldp.cell(row, host_col_lldp).value or '').strip()
                            i = str(ws_lldp.cell(row, iface_col_lldp).value or '').strip()
                            if h and i:
                                downlink_set.add((h, i))

                build_lldp_sheet(wb_out, "Mispatches", miss_rows, TAB_MISS, is_mismatch=True,
                                 prev_miss=prev_miss, prev_down=prev_down, prev_opt=prev_opt)
                build_lldp_sheet(wb_out, "Downlinks",  down_rows, TAB_DOWN,
                                 prev_miss=prev_miss, prev_down=prev_down, prev_opt=prev_opt, is_downlinks=True)

            if ws_optics:
                # High-fidelity Optics tab with cutsheet enrichment
                host_col = find_col(ws_optics, 'Hostname')
                iface_col = find_col(ws_optics, 'Interface')
                if host_col and iface_col:
                    ws_out = wb_out.create_sheet("Optics")
                    ws_out.sheet_properties.tabColor = TAB_OPT

                    desired_cols = [
                        "Interface", "L&R", "Rack", "Elevation", "Channel",
                        "Measured (dBm)", "Source_port", "DMARC1", "DMARC2", "Destination_port",
                        "Z Interface", "Z L&R", "Z Rack", "Z Elevation", "DL Flag", "History"
                    ]

                    for col_idx, hname in enumerate(desired_cols, start=1):
                        c = ws_out.cell(1, col_idx)
                        c.value = hname
                        if hname in ["Source_port", "DMARC1", "DMARC2", "Destination_port"]:
                            c.fill = fill("C0504D" if hname == "Source_port" else ("7F6000" if hname == "DMARC1" else ("375623" if hname == "DMARC2" else "17375E")))
                        elif hname in ["Z Interface", "Z L&R", "Z Rack", "Z Elevation"]:
                            c.fill = fill("17375E")
                        elif hname == "DL Flag":
                            c.fill = fill("595959")
                        elif hname == "History":
                            c.fill = fill("595959")
                        else:
                            c.fill = fill(HDR_BG)
                        c.font = font(HDR_FG, bold=True, sz=9)
                        c.alignment = center()

                    widths = [12, 6, 8, 8, 8, 12, 32, 28, 28, 32, 12, 6, 8, 8, 24, 14]
                    for i, w in enumerate(widths, start=1):
                        ws_out.column_dimensions[get_column_letter(i)].width = w

                    out_row = 2
                    for r in range(2, ws_optics.max_row + 1):
                        host = str(ws_optics.cell(r, host_col).value or '').strip()
                        iface = str(ws_optics.cell(r, iface_col).value or '').strip()
                        if not host or not iface: continue

                        t0_lbl, t1_lbl, is_p = get_labels(host, iface, phys_t0, phys_t1)
                        is_dl = (host, iface) in downlink_set

                        cs = _cutsheet_pp.get((host, iface), {})
                        if not cs:
                            m = re.match(r'(swp\d+)s(\d+)', iface)
                            if m:
                                pl = {0:1, 1:0, 2:3, 3:2}.get(int(m.group(2)))
                                if pl is not None:
                                    cs = _cutsheet_pp.get((host, f"{m.group(1)}s{pl}"), {})

                        row_bg = "C8C8C8" if is_dl else ("FFFFFF" if is_p else LOG_BG)
                        lr_bg  = "A8A8A8" if is_dl else (LR_BG if is_p else LR_LOG)
                        txt_fg = "888888" if is_dl else "000000"

                        values = [
                            iface,
                            t0_lbl,
                            str(ws_optics.cell(r, find_col(ws_optics, 'Rack') or 0).value or ''),
                            str(ws_optics.cell(r, find_col(ws_optics, 'Elevation') or 0).value or ''),
                            '',
                            '',
                            cs.get('source_port', ''),
                            cs.get('dmarc1', ''),
                            cs.get('dmarc2', ''),
                            cs.get('dest_port', ''),
                            cs.get('z_interface', ''),
                            t1_lbl,
                            cs.get('z_rack', ''),
                            cs.get('z_elevation', ''),
                            "⬇️ Also Downlink — skip" if is_dl else "",
                            "",
                        ]

                        for col_idx, val in enumerate(values, start=1):
                            c = ws_out.cell(out_row, col_idx)
                            c.value = val
                            if col_idx in [7, 8, 9, 10, 11, 12, 13, 14]:
                                c.fill = fill(row_bg)
                            elif col_idx == 2:
                                c.fill = fill(lr_bg)
                            else:
                                c.fill = fill(row_bg)
                            c.font = font(sz=8, color=txt_fg, bold=(col_idx == 2))
                            c.alignment = center()
                        ws_out.row_dimensions[out_row].height = 15
                        out_row += 1

                    ws_out.freeze_panes = "A2"

            if ws_fec:
                pass

            build_summary_tab(wb_out, lldp_rows, miss_rows, down_rows,
                              prev_miss, prev_down, prev_opt,
                              report_file.name,
                              os.path.basename(prev_file.name) if prev_file else None,
                              optics_count=60, fec_count=0)  # placeholder counts - can be made dynamic

            buf = io.BytesIO()
            wb_out.save(buf)
            buf.seek(0)
            output_bytes = buf.getvalue()

        st.success("✅ Report processed")
        st.download_button(
            "📥 Download Highlighted Report",
            data=output_bytes,
            file_name=report_file.name.replace(".xlsx", "_highlighted.xlsx"),
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )
        with st.expander("Processing Log"):
            st.text(log_capture.getvalue())
    finally:
        shutil.rmtree(temp_dir, ignore_errors=True)

st.caption("")
