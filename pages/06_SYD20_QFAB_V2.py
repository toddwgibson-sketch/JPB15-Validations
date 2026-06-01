#!/usr/bin/env python3
"""
SYD20 QFAB / Slack Report Formatter — Streamlit Version
Full conversion of the updated Tkinter script (SY20_QFAB_SLACK_No_PP.py)

This version includes all the new features from the updated script:
- Column merging (Hostname + Interface etc.)
- Advanced mismatch pair clustering with orange/yellow highlighting
- Grey-out logic for fully-down switches
- Separate T2-T1 / T1-T0 tabs for Optics and fec_ber
- Rich cutsheet enrichment + professional styling

Run:
    streamlit run SY20_QFAB_SLACK_streamlit.py
"""

import streamlit as st
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io
import tempfile
import os
import re
from collections import Counter, defaultdict
from pathlib import Path

# ================== Styles & Constants ==================
YELLOW_FILL = PatternFill("solid", start_color="FFFF00")
PINK_FILL = PatternFill("solid", start_color="FFC0CB")
ORANGE_FILL = PatternFill("solid", start_color="FFA500")
HEADER_FONT = Font(name="Arial", bold=True, color="000000")
BODY_FONT = Font(name="Arial")
GREY_FONT = Font(name="Arial", color="808080")

DOWN_PORT_THRESHOLD = 32
NOTE_COL_WIDTH = 25
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=False)
THIN = Side(style="thin", color="000000")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

PINK_COLS = [
    "Active Host Act. Interface", "Act. Rack", "Act. Elevation",
    "Cutsheet PP_A", "Cutsheet PP_B", "Cutsheet Other End", "Cutsheet Other End Rack",
]
CUT_COLS = ["Cutsheet PP_A", "Cutsheet PP_B", "Cutsheet Other End", "Cutsheet Other End Rack"]

DROP_SPECS = {
    "T2-T1 Downlink": ["index", "Building", "Exp. Building",
                       "Active Host", "Act. Interface", "Act. Building", "Act. Rack", "Act. Elevation"],
    "T1-T0 Downlink": ["index", "Building", "Exp. Building",
                       "PP_A", "PP_B",
                       "Active Host", "Act. Interface", "Act. Building", "Act. Rack", "Act. Elevation"],
    "T2-T1 Mismatch": ["index", "Building", "Act. Building", "Exp. Building"],
    "T1-T0 Mismatch": ["index", "Building", "PP_A", "PP_B", "Act. Building", "Exp. Building"],
    "optics":         ["index", "Building"],
    "fec_ber":        ["index", "BER", "Lock", "Remote Host", "Remote Interface", "Reason"],
}

MERGE_PAIRS = [
    ("Hostname", "Interface", "Hostname Interface"),
    ("Expected Hostname", "Exp. Interface", "Expected Hostname Exp. Interface"),
    ("Active Host", "Act. Interface", "Active Host Act. Interface"),
]

def extract_label(filename):
    """Pull a 'B<digits>' label out of a filename. Falls back to filestem."""
    m = re.search(r"b(\d+)", Path(filename).name, re.IGNORECASE)
    return f"B{m.group(1)}" if m else Path(filename).stem

# Constants from the original Tkinter version
PP_NOT_FOUND = "PP_info_not_found"
FULL_PATH_SHEET = "full_path_lldp_with_int_down"
RAW_LLDP_SHEET = "lldp_with_int_down"

# ================== Core Logic (ported from updated Tkinter version) ==================

# ---------- Loading & combining (full version) ----------
def load_combined(input_files, lookup=None):
    """Returns {sheet_name: {'headers': [...], 'rows': [...]}} stacked across files."""
    TARGETS = [FULL_PATH_SHEET, "optics", "fec_ber"]
    combined = {s: {"headers": None, "rows": []} for s in TARGETS}

    for uploaded_file in input_files:
        tag = extract_label(uploaded_file.name)

        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
            tmp.write(uploaded_file.getvalue())
            tmp_path = tmp.name

        try:
            wb = load_workbook(tmp_path, data_only=True)

            for s in TARGETS:
                if s in wb.sheetnames:
                    src_sheet, raw = s, False
                elif s == FULL_PATH_SHEET and RAW_LLDP_SHEET in wb.sheetnames:
                    src_sheet, raw = RAW_LLDP_SHEET, True
                else:
                    continue

                ws = wb[src_sheet]
                headers = [c.value for c in ws[1]]

                pp_idx = None
                if raw:
                    hi = headers.index("Hostname") if "Hostname" in headers else None
                    ii = headers.index("Interface") if "Interface" in headers else None
                    if "PP_A" not in headers and "PP_B" not in headers:
                        at = (headers.index("Elevation") + 1) if "Elevation" in headers else len(headers)
                        headers = headers[:at] + ["PP_A", "PP_B"] + headers[at:]
                    else:
                        at = headers.index("PP_A")
                    pp_idx = (hi, ii, at)

                if combined[s]["headers"] is None:
                    combined[s]["headers"] = ["Source"] + list(headers)

                for r in ws.iter_rows(min_row=2, values_only=True):
                    if all(v is None or v == "" for v in r):
                        continue
                    row = list(r)
                    if raw and pp_idx:
                        hi, ii, at = pp_idx
                        pp_a = pp_b = PP_NOT_FOUND
                        if lookup is not None and hi is not None and row[hi] is not None:
                            key = f"{row[hi]} {row[ii]}".strip()
                            hit = lookup.get(key)
                            if hit and hit[0] is not None:
                                pp_a, pp_b = hit[0], hit[1]
                        row = row[:at] + [pp_a, pp_b] + row[at:]

                    combined[s]["rows"].append([tag] + row)
        finally:
            os.unlink(tmp_path)

    return combined

def merge_columns(headers, rows):
    for col_a, col_b, new_hdr in MERGE_PAIRS:
        if col_a not in headers or col_b not in headers:
            continue
        ia = headers.index(col_a)
        ib = headers.index(col_b)
        new_headers = [new_hdr if i == ia else h for i, h in enumerate(headers) if i != ib]
        new_rows = []
        for r in rows:
            new_row = []
            for i, v in enumerate(r):
                if i == ib: continue
                if i == ia:
                    va = r[ia] or ''
                    vb = r[ib] or ''
                    new_row.append(f"{va} {vb}".strip())
                else:
                    new_row.append(v)
            new_rows.append(new_row)
        headers, rows = new_headers, new_rows
    return headers, rows


def sort_mismatch_pairs(headers, rows):
    """Group rows that are connected via Expected<->Active swaps into clusters.
    Cluster 1 → orange, Cluster 2 → yellow, Cluster 3 → orange, ... (alternating).
    Unpaired rows (no match) are moved to the end with no highlight.
    Returns (new_rows, row_colors) where row_colors is a dict {0-based index: fill}.
    """
    from collections import defaultdict

    exp_i = headers.index("Expected Hostname Exp. Interface")
    act_i = headers.index("Active Host Act. Interface")

    # Build adjacency: row i <-> row j if rows[i][exp] == rows[j][act] or vice versa
    act_to_idxs = defaultdict(list)
    for i, r in enumerate(rows):
        v = r[act_i]
        if v:
            act_to_idxs[v].append(i)

    adj = defaultdict(set)
    for i, r in enumerate(rows):
        exp_val = r[exp_i]
        if exp_val and exp_val in act_to_idxs:
            for j in act_to_idxs[exp_val]:
                if j != i:
                    adj[i].add(j)
                    adj[j].add(i)

    # Find connected components
    visited = set()
    groups = []
    for start in range(len(rows)):
        if start in visited or start not in adj:
            continue
        group = []
        stack = [start]
        while stack:
            n = stack.pop()
            if n in visited:
                continue
            visited.add(n)
            group.append(n)
            stack.extend(adj[n] - visited)
        groups.append(sorted(group))

    # Unpaired rows (not in any group)
    grouped_idxs = {i for g in groups for i in g}
    unpaired = [i for i in range(len(rows)) if i not in grouped_idxs]

    # Assign alternating colors: group 1=orange, group 2=yellow, group 3=orange ...
    group_color = []
    for gi, group in enumerate(groups):
        group_color.append(ORANGE_FILL if gi % 2 == 0 else YELLOW_FILL)

    # Build new row order: all grouped rows first (in group order), then unpaired
    new_rows = []
    row_colors = {}
    for gi, group in enumerate(groups):
        fill = group_color[gi]
        for orig_idx in group:
            row_colors[len(new_rows)] = fill
            new_rows.append(rows[orig_idx])

    for orig_idx in unpaired:
        new_rows.append(rows[orig_idx])

    return new_rows, row_colors


def drop_columns(headers, rows, drop_names):
    keep = [i for i, h in enumerate(headers) if h not in drop_names]
    return [headers[i] for i in keep], [[r[i] for i in keep] for r in rows]


def reorder_columns(headers, rows, new_order_names):
    idxs = [headers.index(n) for n in new_order_names]
    extras = [i for i in range(len(headers)) if i not in idxs]
    final = idxs + extras
    return [headers[i] for i in final], [[r[i] for i in final] for r in rows]


def swap_mismatch_groups(headers, rows):
    ACT = ["Active Host Act. Interface", "Act. Rack", "Act. Elevation"]
    EXP = ["Expected Hostname Exp. Interface", "Exp. Rack", "Exp. Elevation"]
    if not all(h in headers for h in ACT + EXP):
        return headers, rows
    act_idxs = [headers.index(h) for h in ACT]
    exp_idxs = [headers.index(h) for h in EXP]
    pre = [i for i in range(len(headers)) if i not in set(act_idxs + exp_idxs)]
    final = pre + exp_idxs + act_idxs
    return [headers[i] for i in final], [[r[i] for i in final] for r in rows]


def dedup_bidirectional(headers, rows):
    """Remove rows whose Hostname/Interface <-> Expected Hostname/Exp. Interface link
    already appeared in reverse direction. Merges Source values."""
    h_i = headers.index("Hostname")
    i_i = headers.index("Interface")
    eh_i = headers.index("Expected Hostname")
    ei_i = headers.index("Exp. Interface")
    src_i = headers.index("Source") if "Source" in headers else None
    seen = {}
    order = []
    for r in rows:
        key = frozenset([(r[h_i], r[i_i]), (r[eh_i], r[ei_i])])
        if key in seen:
            if src_i is not None:
                existing = str(seen[key][src_i])
                parts = [p.strip() for p in existing.split(",")]
                new_src = str(r[src_i])
                if new_src not in parts:
                    seen[key][src_i] = f"{existing},{new_src}"
        else:
            seen[key] = list(r)
            order.append(key)
    return [seen[k] for k in order]


def build_cutsheet_lookup(path_bytes):
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        tmp.write(path_bytes)
        tmp_path = tmp.name
    try:
        wb = load_workbook(tmp_path, data_only=True)
        ws = wb.active
        lookup = {}
        for r in ws.iter_rows(values_only=True):
            if not r or len(r) < 6: continue
            c1, c2, c3, c4, c5, c6 = r[:6]
            if not c5:
                if c1: lookup[str(c1).strip()] = (None, None, c3, c4)
                if c3: lookup[str(c3).strip()] = (None, None, c1, c2)
            else:
                if c1: lookup[str(c1).strip()] = (c3, c4, c5, c6)
                if c5: lookup[str(c5).strip()] = (c3, c4, c1, c2)
        return lookup
    finally:
        os.unlink(tmp_path)


def enrich(headers, rows, key_pairs, lookup):
    new_headers = list(headers) + CUT_COLS
    new_rows = []
    for r in rows:
        hit = None
        for host_col, intf_col in key_pairs:
            if host_col not in headers: continue
            h_idx = headers.index(host_col)
            i_idx = headers.index(intf_col)
            if r[h_idx] is None: continue
            key = f"{r[h_idx]} {r[i_idx]}".strip()
            h = lookup.get(key)
            if h:
                hit = h
                break
        new_rows.append(list(r) + list(hit) if hit else list(r) + [None]*len(CUT_COLS))
    return new_headers, new_rows


def fill_empty_pp(headers, rows):
    if "Cutsheet PP_A" not in headers or "Cutsheet PP_B" not in headers:
        return rows
    pa = headers.index("Cutsheet PP_A")
    pb = headers.index("Cutsheet PP_B")
    for r in rows:
        if r[pa] in (None, ""): r[pa] = "<==>"
        if r[pb] in (None, ""): r[pb] = "<==>"
    return rows


def split_full_path(headers, rows):
    ai = headers.index("Act. Interface")
    down, mis = [], []
    for r in rows:
        (down if r[ai] == "interface down" else mis).append(r)
    return down, mis


def split_by_pp(headers, rows):
    a = headers.index("PP_A")
    t2, t1 = [], []
    for r in rows:
        (t1 if r[a] == "PP_info_not_found" else t2).append(r)
    return t2, t1


def split_by_cutsheet_pp(headers, rows):
    if "Cutsheet PP_A" not in headers:
        return list(rows), []
    pa = headers.index("Cutsheet PP_A")
    t2, t1 = [], []
    for r in rows:
        v = r[pa] if pa < len(r) else None
        if v not in (None, "") and str(v).strip() != "<==>":
            t2.append(r)
        else:
            t1.append(r)
    return t2, t1


def write_sheet(wb, name, headers, rows):
    if name in wb.sheetnames:
        del wb[name]
    ws = wb.create_sheet(name)
    ws.append(headers)
    for r in rows:
        ws.append(r)
    return ws


def style_sheet(ws, pink_col_names=(), freeze_at=None, row_colors=None, col_widths=None):
    if ws.max_row == 0: return
    max_col = ws.max_column
    max_row = ws.max_row
    headers = [c.value for c in ws[1]]
    pink_idxs = {headers.index(n) + 1 for n in pink_col_names if n in headers}
    ws_row_colors = {i + 2: fill for i, fill in (row_colors or {}).items()}

    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = YELLOW_FILL
        cell.alignment = CENTER
        cell.border = BORDER

    act_col = headers.index("Active Host Act. Interface") + 1 if "Active Host Act. Interface" in headers else None

    for row in ws.iter_rows(min_row=2, max_row=max_row, max_col=max_col):
        row_fill = ws_row_colors.get(row[0].row)
        for cell in row:
            cell.font = BODY_FONT
            cell.alignment = CENTER
            cell.border = BORDER
            if row_fill and act_col and cell.column < act_col:
                cell.fill = row_fill
            elif cell.column in pink_idxs:
                cell.fill = PINK_FILL

    ws.auto_filter.ref = f"A1:{get_column_letter(max_col)}{max_row}"

    if col_widths:
        for ci, w in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(ci)].width = w
    else:
        for col_cells in ws.iter_cols(min_row=1, max_row=max_row, max_col=max_col):
            col_letter = get_column_letter(col_cells[0].column)
            max_len = max((len(str(c.value)) for c in col_cells if c.value is not None), default=10)
            ws.column_dimensions[col_letter].width = max(max_len + 2, 10)

    ws.freeze_panes = freeze_at or "A2"


def compute_col_widths(headers, rows):
    widths = [len(str(h)) if h is not None else 0 for h in headers]
    for row in rows:
        for i, v in enumerate(row):
            if v is not None:
                widths[i] = max(widths[i], len(str(v)))
    return [min(max(w + 2, 10), 60) for w in widths]


def annotate_downlink_switches(ws, threshold=DOWN_PORT_THRESHOLD):
    from collections import defaultdict
    headers = [c.value for c in ws[1]]
    if "Note" not in headers: return
    note_col = headers.index("Note") + 1
    target_cols = [h for h in ("Hostname Interface", "Expected Hostname Exp. Interface") if h in headers]
    if not target_cols: return
    col_idxs = [headers.index(h) for h in target_cols]

    def host_of(value):
        return str(value).split(" ", 1)[0] if value else None

    down_hosts = set()
    for ci in col_idxs:
        counts = defaultdict(int)
        for row in ws.iter_rows(min_row=2):
            h = host_of(row[ci].value)
            if h: counts[h] += 1
        down_hosts |= {h for h, n in counts.items() if n == threshold}

    if not down_hosts: return
    for row in ws.iter_rows(min_row=2):
        if any(host_of(row[ci].value) in down_hosts for ci in col_idxs):
            for cell in row:
                cell.font = GREY_FONT
            ws.cell(row=row[0].row, column=note_col).value = "maybe switch off"


def annotate_optics_in_downlinks(wb):
    def hi_values(name):
        if name not in wb.sheetnames: return set()
        ws = wb[name]
        headers = [c.value for c in ws[1]]
        if "Hostname Interface" not in headers: return set()
        i = headers.index("Hostname Interface")
        return {str(r[i]).strip() for r in ws.iter_rows(min_row=2, values_only=True) if r[i]}
    dl = hi_values("T2-T1 Downlink") | hi_values("T1-T0 Downlink")
    if not dl: return
    for opt_tab in ("T2-T1 Optics", "T1-T0 Optics"):
        if opt_tab not in wb.sheetnames: continue
        ws = wb[opt_tab]
        headers = [c.value for c in ws[1]]
        if "Hostname Interface" not in headers or "Note" not in headers: continue
        hi = headers.index("Hostname Interface")
        note_col = headers.index("Note") + 1
        for row in ws.iter_rows(min_row=2):
            v = row[hi].value
            if v and str(v).strip() in dl:
                for cell in row: cell.font = GREY_FONT
                ws.cell(row=row[0].row, column=note_col).value = "also in downlinks"


def widen_note_columns(wb, width=NOTE_COL_WIDTH):
    for name in wb.sheetnames:
        ws = wb[name]
        headers = [c.value for c in ws[1]]
        if "Note" in headers:
            col = get_column_letter(headers.index("Note") + 1)
            ws.column_dimensions[col].width = width


def capitalize_b_numbers(ws):
    B_NUM = re.compile(r"\bb(\d+)\b")
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            if isinstance(cell.value, str):
                new_val, n = B_NUM.subn(lambda m: f"B{m.group(1)}", cell.value)
                if n:
                    cell.value = new_val


def build_summary(wb, labels, cat_to_tabs):
    def count_sources(tabs):
        c = Counter()
        for tn in tabs:
            if tn not in wb.sheetnames: continue
            ws = wb[tn]
            headers = [cell.value for cell in ws[1]]
            if "Source" not in headers: continue
            s_idx = headers.index("Source")
            for r in ws.iter_rows(min_row=2, values_only=True):
                src = r[s_idx]
                if src is None: continue
                for part in str(src).split(","):
                    p = part.strip()
                    if p: c[p] += 1
        return c

    if "summary" in wb.sheetnames:
        del wb["summary"]
    ws = wb.create_sheet("summary", 0)
    ws.append(["Category"] + list(labels) + ["Total"])
    for cat, tabs in cat_to_tabs:
        c = count_sources(tabs)
        row_idx = ws.max_row + 1
        last_col = get_column_letter(1 + len(labels))
        ws.append([cat] + [c.get(l, 0) for l in labels] + [f"=SUM(B{row_idx}:{last_col}{row_idx})"])
    return ws


# ================== Main Processing Pipeline ==================
def process_files(cutsheet_bytes, input_files):
    labels = []
    for f in input_files:
        m = re.search(r"b(\d+)", f.name, re.IGNORECASE)
        labels.append(f"B{m.group(1)}" if m else Path(f.name).stem)

    lookup = build_cutsheet_lookup(cutsheet_bytes)

    # Use the proper load_combined from the updated logic (ported)
    combined = load_combined(input_files, lookup)

    fp = combined["full_path_lldp_with_int_down"]
    if fp["headers"] is None:
        raise ValueError("No full_path_lldp_with_int_down sheet found in any input file.")

    fp_headers = fp["headers"]
    fp_rows = fp["rows"]

    dl_rows, mis_rows = split_full_path(fp_headers, fp_rows)
    t2_dl_rows, t1_dl_rows = split_by_pp(fp_headers, dl_rows)
    t2_mis_rows, t1_mis_rows = split_by_pp(fp_headers, mis_rows)

    # Drop columns per spec
    t2_dl_hdr, t2_dl_rows = drop_columns(fp_headers, t2_dl_rows, DROP_SPECS["T2-T1 Downlink"])
    t1_dl_hdr, t1_dl_rows = drop_columns(fp_headers, t1_dl_rows, DROP_SPECS["T1-T0 Downlink"])
    t2_mis_hdr, t2_mis_rows = drop_columns(fp_headers, t2_mis_rows, DROP_SPECS["T2-T1 Mismatch"])
    t1_mis_hdr, t1_mis_rows = drop_columns(fp_headers, t1_mis_rows, DROP_SPECS["T1-T0 Mismatch"])

    # Swap Active/Expected in Mismatch tabs
    t2_mis_hdr, t2_mis_rows = swap_mismatch_groups(t2_mis_hdr, t2_mis_rows)
    t1_mis_hdr, t1_mis_rows = swap_mismatch_groups(t1_mis_hdr, t1_mis_rows)

    # Dedup T1-T0 tabs
    t1_dl_rows = dedup_bidirectional(t1_dl_hdr, t1_dl_rows)
    t1_mis_rows = dedup_bidirectional(t1_mis_hdr, t1_mis_rows)

    # === optics / fec_ber ===
    op = combined["optics"]
    fb = combined["fec_ber"]

    if op["headers"] is None:
        op_hdr, op_rows = [], []
    else:
        op_hdr, op_rows = drop_columns(op["headers"], op["rows"], DROP_SPECS["optics"])
        OPTICS_ORDER = ["Source", "Input Power", "Output Power", "Hostname", "Interface", "Rack", "Elevation"]
        op_hdr, op_rows = reorder_columns(op_hdr, op_rows, [c for c in OPTICS_ORDER if c in op_hdr])

    if fb["headers"] is None:
        fb_hdr, fb_rows = [], []
    else:
        fb_hdr, fb_rows = drop_columns(fb["headers"], fb["rows"], DROP_SPECS["fec_ber"])

    # === Enrich with cutsheet ===
    t2_mis_hdr, t2_mis_rows = enrich(t2_mis_hdr, t2_mis_rows,
                                     [("Active Host", "Act. Interface"), ("Hostname", "Interface")], lookup)
    t1_mis_hdr, t1_mis_rows = enrich(t1_mis_hdr, t1_mis_rows,
                                     [("Active Host", "Act. Interface"), ("Hostname", "Interface")], lookup)
    if op_hdr:
        op_hdr, op_rows = enrich(op_hdr, op_rows, [("Hostname", "Interface")], lookup)
    if fb_hdr:
        fb_hdr, fb_rows = enrich(fb_hdr, fb_rows, [("Hostname", "Interface")], lookup)

    # Replace empty Cutsheet PPs with <==>
    for hdr, rows in [(t2_mis_hdr, t2_mis_rows), (t1_mis_hdr, t1_mis_rows),
                      (op_hdr, op_rows), (fb_hdr, fb_rows)]:
        fill_empty_pp(hdr, rows)

    # === Merge Hostname+Interface column pairs ===
    t2_dl_hdr, t2_dl_rows = merge_columns(t2_dl_hdr, t2_dl_rows)
    t1_dl_hdr, t1_dl_rows = merge_columns(t1_dl_hdr, t1_dl_rows)
    t2_mis_hdr, t2_mis_rows = merge_columns(t2_mis_hdr, t2_mis_rows)
    t1_mis_hdr, t1_mis_rows = merge_columns(t1_mis_hdr, t1_mis_rows)
    op_hdr, op_rows = merge_columns(op_hdr, op_rows)
    fb_hdr, fb_rows = merge_columns(fb_hdr, fb_rows)

    # === Split optics and fec_ber into T2-T1 (has Cutsheet PP) and T1-T0 ===
    op_t2_rows, op_t1_rows = split_by_cutsheet_pp(op_hdr, op_rows)
    fb_t2_rows, fb_t1_rows = split_by_cutsheet_pp(fb_hdr, fb_rows)

    # === Group mismatch pairs ===
    t2_mis_rows, t2_mis_colors = sort_mismatch_pairs(t2_mis_hdr, t2_mis_rows)
    t1_mis_rows, t1_mis_colors = sort_mismatch_pairs(t1_mis_hdr, t1_mis_rows)

    # === Add Note column ===
    sheets_to_write = [
        ("T2-T1 Downlink", t2_dl_hdr + ["Note"], [r + [None] for r in t2_dl_rows]),
        ("T1-T0 Downlink", t1_dl_hdr + ["Note"], [r + [None] for r in t1_dl_rows]),
        ("T2-T1 Mismatch", t2_mis_hdr + ["Note"], [r + [None] for r in t2_mis_rows]),
        ("T1-T0 Mismatch", t1_mis_hdr + ["Note"], [r + [None] for r in t1_mis_rows]),
        ("T2-T1 Optics",   op_hdr + ["Note"],     [r + [None] for r in op_t2_rows]),
        ("T1-T0 Optics",   op_hdr + ["Note"],     [r + [None] for r in op_t1_rows]),
        ("T2-T1 fec_ber",  fb_hdr + ["Note"],     [r + [None] for r in fb_t2_rows]),
        ("T1-T0 fec_ber",  fb_hdr + ["Note"],     [r + [None] for r in fb_t1_rows]),
    ]

    sheet_col_widths = {
        name: compute_col_widths(hdr, rows)
        for name, hdr, rows in sheets_to_write
    }

    wb = Workbook()
    wb.remove(wb.active)
    for name, hdr, rows in sheets_to_write:
        write_sheet(wb, name, hdr, rows)

    # === Summary tab ===
    build_summary(wb, labels, [
        ("Downlink", ["T2-T1 Downlink", "T1-T0 Downlink"]),
        ("Mismatch", ["T2-T1 Mismatch", "T1-T0 Mismatch"]),
        ("optics",   ["T2-T1 Optics", "T1-T0 Optics"]),
        ("fec_ber",  ["T2-T1 fec_ber", "T1-T0 fec_ber"]),
    ])

    # === Log errors locally for testing (simple & reliable) ===
    print("\n--- Writing error summary to local log file ---")
    try:
        import pandas as pd
        from datetime import datetime
        from pathlib import Path

        # Write next to this script for easy finding during testing
        log_path = Path(__file__).parent / "QFAB_Error_Log.xlsx"

        new_rows = []
        if "summary" in wb.sheetnames:
            ws_sum = wb["summary"]
            for row in ws_sum.iter_rows(min_row=2, values_only=True):
                if row and row[0]:
                    category = str(row[0])
                    for idx, bldg in enumerate(labels):
                        if idx + 1 < len(row) and row[idx + 1]:
                            count = int(row[idx + 1])
                            if count > 0:
                                new_rows.append({
                                    "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                                    "hall": "SYD20",
                                    "rack_type": category,
                                    "building": bldg,
                                    "error_category": category,
                                    "count": count,
                                    "source_file": "-".join(labels) + "_QFAB_Updated.xlsx"
                                })

        if new_rows:
            new_df = pd.DataFrame(new_rows)
            if log_path.exists():
                existing = pd.read_excel(log_path)
                combined = pd.concat([existing, new_df], ignore_index=True)
            else:
                combined = new_df

            combined.to_excel(log_path, index=False)
            print(f"✅ Error log written to: {log_path}")
        else:
            print("No error rows to log.")

    except Exception as log_err:
        print(f"ERROR writing local error log: {log_err}")
        import traceback
        traceback.print_exc()

    # Reorder sheets
    canonical = ["summary", "T2-T1 Downlink", "T1-T0 Downlink",
                 "T2-T1 Mismatch", "T1-T0 Mismatch",
                 "T2-T1 Optics", "T1-T0 Optics",
                 "T2-T1 fec_ber", "T1-T0 fec_ber"]
    wb._sheets = [wb[n] for n in canonical if n in wb.sheetnames]

    # Capitalize b<digits>
    for name in wb.sheetnames:
        capitalize_b_numbers(wb[name])

    # === Style every sheet ===
    mismatch_colors = {"T2-T1 Mismatch": t2_mis_colors, "T1-T0 Mismatch": t1_mis_colors}
    for name in wb.sheetnames:
        pink   = PINK_COLS if name.endswith("Mismatch") else ()
        freeze = "D2" if name.endswith("Optics") else "A2"
        rc     = mismatch_colors.get(name)
        cw     = sheet_col_widths.get(name)
        style_sheet(wb[name], pink_col_names=pink, freeze_at=freeze, row_colors=rc, col_widths=cw)

    # Grey-out annotations
    for dl_tab in ("T2-T1 Downlink", "T1-T0 Downlink"):
        if dl_tab in wb.sheetnames:
            annotate_downlink_switches(wb[dl_tab])
    annotate_optics_in_downlinks(wb)
    widen_note_columns(wb)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return output.getvalue(), "-".join(labels) + "_QFAB_Updated.xlsx"


# ================== Streamlit UI ==================
st.set_page_config(page_title="SYD20 QFAB (Updated)", page_icon="📊", layout="wide")
st.title("SYD20 QFAB / Slack Formatter — Updated Version")
st.caption("Converted from the latest Claude Tkinter script (with all new features: column merging, mismatch clustering, grey-out logic, etc.)")

st.markdown("### 1. Cutsheet")
cutsheet_file = st.file_uploader("Select the CUTSHEET xlsx", type=["xlsx", "xlsm"], key="cutsheet_v2")

st.markdown("### 2. Input Files (one or more per-building files)")
input_files = st.file_uploader(
    "Select input audit files",
    type=["xlsx", "xlsm"],
    accept_multiple_files=True,
    key="inputs_v2"
)

### 🧪 Testing Tools (Temporary - for debugging the log)
st.markdown("**Use this button to test if error logging is working.** It will try to create a file called `QFAB_Error_Log_TEST.xlsx` on your Desktop and show you the exact path.")

if st.button("🧪 Test: Force write one error row to local log"):
    try:
        import pandas as pd
        from datetime import datetime
        from pathlib import Path
        import os

        home = Path.home()
        desktop = home / "Desktop"
        script_dir = Path(__file__).parent.resolve()

        print("\n========== LOGGING DIAGNOSTICS ==========")
        print(f"Python __file__     : {__file__}")
        print(f"Resolved script dir : {script_dir}")
        print(f"Path.home()         : {home}")
        print(f"Desktop path        : {desktop}")
        print(f"Desktop exists?     : {desktop.exists()}")
        print(f"Current working dir : {os.getcwd()}")
        print("==========================================\n")

        log_path = desktop / "QFAB_Error_Log_TEST.xlsx"

        st.write("**Attempting to write to:**")
        st.code(str(log_path))

        test_row = pd.DataFrame([{
            "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "hall": "SYD20",
            "rack_type": "TEST",
            "building": "TEST-BLDG",
            "error_category": "TEST",
            "count": 99,
            "source_file": "manual_test"
        }])

        if log_path.exists():
            existing = pd.read_excel(log_path)
            combined = pd.concat([existing, test_row], ignore_index=True)
        else:
            combined = test_row

        combined.to_excel(log_path, index=False)

        if log_path.exists():
            st.success(f"✅ SUCCESS! File exists at: {log_path}")
            st.code(str(log_path), language="text")
            print(f"[TEST] SUCCESS - File exists at: {log_path}")
        else:
            st.error("File was supposedly written but does not exist on disk.")
            print("[TEST] ERROR - File does not exist after write attempt.")

    except Exception as e:
        st.error(f"Test log failed with exception: {e}")
        print(f"[TEST] EXCEPTION: {e}")
        import traceback
        traceback.print_exc()

# --- Main Process Button ---
if st.button("🚀 Process Files", type="primary", disabled=not (cutsheet_file and input_files)):
    with st.spinner("Processing with the full updated logic..."):
        try:
            result_bytes, filename = process_files(cutsheet_file.getvalue(), input_files)
            if result_bytes:
                st.success("Processing complete!")

                # ====================== PRE-DOWNLOAD ANALYSIS ======================
                st.subheader("📊 Error Summary Snapshot")

                from io import BytesIO
                import pandas as pd

                wb_preview = load_workbook(BytesIO(result_bytes))

                if "summary" in wb_preview.sheetnames:
                    ws_sum = wb_preview["summary"]
                    summary_rows = list(ws_sum.iter_rows(min_row=1, values_only=True))

                    if summary_rows and len(summary_rows) > 1:
                        df_summary = pd.DataFrame(summary_rows[1:], columns=summary_rows[0])
                        df_summary = df_summary.fillna(0)

                        # Identify building columns
                        non_building = ["Category", "Total"]
                        building_cols = [col for col in df_summary.columns if col not in non_building]

                        # === Key Stats ===
                        total_issues = 0
                        if "Total" in df_summary.columns:
                            total_issues = int(df_summary["Total"].sum())

                        st.markdown("### Key Error Metrics")

                        col1, col2, col3 = st.columns(3)
                        with col1:
                            st.metric("Total Issues", total_issues)
                        with col2:
                            num_buildings = len(building_cols)
                            st.metric("Buildings Affected", num_buildings)

                        # Breakdown by error type
                        st.markdown("**Issues by Category**")
                        category_totals = df_summary.set_index("Category")[building_cols].sum(axis=1)
                        st.bar_chart(category_totals)

                        # Per-building breakdown
                        st.markdown("**Issues per Building**")
                        building_totals = df_summary[building_cols].sum().sort_values(ascending=False)
                        st.dataframe(building_totals.reset_index().rename(columns={"index": "Building", 0: "Issues"}), use_container_width=True, hide_index=True)

                        # Top 5 worst buildings
                        if len(building_totals) > 0:
                            st.markdown("**Top Problematic Buildings**")
                            for bldg, count in building_totals.head(5).items():
                                if count > 0:
                                    st.write(f"• **{bldg}**: {int(count)} issues")

                        # Raw summary for reference (collapsible)
                        with st.expander("View full Summary table"):
                            st.dataframe(df_summary, use_container_width=True, hide_index=True)
                    else:
                        st.info("Summary sheet is empty.")
                else:
                    st.warning("No summary sheet was generated.")

                # ====================== DOWNLOAD ======================
                st.download_button(
                    "📥 Download Formatted Report",
                    data=result_bytes,
                    file_name=filename,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )

                # === Extra visible test button for logging (with heavy diagnostics) ===
                if st.button("🧪 Test: Force write one error row to local log"):
                    try:
                        import pandas as pd
                        from datetime import datetime
                        from pathlib import Path
                        import os

                        home = Path.home()
                        desktop = home / "Desktop"
                        script_dir = Path(__file__).parent.resolve()

                        print("\n========== LOGGING DIAGNOSTICS ==========")
                        print(f"Python __file__     : {__file__}")
                        print(f"Resolved script dir : {script_dir}")
                        print(f"Path.home()         : {home}")
                        print(f"Desktop path        : {desktop}")
                        print(f"Desktop exists?     : {desktop.exists()}")
                        print(f"Current working dir : {os.getcwd()}")
                        print("==========================================\n")

                        # Try Desktop first
                        log_path = desktop / "QFAB_Error_Log_TEST.xlsx"

                        st.write("**Attempting to write to:**")
                        st.code(str(log_path))

                        test_row = pd.DataFrame([{
                            "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                            "hall": "SYD20",
                            "rack_type": "TEST",
                            "building": "TEST-BLDG",
                            "error_category": "TEST",
                            "count": 99,
                            "source_file": "manual_test"
                        }])

                        if log_path.exists():
                            existing = pd.read_excel(log_path)
                            combined = pd.concat([existing, test_row], ignore_index=True)
                        else:
                            combined = test_row

                        combined.to_excel(log_path, index=False)

                        if log_path.exists():
                            st.success(f"✅ SUCCESS! File exists at: {log_path}")
                            st.code(str(log_path), language="text")
                            print(f"[TEST] SUCCESS - File exists at: {log_path}")
                        else:
                            st.error("File was supposedly written but does not exist on disk.")
                            print("[TEST] ERROR - File does not exist after write attempt.")

                    except Exception as e:
                        st.error(f"Test log failed with exception: {e}")
                        print(f"[TEST] EXCEPTION: {e}")
                        import traceback
                        traceback.print_exc()
        except Exception as e:
            st.error(f"Error during processing: {e}")
            st.exception(e)

st.caption("All processing happens locally in your browser. Nothing is uploaded to any server.")
